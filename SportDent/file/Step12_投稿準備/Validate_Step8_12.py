"""Step8〜12とStudyAIの最終整合性・情報管理を読み取り専用で検証する。"""
from __future__ import annotations

import hashlib
import math
import re
from pathlib import Path

import openpyxl
import pandas as pd


HERE = Path(__file__).resolve().parent
PROJECT = HERE.parent.parent
CREATE = PROJECT / "CreateData"
STEP8 = CREATE / "Step8_多変量解析"
STEP9 = CREATE / "Step9_感度分析"
STEP10 = CREATE / "Step10_統合論文原稿"
STEP11 = CREATE / "Step11_通学中臨床レビュー"
STEP12 = CREATE / "Step12_投稿準備"
STUDY_AI = PROJECT / "StudyAI"


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def csv(path: Path) -> pd.DataFrame:
    require(path.exists(), f"必須CSVがありません: {path}")
    return pd.read_csv(path, encoding="utf-8-sig", dtype=str, keep_default_na=False)


def number(value: object) -> float:
    return float(str(value).replace(",", ""))


def close(actual: object, expected: float, label: str, tolerance: float = 1e-6) -> None:
    require(
        math.isclose(number(actual), expected, rel_tol=tolerance, abs_tol=tolerance),
        f"{label}が不一致です: actual={actual}, expected={expected}",
    )


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def one(frame: pd.DataFrame, mask: pd.Series, label: str) -> pd.Series:
    selected = frame.loc[mask]
    require(len(selected) == 1, f"{label}の該当行が1行ではありません: {len(selected)}")
    return selected.iloc[0]


def validate_statistics() -> None:
    odds = csv(STEP8 / "Step8-03_調整オッズ比.csv")
    commute = one(
        odds,
        odds["モデル"].eq("主モデル（大分類）")
        & odds["項目（比較対象／基準）"].str.startswith("場合別1：通学中（"),
        "通学中の主モデル",
    )
    close(commute["調整オッズ比"], 3.052161028, "通学中aOR")
    close(commute["95%CI下限"], 2.512837076, "通学中95%CI下限")
    close(commute["95%CI上限"], 3.707238734, "通学中95%CI上限")

    walking = one(
        odds,
        odds["モデル"].eq("通学中サブグループ")
        & odds["項目（比較対象／基準）"].str.startswith("通学方法：徒歩（"),
        "通学中サブグループの徒歩",
    )
    close(walking["調整オッズ比"], 0.348486991, "徒歩aOR")
    close(walking["95%CI下限"], 0.175364463, "徒歩95%CI下限")
    close(walking["95%CI上限"], 0.692518774, "徒歩95%CI上限")

    design = csv(STEP8 / "Step8-06_モデル設計と基準カテゴリ.csv")
    require(len(design) == 3, "Step8のモデル設計表は3モデルである必要があります")
    require(design["給付年度の1標準偏差（年）"].nunique() == 1, "給付年度SDがモデル間で不一致です")
    close(design.iloc[0]["給付年度の1標準偏差（年）"], 5.376479410, "給付年度1SD")

    interaction = csv(STEP9 / "Step9-03_通学中と学校種_交互作用.csv")
    require(set(interaction["解析対象件数"]) == {"7104"}, "交互作用解析が7,104件に限定されていません")
    require(set(interaction["解析対象学校種"]) == {"高（基準）・中・小のみ"}, "交互作用の学校種表記が不一致です")
    require(set(interaction["LR検定自由度"]) == {"2"}, "交互作用LR検定の自由度が2ではありません")
    close(interaction.iloc[0]["交互作用全体のLR検定p値"], 0.000370807246, "交互作用LR検定p値")
    middle = one(interaction, interaction["項目"].str.startswith("交互作用：通学中×中（"), "通学中×中")
    elementary = one(interaction, interaction["項目"].str.startswith("交互作用：通学中×小（"), "通学中×小")
    close(middle["調整オッズ比"], 1.398066629, "通学中×中aOR")
    close(elementary["調整オッズ比"], 0.456993138, "通学中×小aOR")

    cross_validation = csv(STEP9 / "Step9-05_5分割交差検証.csv")
    average = one(cross_validation, cross_validation["分割"].eq("5分割平均"), "5分割平均")
    close(average["AUC"], 0.682438922, "5分割平均AUC")
    close(average["Brierスコア"], 0.150949255, "5分割平均Brier")
    close(average["LogLoss"], 0.470905975, "5分割平均LogLoss")
    require(
        cross_validation["前処理"].str.contains("学習fold内").all(),
        "交差検証の前処理が学習fold内と明記されていません",
    )

    two_group = csv(STEP9 / "Step9-10_通学方法二群_感度分析.csv")
    require(set(two_group["解析対象件数"]) == {"558"}, "通学方法二群感度分析が558件ではありません")
    walking_sensitivity = one(
        two_group,
        two_group["項目（比較対象／基準）"].str.startswith("通学方法：徒歩（"),
        "二群感度分析の徒歩",
    )
    close(walking_sensitivity["調整オッズ比"], 0.363408317, "二群感度分析の徒歩aOR")
    close(walking_sensitivity["95%CI下限"], 0.182372267, "二群感度分析の徒歩95%CI下限")
    close(walking_sensitivity["95%CI上限"], 0.724153992, "二群感度分析の徒歩95%CI上限")


def validate_step10_to_step12() -> None:
    for name in [
        "Table10-1_多変量解析主要結果.csv",
        "Table10-2_モデル診断.csv",
        "Table10-3_交差検証.csv",
        "Table10-4_通学方法二群感度分析.csv",
    ]:
        source = STEP10 / "Table" / name
        copied = STEP12 / "Table" / name
        require(source.exists() and copied.exists(), f"Step10→12の表が不足しています: {name}")
        require(sha256(source) == sha256(copied), f"Step10→12の表が一致しません: {name}")

    for extension in ["png", "svg"]:
        source = STEP10 / "Figure" / f"Figure10-1_調整オッズ比.{extension}"
        copied = STEP12 / "Figure" / f"Figure1_調整オッズ比.{extension}"
        require(source.exists() and copied.exists(), f"Step10→12の図が不足しています: {extension}")
        require(sha256(source) == sha256(copied), f"Step10→12の図が一致しません: {extension}")

    workbook_path = STEP12 / "Table" / "Step12_論文用表統合.xlsx"
    require(workbook_path.exists(), "Step12統合Excelがありません")
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    require(workbook.sheetnames == ["Table1", "Table2", "Table3", "Table4"], "統合Excelのシート構成が不一致です")
    workbook.close()

    inventory = csv(STEP12 / "Step12-04_図表・公開可否一覧.csv")
    expected = {
        "Figure1_調整オッズ比.png",
        "Figure1_調整オッズ比.svg",
        "Table10-1_多変量解析主要結果.csv",
        "Table10-2_モデル診断.csv",
        "Table10-3_交差検証.csv",
        "Table10-4_通学方法二群感度分析.csv",
        "Step12_論文用表統合.xlsx",
    }
    require(expected.issubset(set(inventory["ファイル"])), "Step12の公開可否一覧に不足があります")


def validate_blinding_and_privacy() -> None:
    internal = STEP11 / "InternalReview"
    after_review = internal / "AfterReview_DoNotOpen"
    stopped = STEP11 / "Internal_DoNotPublish"
    private_dirs = [
        STEP9 / "Internal_DoNotPublish", internal, after_review, stopped,
        STEP12 / "Internal_DoNotPublish",
    ]
    backup_root = CREATE / "RegenerationBackups"
    if backup_root.exists():
        private_dirs.append(backup_root)
    readonly_roots: list[Path] = []
    for root in [internal / "FrozenOriginals", internal / "FrozenConsensus"]:
        if root.exists():
            readonly_roots.extend(path for path in root.iterdir() if path.is_dir())
    if (internal / "FinalRules").exists():
        readonly_roots.append(internal / "FinalRules")

    def is_frozen(path: Path) -> bool:
        return any(path == root or root in path.parents for root in readonly_roots)

    for directory in private_dirs:
        require(directory.exists(), f"非公開ディレクトリがありません: {directory}")
        expected_directory_mode = 0o500 if is_frozen(directory) else 0o700
        require(
            directory.stat().st_mode & 0o777 == expected_directory_mode,
            f"非公開ディレクトリの権限が{expected_directory_mode:04o}ではありません: {directory}",
        )
        for path in directory.rglob("*"):
            require(not path.is_symlink(), f"非公開領域にシンボリックリンクがあります: {path}")
            if path.is_dir():
                expected_mode = 0o500 if is_frozen(path) else 0o700
                require(
                    path.stat().st_mode & 0o777 == expected_mode,
                    f"非公開サブディレクトリの権限が{expected_mode:04o}ではありません: {path}",
                )
            elif path.is_file():
                expected_mode = 0o400 if is_frozen(path) else 0o600
                require(
                    path.stat().st_mode & 0o777 == expected_mode,
                    f"非公開ファイルの権限が{expected_mode:04o}ではありません: {path}",
                )

    full = csv(internal / "Step11-01_通学中590件_盲検レビューシート_内部用.csv")
    require(len(full) == 590, "590件盲検シートの行数が不一致です")
    require(not any(token in column for column in full.columns for token in ["内部ID", "歯牙障害", "自動："]), "590件盲検シートに伏せるべき列があります")
    require(full["レビューID"].str.fullmatch(r"R-[0-9A-F]{10}").all(), "レビューIDが不透明ID形式ではありません")
    require(full["レビューID"].nunique() == 590, "レビューIDが重複しています")
    step11_source = (PROJECT / "file" / "Step11_通学中臨床レビュー" / "Step11_Main.py").read_text(encoding="utf-8")
    require("secrets.token_hex" in step11_source, "レビューIDが暗号学的乱数で生成されていません")
    require("SportDent-Step11-" not in step11_source, "レビューIDに公開された決定的対応規則が残っています")
    full_lookup = full.set_index("レビューID")["原文（内部確認用）"]
    answer_columns = [column for column in full.columns if column not in ["レビューID", "原文（内部確認用）"]]
    require((full[answer_columns] == "").all().all(), "590件盲検シートに既存の人手回答があります")

    phases: dict[str, set[str]] = {}
    for phase, step_number in [("開発用", "08"), ("最終評価用", "16")]:
        reviewer_frames = {}
        for reviewer in ["A", "B"]:
            path = internal / f"Step11-{step_number}{reviewer}_{phase}_評価者{reviewer}_盲検100件.csv"
            frame = csv(path)
            reviewer_frames[reviewer] = frame
            require(len(frame) == 100, f"{phase}評価者{reviewer}が100件ではありません")
            prohibited = ["内部ID", "歯牙障害", "自動", f"評価者{'B' if reviewer == 'A' else 'A'}："]
            require(
                not any(token in column for column in frame.columns for token in prohibited),
                f"{phase}評価者{reviewer}ファイルに伏せるべき列があります",
            )
            answers = [column for column in frame.columns if column.startswith(f"評価者{reviewer}：")]
            require(answers and (frame[answers] == "").all().all(), f"{phase}評価者{reviewer}に既存回答があります")
            require(set(frame["レビューID"]).issubset(set(full["レビューID"])), f"{phase}評価者{reviewer}に全590件外のIDがあります")
            expected_text = frame["レビューID"].map(full_lookup)
            require(expected_text.equals(frame["原文（内部確認用）"]), f"{phase}評価者{reviewer}のIDと原文が対応しません")
        require(
            reviewer_frames["A"][["レビューID", "原文（内部確認用）"]].equals(
                reviewer_frames["B"][["レビューID", "原文（内部確認用）"]]
            ),
            f"{phase}の評価者A/Bで標本が一致しません",
        )
        phases[phase] = set(reviewer_frames["A"]["レビューID"])
    require(phases["開発用"].isdisjoint(phases["最終評価用"]), "開発用と最終評価用の標本が重複しています")
    development_text = set(full_lookup.loc[list(phases["開発用"])])
    validation_text = set(full_lookup.loc[list(phases["最終評価用"])])
    require(development_text.isdisjoint(validation_text), "開発用と最終評価用に同一原文があります")

    codebook = csv(STEP11 / "Step11-18_固定コードブック候補.csv")
    require(
        list(codebook.columns) == ["コードブック版", "項目", "選択値", "定義", "コメント必須", "状態"],
        "Step11固定コードブック候補の列が不一致です",
    )
    require(set(codebook["コードブック版"]) == {"Step11-CB-1.0.0-rc1"}, "Step11コードブック版が不一致です")
    approval_sheet_path = internal / "Step11-19_コードブック承認チェックシート.csv"
    require(approval_sheet_path.exists(), "Step11コードブック承認チェックシートがありません")
    approval_sheet = csv(approval_sheet_path)
    require(
        list(approval_sheet.columns) == ["項目ID", "確認項目", "入力値", "必須", "入力例・推奨案", "説明"],
        "Step11コードブック承認チェックシートの列が不一致です",
    )
    expected_approval_ids = [
        "sheet_version", "codebook_version", "codebook_read_confirmed", "approved_by",
        "reviewer_a_confirmed", "reviewer_b_confirmed", "adjudicator", "protection_na_rule",
        "agreement_threshold", "sparse_category_rule", "approval_date", "approval_intent",
    ]
    require(approval_sheet["項目ID"].tolist() == expected_approval_ids, "Step11承認チェックシートの項目または行順が不一致です")
    approval_values = dict(zip(approval_sheet["項目ID"], approval_sheet["入力値"], strict=True))
    require(approval_values["sheet_version"] == "Step11-ApprovalSheet-1.0", "Step11承認チェックシート版が不一致です")
    require(approval_values["codebook_version"] == "Step11-CB-1.0.0-rc1", "Step11承認対象コードブック版が不一致です")
    require(
        not any(value.lstrip().startswith(("=", "+", "-", "@")) for value in approval_values.values()),
        "Step11承認チェックシートに数式として解釈され得る入力があります",
    )
    approval_record = internal / "CODEBOOK_APPROVED.json"
    human_approval_ids = expected_approval_ids[2:]
    if approval_record.exists():
        require(all(approval_values[item_id].strip() for item_id in human_approval_ids), "承認済みなのに承認チェックシートに空欄があります")
        require(
            approval_values["approval_intent"] == "この内容で開発用コードブックを承認する",
            "Step11承認チェックシートの明示承認文が不一致です",
        )
    else:
        require(
            approval_values["codebook_read_confirmed"] in {"", "Step11-02を読んで確認済み"},
            "Step11承認チェックシートの確認文が不正です",
        )
        require(
            approval_values["approval_intent"] in {"", "この内容で開発用コードブックを承認する"},
            "Step11承認チェックシートの明示承認文が不正です",
        )
        require(
            not approval_values["approval_date"] or re.fullmatch(r"\d{4}-\d{2}-\d{2}", approval_values["approval_date"]) is not None,
            "Step11承認チェックシートの承認日形式が不正です",
        )
    require((internal / "REVIEW_WORKFLOW_INITIALIZED.lock").exists(), "Step11 Excel初期化記録がありません")
    expected_sheets = {"入力", "判定基準", "選択肢", "メタデータ"}
    expected_input_columns = [
        "レビューID", "原文（内部確認用）", "起点機転", "出来事の順序", "最終接触対象",
        "口腔・顔面への直接外力", "前歯部の記載", "その他歯牙の記載", "歯牙・部位不明の記載",
        "口腔・口唇の記載", "顔面・顎の記載", "その他部位の記載", "ヘルメット", "マウスガード",
        "予防可能性", "予防可能性の根拠区分", "判定不能理由", "コメント", "行確認",
    ]
    for reviewer in ["A", "B"]:
        excel_path = internal / f"Step11-08{reviewer}_開発用_評価者{reviewer}_盲検100件.xlsx"
        require(excel_path.exists(), f"開発用評価者{reviewer}のExcelがありません")
        workbook = openpyxl.load_workbook(excel_path, data_only=False, keep_links=True)
        require(set(workbook.sheetnames) == expected_sheets, f"評価者{reviewer}Excelのシート構成が不一致です")
        require(not getattr(workbook, "_external_links", []), f"評価者{reviewer}Excelに外部リンクがあります")
        for sheet in workbook.worksheets:
            require(sheet.sheet_state == "visible", f"評価者{reviewer}Excelに非表示シートがあります: {sheet.title}")
            require(
                not any(cell.data_type == "f" for row in sheet.iter_rows() for cell in row),
                f"評価者{reviewer}Excelにセル計算式があります: {sheet.title}",
            )
        input_sheet = workbook["入力"]
        headers = [input_sheet.cell(3, column).value for column in range(1, len(expected_input_columns) + 1)]
        require(headers == expected_input_columns, f"評価者{reviewer}Excelの入力列が不一致です")
        require(input_sheet.protection.sheet, f"評価者{reviewer}Excelの入力シート保護が解除されています")
        excel_ids = [str(input_sheet.cell(row, 1).value or "") for row in range(4, 104)]
        excel_texts = [str(input_sheet.cell(row, 2).value or "") for row in range(4, 104)]
        source = csv(internal / f"Step11-08{reviewer}_開発用_評価者{reviewer}_盲検100件.csv")
        require(excel_ids == source["レビューID"].tolist(), f"評価者{reviewer}ExcelのIDが元CSVと一致しません")
        require(excel_texts == source["原文（内部確認用）"].tolist(), f"評価者{reviewer}Excelの原文が元CSVと一致しません")
        meta = {
            str(row[0]): "" if row[1] is None else str(row[1])
            for row in workbook["メタデータ"].iter_rows(min_col=1, max_col=2, values_only=True)
            if row[0] is not None
        }
        require(meta.get("コードブック版") == "Step11-CB-1.0.0-rc1", f"評価者{reviewer}Excelのコードブック版が不一致です")
        require(meta.get("承認状態") in {"研究者確認待ち（入力禁止）", "承認済み"}, f"評価者{reviewer}Excelの承認状態が不正です")
        expected_answer_locked = meta.get("承認状態") != "承認済み"
        require(
            input_sheet["A4"].protection.locked and input_sheet["B4"].protection.locked
            and input_sheet["C4"].protection.locked == expected_answer_locked,
            f"評価者{reviewer}Excelのセル保護が承認状態と不一致です",
        )
        workbook.close()

    mapping = csv(after_review / "Step11-01B_レビューID対応・自動検索補助_評価完了まで非表示.csv")
    require(mapping["レビューID"].nunique() == 590, "内部対応表のレビューIDが重複しています")
    ordered = mapping.sort_values("レビューID")["歯牙障害"].eq("歯牙障害").tolist()
    best_threshold_accuracy = 0.0
    for threshold in range(len(ordered) + 1):
        forward = sum((index < threshold) == outcome for index, outcome in enumerate(ordered)) / len(ordered)
        reverse = sum((index >= threshold) == outcome for index, outcome in enumerate(ordered)) / len(ordered)
        best_threshold_accuracy = max(best_threshold_accuracy, forward, reverse)
    require(best_threshold_accuracy < 0.70, f"レビューIDの単純な番号帯が群を予測します: {best_threshold_accuracy:.3f}")
    longest_run = 1
    current_run = 1
    for previous, current in zip(ordered, ordered[1:]):
        current_run = current_run + 1 if current == previous else 1
        longest_run = max(longest_run, current_run)
    require(longest_run < 30, f"レビューID順に群の長い連続があります: {longest_run}")
    for phase, filename in [
        ("開発用", "Step11-14_開発用100件_自動回答_評価完了まで非表示.csv"),
        ("最終評価用", "Step11-17_最終評価用100件_自動回答_評価完了まで非表示.csv"),
    ]:
        reference = csv(after_review / filename)
        require(set(reference["レビューID"]) == phases[phase], f"{phase}の自動回答参照ファイルと盲検標本が一致しません")

    require(
        not list(STEP11.glob("Step11-1[12]*自動*.csv")),
        "使用停止中のStep11-11/12自動分類表が公開直下に残っています",
    )
    require(
        not list(STEP11.glob("Step11-15*自動*.md")),
        "使用停止中のStep11-15自動分類レポートが公開直下に残っています",
    )
    for private_name in [
        "Step11-11_自動レビュー結果集計_使用停止.csv",
        "Step11-12_自動暫定分類別集計_使用停止.csv",
        "Step11-15_自動レビュー暫定解析レポート_使用停止.md",
    ]:
        require((stopped / private_name).exists(), f"使用停止成果物の隔離版がありません: {private_name}")
    for extension in ["png", "svg"]:
        require(
            not (STEP11 / "Figure" / f"Step11-13_自動暫定_主受傷機転別_歯牙障害割合.{extension}").exists(),
            f"使用停止図が公開Figureに残っています: {extension}",
        )
        require(
            (stopped / f"Step11-13_自動暫定_主受傷機転別_歯牙障害割合.{extension}").exists(),
            f"使用停止図の隔離版がありません: {extension}",
        )

    def validate_public_outcome_table(path: Path, group_column: str | None = None) -> pd.DataFrame:
        frame = csv(path)
        total = pd.to_numeric(frame["全体件数"], errors="coerce")
        dental = pd.to_numeric(frame["歯牙障害件数"], errors="coerce")
        require(not total.between(1, 4).any(), f"公開表に全体1〜4件セルがあります: {path}")
        visible = dental.notna()
        require(not dental.loc[visible].between(1, 4).any(), f"公開表に歯牙障害1〜4件セルがあります: {path}")
        non_dental = total.loc[visible] - dental.loc[visible]
        require(not non_dental.between(1, 4).any(), f"公開表に歯牙障害以外1〜4件セルがあります: {path}")
        hidden = ~visible
        if hidden.any():
            require(
                pd.to_numeric(frame.loc[hidden, "歯牙障害割合（％）"], errors="coerce").isna().all(),
                f"小集計を伏せた行に割合が残っています: {path}",
            )
            groups = frame[group_column] if group_column else pd.Series("全体", index=frame.index)
            for group in groups.unique():
                group_hidden = hidden.loc[groups.eq(group)]
                require(int(group_hidden.sum()) != 1, f"二次セル抑制が不足しています: {path} / {group}")
        return frame

    validate_public_outcome_table(STEP8 / "Step8-01_上位5因子カテゴリ別集計.csv", "因子")
    commute_public = validate_public_outcome_table(STEP11 / "Step11-04_通学方法別集計.csv")
    validate_public_outcome_table(STEP11 / "Step11-05_登下校別集計.csv")
    hidden_commute_categories = commute_public.loc[
        pd.to_numeric(commute_public["歯牙障害件数"], errors="coerce").isna(), "通学方法"
    ]
    commute_svg = (STEP11 / "Figure" / "Step11-06_通学方法別_歯牙障害割合.svg").read_text(encoding="utf-8")
    for category in hidden_commute_categories:
        require(str(category) not in commute_svg, f"小集計カテゴリが公開図に残っています: {category}")

    for filename in [
        "Step11-03_複数ラベル語彙検出集計_非識別.csv",
        "Step11-03A_歯牙障害群内_複数ラベル語彙検出集計.csv",
    ]:
        frame = csv(STEP11 / filename)
        target = pd.to_numeric(frame["対象件数"], errors="coerce")
        detected = pd.to_numeric(frame["検出件数"], errors="coerce")
        visible = detected.notna()
        require(not target.between(1, 4).any(), f"語彙検出表に対象1〜4件セルがあります: {filename}")
        require(not detected.loc[visible].between(1, 4).any(), f"語彙検出表に検出1〜4件セルがあります: {filename}")
        require(not (target.loc[visible] - detected.loc[visible]).between(1, 4).any(), f"語彙検出表に非検出1〜4件セルがあります: {filename}")
        if "群" in frame and frame["群"].nunique() > 1:
            for _, part in frame.groupby("複数ラベル語彙検出"):
                require(int(pd.to_numeric(part["検出件数"], errors="coerce").isna().sum()) in {0, len(part)}, f"語彙検出表の補完防止が不十分です: {filename}")

    for path in list(STEP11.glob("*.csv")) + list((STEP11 / "Figure").glob("*.csv")):
        columns = csv(path).columns
        require(not any(column in {"内部ID", "記号", "原文（内部確認用）"} for column in columns), f"公開CSVに機微列があります: {path}")

    internal_ids = set(mapping["内部ID"].astype(str))
    original_texts = {text for text in mapping["原文（内部確認用）"].astype(str) if text}
    public_files: list[Path] = []
    for root in [STEP8, STEP9, STEP10, STEP11, STEP12, STUDY_AI]:
        for path in root.rglob("*"):
            if not path.is_file() or path.suffix.lower() not in {".csv", ".md", ".svg", ".xlsx"}:
                continue
            if any(private in path.parents for private in [internal, stopped, STEP9 / "Internal_DoNotPublish", STEP12 / "Internal_DoNotPublish"]):
                continue
            public_files.append(path)
    for path in public_files:
        values: list[str] = []
        if path.suffix.lower() == ".csv":
            frame = csv(path)
            values = [str(value) for value in frame.to_numpy().ravel()]
        elif path.suffix.lower() == ".xlsx":
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            values = [str(cell) for sheet in workbook for row in sheet.iter_rows(values_only=True) for cell in row if cell is not None]
            workbook.close()
        else:
            values = [path.read_text(encoding="utf-8")]
        joined = "\n".join(values)
        require(not any(identifier and identifier in joined for identifier in internal_ids), f"公開成果物に内部IDがあります: {path}")
        if path.suffix.lower() in {".csv", ".xlsx"}:
            require(set(values).isdisjoint(original_texts), f"公開成果物に事例原文があります: {path}")
        else:
            require(not any(text in joined for text in original_texts if len(text) >= 10), f"公開成果物に事例原文があります: {path}")


def validate_traceability() -> None:
    evidence = csv(STUDY_AI / "05_根拠データ対応表.csv")
    require(list(evidence.columns) == ["区分", "所見", "主な根拠ファイル", "確定度", "必要な確認"], "StudyAI根拠表が5列ではありません")
    for reference in evidence["主な根拠ファイル"]:
        require((PROJECT / reference).exists(), f"StudyAIの根拠参照先がありません: {reference}")

    manifest = csv(STEP12 / "Step12-06_入力・コードハッシュ.csv")
    require(list(manifest.columns) == ["区分", "ファイル", "SHA-256", "バイト数"], "Step12ハッシュ表の列が不一致です")
    for _, row in manifest.iterrows():
        path = PROJECT / row["ファイル"]
        require(path.exists(), f"ハッシュ対象がありません: {path}")
        require(sha256(path) == row["SHA-256"], f"ハッシュが変化しています: {path}")
        require(path.stat().st_size == int(row["バイト数"]), f"バイト数が変化しています: {path}")
    manifest_files = set(manifest["ファイル"])
    for required_code in [
        "file/Step7_探索的解析/Step7_Main.py",
        "file/Step8_多変量解析/Step8_Main.py",
        "file/Step9_感度分析/Step9_Main.py",
        "file/Step10_統合論文原稿/Step10_Main.py",
        "file/Step11_通学中臨床レビュー/Step11_Main.py",
        "file/Step11_通学中臨床レビュー/Step11_ReviewWorkflow.py",
        "file/Step11_通学中臨床レビュー/Step11_PostReviewWorkflow.py",
        "file/Step11_通学中臨床レビュー/Step11_ReviewAnalysis.py",
        "file/Step11_通学中臨床レビュー/Step11_ReviewCodebook_candidate.md",
        "file/Step12_投稿準備/Step12_Main.py",
        "file/Step12_投稿準備/Validate_Step8_12.py",
    ]:
        require(required_code in manifest_files, f"再現用ハッシュ表に依存コードがありません: {required_code}")
    require(
        sha256(HERE / "requirements-lock.txt") == sha256(STEP12 / "Step12-07_requirements-lock.txt"),
        "requirements-lockの元ファイルとStep12コピーが一致しません",
    )

    study_text = "\n".join(path.read_text(encoding="utf-8") for path in sorted(STUDY_AI.glob("*.md")))
    for required_text in [
        "給付年度（西暦換算）",
        "7,104",
        "0.000371",
        "558件",
        "5.38年",
        "通学（通園）に準ずるとき",
    ]:
        require(required_text in study_text, f"StudyAIへの反映表現が見つかりません: {required_text}")


def main() -> None:
    validate_statistics()
    validate_step10_to_step12()
    validate_blinding_and_privacy()
    validate_traceability()
    print("VALIDATION_OK: Step8〜12 / StudyAI / Step11盲検・非公開管理")


if __name__ == "__main__":
    main()
