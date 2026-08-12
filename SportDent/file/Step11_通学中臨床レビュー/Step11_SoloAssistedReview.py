"""Step11を一人で進めるための、厳格なAI仮分類と確認用Excelを作る。

この工程は2名独立評価の代替ではない。A/B原本は変更せず、AI案を研究者1名が
原文と照合・修正して確定候補にする。曖昧な記述は積極的に判定不能へ送る。
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

import Step11_ReviewWorkflow as workflow


HERE = Path(__file__).resolve().parent
OUTPUT = workflow.INTERNAL / "SoloAssistedReview"
PROPOSAL_CSV = OUTPUT / "Step11-31_Codex暫定分類_開発用100件.csv"
REVIEW_XLSX = OUTPUT / "Step11-32_一人研究者確認用_開発用100件.xlsx"
SUMMARY_CSV = OUTPUT / "Step11-33_自動仮分類サマリー.csv"
PRIORITY_CSV = OUTPUT / "Step11-34_要確認事例.csv"
MANIFEST = OUTPUT / "Step11-35_AI仮分類マニフェスト.json"
CONFIRMED_CSV = OUTPUT / "Step11-36_一人研究者確認済み_確定候補.csv"
AI_REVIEWED_CSV = OUTPUT / "Step11-37_Codex二次検査済み_AI暫定版.csv"
AI_AGGREGATE_CSV = OUTPUT / "Step11-38_AI暫定分類_項目別集計.csv"
AI_REPORT = OUTPUT / "Step11-39_AI暫定分類_結果レポート.md"
AI_REVIEW_MANIFEST = OUTPUT / "Step11-40_AI二次検査記録.json"

SCHEMA_VERSION = "Step11-SoloAssistedReview-1.0"
STATUS = "AI_PROVISIONAL_NOT_GOLD"
SHEET = "確認入力"
HEADER_ROW = 3
FIRST_ROW = 4
AI_META_COLUMNS = ["AI確信度", "AI確認優先度", "AI判定根拠"]
REVIEW_COLUMNS = [workflow.ID_COLUMN, workflow.TEXT_COLUMN, *AI_META_COLUMNS, *workflow.CHOICES.keys()]
REVIEW_COLUMNS.insert(-1, workflow.COMMENT_COLUMN)


class SoloReviewError(RuntimeError):
    """原本を変更せず処理を止めるべき問題。"""


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def secure_output() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    OUTPUT.chmod(0o700)
    for path in OUTPUT.rglob("*"):
        path.chmod(0o700 if path.is_dir() else 0o600)


def write_json(path: Path, value: object) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    path.chmod(0o600)


def normalized(text: str) -> str:
    return re.sub(r"\s+", "", str(text)).replace("くう", "腔")


def hits(text: str, patterns: dict[str, str]) -> list[tuple[int, str, str]]:
    found: list[tuple[int, str, str]] = []
    for label, pattern in patterns.items():
        for match in re.finditer(pattern, text):
            found.append((match.start(), label, match.group(0)))
    return sorted(found)


def classify_origin(text: str) -> tuple[str, str, list[str]]:
    collision = r"正面衝突|追突|衝突|接触|激突|轢かれ|ひかれ|蹴られ|叩かれ|暴行を受け|ぶつかった|ぶつかって|ぶつかり|当たった|当たり"
    fall = r"つまず|バランスを崩|スリップ|滑って|滑り|転倒|転んだ|転び|前のめり|放り出され|投げ出され|飛ばされ|一回転"
    drop = r"転落|(?:階|側溝|用水路|土手|斜面|窓|溝).{0,30}(?:落ち|落下|突っ込)"
    events = hits(text, {"衝突・接触": collision, "転倒・つまずき": fall, "転落・落下": drop})
    if not events:
        if re.search(r"投げ合っていた石|傘を振った|傘が開いて.*眼|支柱.*眼|顔を負傷|傷を負った", text):
            return "衝突・接触", "単一機転", ["物体との接触が明記"]
        return "判定不能", "判定不能", ["起点となる物理的事象を特定できず"]

    # 「転落」は高低差を明示するため、先行する衝突がなければ転落を優先する。
    drop_events = [event for event in events if event[1] == "転落・落下"]
    collision_events = [event for event in events if event[1] == "衝突・接触"]
    if drop_events and not any(event[0] < drop_events[0][0] for event in collision_events):
        origin = drop_events[0]
    else:
        origin = events[0]

    # 滑る→バランスを崩す→転倒は同じ起点機転の記述として一つにまとめる。
    distinct: list[tuple[int, str, str]] = []
    for event in events:
        if event[1] not in {known[1] for known in distinct}:
            distinct.append(event)
    if len(distinct) == 1:
        order = "単一機転"
    else:
        order = "複数機転・順序明記" if distinct == sorted(distinct) else "複数機転・順序不明"
    evidence = [f"起点候補:{origin[2]}"]
    if len(distinct) > 1:
        evidence.append("複数出来事:" + "→".join(event[2] for event in distinct[:4]))
    return origin[1], order, evidence


SITE_PATTERNS = {
    "前歯部の記載": r"前歯|門歯|切歯|犬歯|上[123一二三１-３]本?の歯|上顎前歯",
    "その他歯牙の記載": r"奥歯|臼歯|小臼歯|大臼歯",
    "口腔・口唇の記載": r"口腔|口内|口元|口部|口の部分|口唇|下唇|上唇|唇|歯肉|舌|口蓋|頬粘膜",
    "顔面・顎の記載": r"顔面|顔|頬|ほお|顎|下顎|上顎|オトガイ|鼻|前額|額|眉|眼|目",
    "その他部位の記載": r"頭の側面|頭部|後頭部|側頭部|頸部|首|肩|鎖骨|腕|上肢|肘|手首|手掌|掌|指|胸部|胸|腹部|背中|背部|腰|骨盤|大腿|下肢|膝|足首|踵|足部",
}


INJURY_CONTEXT = (
    r"強打|打ちつけ|打ち付け|打った|ぶつけ|ぶつかった|当たった|当たり|衝突|接触|激突|"
    r"負傷|損傷|骨折|破折|脱臼|欠損|抜け|折れ|創|裂傷|擦り傷|切り|切っ|ひっかかれ|"
    r"出血|瘢痕|疼痛|麻痺|障害|捻挫|歯科補綴"
)


def affirmative_site_matches(text: str, pattern: str) -> list[str]:
    """単なる登場語でなく、今回事故の受傷・治療文脈にある部位だけを返す。"""
    found: list[str] = []
    for clause in re.split(r"[。；]", text):
        if not re.search(INJURY_CONTEXT, clause):
            continue
        for match in re.finditer(pattern, clause):
            start, end = match.span()
            context = clause[max(0, start - 24): min(len(clause), end + 24)]
            if re.search(INJURY_CONTEXT, context):
                found.append(match.group(0))
    return found


def classify_sites(text: str) -> tuple[dict[str, str], list[str], list[str]]:
    values: dict[str, str] = {}
    evidence: list[str] = []
    other_names: list[str] = []
    for field in workflow.SITE_COLUMNS:
        pattern = SITE_PATTERNS.get(field)
        matched = affirmative_site_matches(text, pattern) if pattern else []
        values[field] = "記載あり" if matched else "記載なし"
        if matched:
            unique = list(dict.fromkeys(matched))[:5]
            evidence.append(f"{field}:{'/'.join(unique)}")
            if field == "その他部位の記載":
                other_names = unique

    specific = values["前歯部の記載"] == "記載あり" or values["その他歯牙の記載"] == "記載あり"
    generic_dental = bool(re.search(r"(?:歯牙|歯科補綴|歯を|歯に|歯が|歯の|歯科).{0,16}(?:負傷|損傷|破折|脱臼|欠損|抜け|折れ|補綴)|(?:負傷|損傷|破折|脱臼|欠損|抜け|折れ).{0,16}(?:歯牙|歯を|歯に|歯が|歯の)", text))
    values["歯牙・部位不明の記載"] = "記載あり" if generic_dental and not specific else "記載なし"
    if values["歯牙・部位不明の記載"] == "記載あり":
        evidence.append("歯牙・部位不明:歯/歯牙/歯科記載")
    return values, evidence, other_names


def direct_force_and_target(text: str, site_values: dict[str, str]) -> tuple[str, str, list[str], bool]:
    site = r"前歯|歯牙|歯|口腔|口内|口元|口部|口唇|唇|顔面|顔|頬|ほお|顎|下顎|上顎|鼻|前額|額|眉|眼|目"
    force = r"強打|打ちつけ|打ち付け|打った|打ち|ぶつけ|ぶつかった|ぶつかって|当たった|当たり|衝突|接触|激突|蹴られ|叩かれ|ひっかかれ"
    direct_matches = re.findall(rf"(?:{site}).{{0,18}}(?:{force})|(?:{force}).{{0,18}}(?:{site})", text)
    surface = r"道路|路面|地面|歩道|アスファルト|床|雪道"
    landing_matches = re.findall(
        rf"(?:{site})(?:から|を|が).{{0,8}}(?:{surface})(?:に|へ|で).{{0,10}}(?:つき|倒れ込|突っ込|落ち)|"
        rf"(?:{surface})(?:に|へ|で).{{0,10}}(?:{site})(?:から|を|が).{{0,10}}(?:つき|倒れ込|突っ込|落ち)",
        text,
    )
    direct_matches.extend(landing_matches)
    face_or_oral = any(site_values[field] == "記載あり" for field in workflow.SITE_COLUMNS[:-1])
    if not direct_matches:
        if face_or_oral:
            return "記載なし", "判定不能", ["部位記載はあるが直接外力・対象の明記不足"], True
        return "記載なし", "該当なし", ["口腔・顔面の記載なし"], False

    target_patterns = {
        "路面・地面・床": r"道路|路面|地面|歩道|アスファルト|床|雪道",
        "車両": r"自動車|車両|電車|レールバス|相手の自転車|友人の自転車",
        "自転車部品": r"自転車のハンドル|ハンドル|前輪|後輪|かご|ペダル|スポーク",
        "人": r"友人|児童|生徒|歩行者|他の人|相手",
        "構造物": r"ガードレール|ポール|塀|縁石|消火器ボックス|ドア|壁|柱|墓|側溝|鉄板",
        "その他": r"傘|石|支柱|ラケット|グリップ|何か",
    }
    categories: list[str] = []
    for label, pattern in target_patterns.items():
        # 助詞を伴う接触構文に限定し、単なる移動手段・登場人物を対象にしない。
        relation = (
            rf"(?:{site}).{{0,10}}(?:を|が)?(?:{pattern})(?:に|で|へ).{{0,10}}(?:{force})|"
            rf"(?:{pattern})(?:に|で|へ|の).{{0,14}}(?:{site}).{{0,10}}(?:{force})|"
            rf"(?:{pattern})(?:に|で|へ).{{0,12}}(?:{force}).{{0,12}}(?:{site})|"
            rf"(?:{site})(?:を|が)?(?:{pattern})(?:に|で|へ).{{0,12}}(?:強く)?(?:打|ぶつ|当)|"
            rf"(?:{site})(?:から|を|が).{{0,8}}(?:{pattern})(?:に|へ|で).{{0,10}}(?:つき|倒れ込|突っ込|落ち)|"
            rf"(?:{pattern})(?:が|は).{{0,18}}(?:{site})(?:に|を).{{0,8}}(?:{force})"
        )
        if re.search(relation, text):
            categories.append(label)
    categories = list(dict.fromkeys(categories))
    explicit_object = re.search(
        rf"(?:傘|石|支柱|ラケット|グリップ|何か)(?:が|は|の[^、。]{{0,8}}が).{{0,18}}(?:{site})(?:に|を).{{0,8}}(?:{force})",
        text,
    )
    if explicit_object:
        categories = ["その他"]
    if len(categories) == 1:
        return "あり（明記）", categories[0], [f"直接外力:{direct_matches[0]}", f"対象:{categories[0]}"], False
    if not categories:
        return "あり（明記）", "判定不能", [f"直接外力:{direct_matches[0]}", "接触対象を特定できず"], True
    return "あり（明記）", "判定不能", [f"直接外力:{direct_matches[0]}", "対象候補:" + "/".join(categories)], True


def protection(text: str) -> tuple[str, str, list[str]]:
    evidence: list[str] = []
    if re.search(r"ヘルメットを(?:かぶ|着用)|ヘルメット着用", text):
        helmet = "使用あり（明記）"
        evidence.append("ヘルメット使用明記")
    elif re.search(r"ヘルメットを(?:かぶらず|着用せず)|ヘルメットなし", text):
        helmet = "使用なし（明記）"
        evidence.append("ヘルメット不使用明記")
    elif "ヘルメット" in text:
        helmet = "言及あり・使用状況不明"
        evidence.append("ヘルメット言及")
    elif re.search(r"徒歩|歩いて|歩行|鉄道|電車|バス", text) and not re.search(r"自転車|原動機付", text):
        helmet = "該当なし"
    else:
        helmet = "記載なし"

    if re.search(r"マウスガードを(?:使用|装着)|マウスピースを(?:使用|装着)", text):
        mouthguard = "使用あり（明記）"
        evidence.append("マウスガード使用明記")
    elif re.search(r"マウスガードを(?:使用せず|装着せず)|マウスガードなし", text):
        mouthguard = "使用なし（明記）"
        evidence.append("マウスガード不使用明記")
    elif re.search(r"マウスガード|マウスピース", text):
        mouthguard = "言及あり・使用状況不明"
        evidence.append("マウスガード言及")
    else:
        mouthguard = "記載なし"
    return helmet, mouthguard, evidence


PREVENTION_PATTERNS = {
    "行動": r"よそ見|前方不注意|スピードの出しすぎ|スピードが出|急ぎ足|走り出|ふざけ|鬼ごっこ|遊んで|暗く|メガネを忘|後ろを振り向|背負い直|戻そう|拾おう",
    "路面・天候": r"凍結|雨|濡れ|ぬれ|強風|砂利|段差|凹凸|くぼみ|亀裂|盛り上が|滑り止め|側溝|マンホール|鉄板|暗かった",
    "車両・操作": r"ブレーキがきか|ブレーキをかけ|ハンドル操作を誤|曲がり切れ|ペダルの不具合|前輪がロック|タイヤをとら|前輪に.*絡|前輪に.*入り|かばん.*前かご",
    "施設・構造物": r"ガードレールが切れ|ふたのない側溝|警報機・遮断機ともになし|ポールに気づかず|車輪止め",
    "保護具": r"ヘルメットをかぶらず|マウスガードを使用せず",
}


def preventability(text: str) -> tuple[str, str, list[str]]:
    found = [label for label, pattern in PREVENTION_PATTERNS.items() if re.search(pattern, text)]
    if not found:
        return "修正可能要因の記載なし", "該当なし", []
    basis = found[0] if len(found) == 1 else "複数"
    return "修正可能要因の明記あり", basis, ["修正可能要因:" + "/".join(found)]


def classify(text_original: str) -> dict[str, str]:
    text = normalized(text_original)
    origin, order, origin_evidence = classify_origin(text)
    sites, site_evidence, other_names = classify_sites(text)
    direct, target, force_evidence, target_uncertain = direct_force_and_target(text, sites)
    helmet, mouthguard, protection_evidence = protection(text)
    prevent, basis, prevention_evidence = preventability(text)

    values: dict[str, str] = {
        "起点機転": origin,
        "出来事の順序": order,
        "最終接触対象": target,
        "口腔・顔面への直接外力": direct,
        **sites,
        "ヘルメット": helmet,
        "マウスガード": mouthguard,
        "予防可能性": prevent,
        "予防可能性の根拠区分": basis,
    }
    uncertain = [field for field, value in values.items() if value == "判定不能"]
    if uncertain:
        if order == "複数機転・順序不明":
            reason = "時系列不明"
        elif target_uncertain and direct == "あり（明記）":
            reason = "複数候補" if "対象候補" in " ".join(force_evidence) else "原文情報不足"
        else:
            reason = "原文情報不足"
    else:
        reason = "該当なし"
    values["判定不能理由"] = reason

    comments: list[str] = []
    if other_names:
        comments.append("その他部位=" + "/".join(other_names))
    if target == "その他":
        comments.append("接触対象は原文の物体名を要確認")
    if origin == "その他":
        comments.append("起点機転は既定3分類外")
    values[workflow.COMMENT_COLUMN] = "；".join(comments)
    values[workflow.CONFIRM_COLUMN] = ""

    evidence = origin_evidence + force_evidence + site_evidence + protection_evidence + prevention_evidence
    if uncertain:
        confidence, priority = "低", "最優先"
    elif target == "その他" or order.startswith("複数") or prevent != "修正可能要因の記載なし":
        confidence, priority = "中", "要確認"
    else:
        confidence, priority = "高", "通常"
    values["AI確信度"] = confidence
    values["AI確認優先度"] = priority
    values["AI判定根拠"] = "｜".join(evidence[:12])

    problems = workflow.row_problems(values)
    if problems:
        raise SoloReviewError("AI仮分類がコードブック検査に失敗: " + " / ".join(problems))
    return values


def read_source() -> pd.DataFrame:
    path = workflow.source_path("A")
    frame = pd.read_csv(path, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    needed = [workflow.ID_COLUMN, workflow.TEXT_COLUMN]
    if not set(needed).issubset(frame.columns) or len(frame) != workflow.EXPECTED_ROWS:
        raise SoloReviewError("開発用100件CSVの構造または件数が不正です")
    if frame[workflow.ID_COLUMN].duplicated().any() or frame[workflow.TEXT_COLUMN].str.strip().eq("").any():
        raise SoloReviewError("開発用100件CSVにID重複または原文空欄があります")
    return frame[needed].copy()


def proposal_frame() -> pd.DataFrame:
    source = read_source()
    rows: list[dict[str, str]] = []
    for _, source_row in source.iterrows():
        values = classify(source_row[workflow.TEXT_COLUMN])
        rows.append({workflow.ID_COLUMN: source_row[workflow.ID_COLUMN], workflow.TEXT_COLUMN: source_row[workflow.TEXT_COLUMN], **values})
    return pd.DataFrame(rows)


def add_choices(book: Workbook) -> dict[str, str]:
    sheet = book.create_sheet("選択肢")
    result: dict[str, str] = {}
    for index, (field, choices) in enumerate(workflow.CHOICES.items(), start=1):
        column = get_column_letter(index)
        sheet.cell(1, index, field)
        for row, choice in enumerate(choices, start=2):
            sheet.cell(row, index, choice)
        name = f"solo_choices_{index:02d}"
        book.defined_names.add(DefinedName(name, attr_text=f"'選択肢'!${column}$2:${column}${len(choices)+1}"))
        result[field] = name
    return result


def build_workbook(frame: pd.DataFrame, destination: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    choice_names = add_choices(book)
    codebook = book.create_sheet("判定基準")
    codebook["A1"] = f"承認済みコードブック {workflow.CODEBOOK_VERSION}"
    codebook["A2"] = "AI案は正解ではありません。原文と照合し、必要なら修正してから行確認を選択してください。"
    for row, line in enumerate(workflow.CODEBOOK_SOURCE.read_text(encoding="utf-8").splitlines(), start=4):
        codebook.cell(row, 1, line)
    codebook.column_dimensions["A"].width = 120

    metadata = book.create_sheet("メタデータ")
    meta_rows = [
        ("方式", "AI仮分類＋一人研究者確認"), ("評価者間一致", "算出不可（2名独立評価ではない）"),
        ("状態", STATUS), ("コードブック版", workflow.CODEBOOK_VERSION),
        ("作成日時UTC", utc_now()), ("元CSV SHA-256", sha256(workflow.source_path("A"))),
        ("分類コード SHA-256", sha256(Path(__file__).resolve())),
    ]
    for row, (key, value) in enumerate(meta_rows, start=1):
        metadata.cell(row, 1, key).font = Font(bold=True)
        metadata.cell(row, 2, value)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REVIEW_COLUMNS))
    sheet["A1"] = "Step11 一人研究者用：AI仮分類を原文と照合して確認・修正"
    sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(REVIEW_COLUMNS))
    sheet["A2"] = "AI案は未確認です。『行確認=確認済み』を選ぶまでは研究結果に使用できません。"
    sheet["A2"].font = Font(bold=True, color="9C0006")
    sheet["A2"].fill = PatternFill("solid", fgColor="FFC7CE")

    for column, header in enumerate(REVIEW_COLUMNS, start=1):
        cell = sheet.cell(HEADER_ROW, column, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for offset, (_, row) in enumerate(frame.iterrows(), start=FIRST_ROW):
        for column, header in enumerate(REVIEW_COLUMNS, start=1):
            value = row.get(header, "")
            cell = sheet.cell(offset, column, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            fixed = header in {workflow.ID_COLUMN, workflow.TEXT_COLUMN, *AI_META_COLUMNS}
            cell.protection = Protection(locked=fixed)
            if fixed:
                cell.fill = PatternFill("solid", fgColor="E7E6E6")
            else:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
        priority = str(row["AI確認優先度"])
        if priority == "最優先":
            sheet.cell(offset, REVIEW_COLUMNS.index("AI確認優先度") + 1).fill = PatternFill("solid", fgColor="F4CCCC")
        elif priority == "要確認":
            sheet.cell(offset, REVIEW_COLUMNS.index("AI確認優先度") + 1).fill = PatternFill("solid", fgColor="FCE5CD")

    for field, name in choice_names.items():
        column = REVIEW_COLUMNS.index(field) + 1
        validation = DataValidation(type="list", formula1=f"={name}", allow_blank=True)
        validation.showErrorMessage = True
        validation.error = "固定選択肢から選んでください。"
        sheet.add_data_validation(validation)
        validation.add(f"{get_column_letter(column)}{FIRST_ROW}:{get_column_letter(column)}{FIRST_ROW + len(frame)-1}")

    sheet.freeze_panes = "F4"
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(REVIEW_COLUMNS))}{FIRST_ROW + len(frame)-1}"
    for index, header in enumerate(REVIEW_COLUMNS, start=1):
        width = 18
        if header == workflow.TEXT_COLUMN:
            width = 80
        elif header == "AI判定根拠" or header == workflow.COMMENT_COLUMN:
            width = 46
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(FIRST_ROW, FIRST_ROW + len(frame)):
        sheet.row_dimensions[row].height = 72
    sheet.protection.sheet = True
    sheet.protection.password = "Step11Solo"
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = True
    for other in book.worksheets[1:]:
        other.protection.sheet = True
        other.protection.password = "Step11Solo"
    book.save(destination)


def prepare() -> None:
    expected = [PROPOSAL_CSV, REVIEW_XLSX, SUMMARY_CSV, PRIORITY_CSV, MANIFEST]
    existing = [path for path in expected if path.exists()]
    if existing:
        raise SoloReviewError("既存成果物を上書きしません: " + ", ".join(path.name for path in existing))
    if not workflow.APPROVAL.exists():
        raise SoloReviewError("コードブック承認記録がありません")
    OUTPUT.mkdir(parents=True, exist_ok=True)
    OUTPUT.chmod(0o700)
    frame = proposal_frame()

    with tempfile.TemporaryDirectory(prefix="step11-solo-", dir=workflow.INTERNAL) as temp_name:
        temp = Path(temp_name)
        proposal = temp / PROPOSAL_CSV.name
        workbook = temp / REVIEW_XLSX.name
        summary = temp / SUMMARY_CSV.name
        priority = temp / PRIORITY_CSV.name
        manifest = temp / MANIFEST.name
        frame.to_csv(proposal, index=False, encoding="utf-8-sig")
        build_workbook(frame, workbook)
        summary_frame = (
            frame.groupby(["AI確認優先度", "AI確信度"], dropna=False).size().reset_index(name="件数")
        )
        summary_frame.to_csv(summary, index=False, encoding="utf-8-sig")
        frame.loc[frame["AI確認優先度"].isin(["最優先", "要確認"]), [
            workflow.ID_COLUMN, "AI確認優先度", "AI確信度", "AI判定根拠",
            "起点機転", "出来事の順序", "最終接触対象", "口腔・顔面への直接外力", "判定不能理由",
        ]].to_csv(priority, index=False, encoding="utf-8-sig")
        record = {
            "schema_version": SCHEMA_VERSION, "status": STATUS, "created_at_utc": utc_now(),
            "warning": "2名独立評価でもGold Standardでもない。研究者確認前は研究結果へ使用しない。",
            "source_csv": workflow.source_path("A").name,
            "source_csv_sha256": sha256(workflow.source_path("A")),
            "codebook_version": workflow.CODEBOOK_VERSION,
            "codebook_sha256": sha256(workflow.CODEBOOK_SOURCE),
            "approval_sha256": sha256(workflow.APPROVAL),
            "script_sha256": sha256(Path(__file__).resolve()),
            "rows": len(frame),
        }
        write_json(manifest, record)
        for path in temp.iterdir():
            path.chmod(0o600)
            os.replace(path, OUTPUT / path.name)
    secure_output()
    print(f"SOLO_PREPARE_OK: {len(frame)}件 / 確認Excel={REVIEW_XLSX}")


def read_review() -> pd.DataFrame:
    if not REVIEW_XLSX.exists():
        raise SoloReviewError("一人研究者確認用Excelがありません")
    book = load_workbook(REVIEW_XLSX, data_only=False, keep_links=True)
    forbidden: list[str] = []
    for current_sheet in book.worksheets:
        if current_sheet.sheet_state != "visible":
            forbidden.append(f"非表示シート:{current_sheet.title}")
        for row in current_sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    forbidden.append(f"数式:{current_sheet.title}!{cell.coordinate}")
                if cell.hyperlink is not None:
                    forbidden.append(f"リンク:{current_sheet.title}!{cell.coordinate}")
    if getattr(book, "_external_links", []):
        forbidden.append("外部リンク")
    if getattr(book, "vba_archive", None) is not None:
        forbidden.append("VBA")
    if forbidden:
        book.close()
        raise SoloReviewError("Excelに数式・リンク等の禁止内容があります: " + " / ".join(forbidden[:10]))
    if set(book.sheetnames) != {SHEET, "選択肢", "判定基準", "メタデータ"}:
        book.close()
        raise SoloReviewError("Excelのシート構成が不正です")
    sheet = book[SHEET]
    headers = [sheet.cell(HEADER_ROW, col).value for col in range(1, len(REVIEW_COLUMNS) + 1)]
    if headers != REVIEW_COLUMNS:
        book.close()
        raise SoloReviewError("Excelの列または列順が不正です")
    rows = []
    for row_number in range(FIRST_ROW, FIRST_ROW + workflow.EXPECTED_ROWS):
        rows.append({
            header: "" if sheet.cell(row_number, col).value is None else str(sheet.cell(row_number, col).value).strip()
            for col, header in enumerate(REVIEW_COLUMNS, start=1)
        })
    book.close()
    frame = pd.DataFrame(rows)
    source = read_source()
    if not frame[[workflow.ID_COLUMN, workflow.TEXT_COLUMN]].equals(source):
        raise SoloReviewError("レビューIDまたは原文が元CSVと一致しません")
    return frame


def validate(require_complete: bool = False) -> pd.DataFrame:
    if not MANIFEST.exists() or not PROPOSAL_CSV.exists():
        raise SoloReviewError("AI仮分類マニフェストまたは暫定分類CSVがありません")
    record = json.loads(MANIFEST.read_text(encoding="utf-8"))
    expected_record = {
        "schema_version": SCHEMA_VERSION,
        "status": STATUS,
        "source_csv_sha256": sha256(workflow.source_path("A")),
        "codebook_version": workflow.CODEBOOK_VERSION,
        "codebook_sha256": sha256(workflow.CODEBOOK_SOURCE),
        "approval_sha256": sha256(workflow.APPROVAL),
        "script_sha256": sha256(Path(__file__).resolve()),
        "rows": workflow.EXPECTED_ROWS,
    }
    mismatched = [key for key, value in expected_record.items() if record.get(key) != value]
    if mismatched:
        raise SoloReviewError("AI仮分類の版・ハッシュが現行仕様と一致しません: " + "/".join(mismatched))
    frame = read_review()
    proposal = pd.read_csv(PROPOSAL_CSV, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    fixed_columns = [workflow.ID_COLUMN, workflow.TEXT_COLUMN, *AI_META_COLUMNS]
    if not set(fixed_columns).issubset(proposal.columns) or not frame[fixed_columns].equals(proposal[fixed_columns]):
        raise SoloReviewError("Excelの固定列がAI仮分類CSVと一致しません")
    errors: list[str] = []
    completed = 0
    for index, row in frame.iterrows():
        values = {column: str(row.get(column, "")).strip() for column in workflow.INPUT_COLUMNS}
        problems = workflow.row_problems(values)
        if values[workflow.CONFIRM_COLUMN] == "確認済み":
            completed += 1
            missing = [column for column in workflow.REQUIRED_COLUMNS if not values.get(column)]
            if missing:
                problems.append("確認済みだが必須欄が空: " + "/".join(missing))
        elif values[workflow.CONFIRM_COLUMN]:
            problems.append("行確認が選択肢外")
        if problems:
            errors.append(f"行{index + FIRST_ROW}: " + " / ".join(problems))
    if require_complete and completed != workflow.EXPECTED_ROWS:
        errors.append(f"確認済み件数が100件ではありません: {completed}/100")
    if errors:
        raise SoloReviewError("\n".join(errors[:30]))
    print(f"SOLO_VALIDATE_OK: 確認済み {completed}/{workflow.EXPECTED_ROWS}件")
    return frame


def export_confirmed() -> None:
    if CONFIRMED_CSV.exists():
        raise SoloReviewError(f"確定候補を上書きしません: {CONFIRMED_CSV}")
    frame = validate(require_complete=True)
    output = frame[[workflow.ID_COLUMN, workflow.TEXT_COLUMN, *workflow.CHOICES.keys()]].copy()
    output.insert(2, "評価方式", "AI仮分類を一人研究者が全件確認")
    output.insert(3, "評価者間一致", "算出不可")
    output.to_csv(CONFIRMED_CSV, index=False, encoding="utf-8-sig")
    CONFIRMED_CSV.chmod(0o600)
    secure_output()
    print(f"SOLO_EXPORT_OK: {CONFIRMED_CSV}")


def export_ai_provisional() -> None:
    """人手確認とは呼ばず、AI案の完全性・整合性を再検査して暫定集計する。"""
    outputs = [AI_REVIEWED_CSV, AI_AGGREGATE_CSV, AI_REPORT, AI_REVIEW_MANIFEST]
    existing = [path for path in outputs if path.exists()]
    if existing:
        raise SoloReviewError("AI二次検査成果物を上書きしません: " + ", ".join(path.name for path in existing))
    # 現行コード・承認記録・固定列の版ずれも含めて先に検査する。
    validate(require_complete=False)
    proposal = pd.read_csv(PROPOSAL_CSV, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    errors: list[str] = []
    for index, row in proposal.iterrows():
        values = {column: str(row.get(column, "")).strip() for column in workflow.INPUT_COLUMNS}
        missing = [
            field for field in workflow.CHOICES
            if field != workflow.CONFIRM_COLUMN and not values.get(field)
        ]
        problems = workflow.row_problems(values)
        if missing:
            problems.append("AI必須分類が空: " + "/".join(missing))
        if problems:
            errors.append(f"事例{index + 1}: " + " / ".join(problems))
    if errors:
        raise SoloReviewError("AI二次検査に失敗:\n" + "\n".join(errors[:30]))

    reviewed = proposal.copy()
    reviewed.insert(2, "評価方式", "承認済みコードブックによるルール分類＋Codex二次整合性検査")
    reviewed.insert(3, "臨床確認状態", "歯科医師による最終確認前")
    reviewed.insert(4, "利用区分", "探索的な内部検討のみ")
    reviewed.to_csv(AI_REVIEWED_CSV, index=False, encoding="utf-8-sig")

    aggregate_rows: list[dict[str, object]] = []
    fields = [
        "起点機転", "出来事の順序", "最終接触対象", "口腔・顔面への直接外力",
        *workflow.SITE_COLUMNS, "ヘルメット", "マウスガード", "予防可能性",
        "予防可能性の根拠区分", "判定不能理由", "AI確認優先度", "AI確信度",
    ]
    for field in fields:
        counts = proposal[field].value_counts(dropna=False)
        for value, count in counts.items():
            aggregate_rows.append({
                "項目": field, "値": str(value), "件数": int(count),
                "割合（%）": round(int(count) / len(proposal) * 100, 1),
            })
    aggregate = pd.DataFrame(aggregate_rows)
    aggregate.to_csv(AI_AGGREGATE_CSV, index=False, encoding="utf-8-sig")

    def count(field: str, value: str) -> int:
        return int((proposal[field] == value).sum())

    report = f"""# Step11 AI暫定分類 結果レポート

## 状態

- 対象：開発用100件
- コードブック：`{workflow.CODEBOOK_VERSION}`
- 評価方式：ルール分類後、全100件についてコード値、必須項目、項目間矛盾、原文と根拠表示をCodexが二次検査
- 臨床確認：歯科医師による最終確認前
- 利用範囲：探索的な内部検討のみ
- 評価者間一致率・Cohenのκ：算出不可

## 暫定的に把握できた内容

- 起点機転は、転倒・つまずき{count('起点機転', '転倒・つまずき')}件、衝突・接触{count('起点機転', '衝突・接触')}件、転落・落下{count('起点機転', '転落・落下')}件、判定不能{count('起点機転', '判定不能')}件だった。
- 口腔・顔面への直接外力が原文に明記された事例は{count('口腔・顔面への直接外力', 'あり（明記）')}件だった。
- 前歯部の記載は{count('前歯部の記載', '記載あり')}件、口腔・口唇の記載は{count('口腔・口唇の記載', '記載あり')}件、顔面・顎の記載は{count('顔面・顎の記載', '記載あり')}件だった。
- 修正可能要因が明記された事例は{count('予防可能性', '修正可能要因の明記あり')}件だった。ただし、予防可能性は文脈依存性が高いため、最終確認前に強い結論を出さない。
- AI確認優先度は、最優先{count('AI確認優先度', '最優先')}件、要確認{count('AI確認優先度', '要確認')}件、通常{count('AI確認優先度', '通常')}件だった。

## 解釈上の注意

この結果は人手の正解データではない。特に、複数機転の順序、最終接触対象、直接外力および予防可能性は、原文の省略や曖昧さの影響を受ける。論文・学会発表の確定値として使用する場合は、一人研究者確認用Excelで全100件を確認し、`Step11-36_一人研究者確認済み_確定候補.csv`を書き出してから再集計する。
"""
    AI_REPORT.write_text(report, encoding="utf-8")

    record = {
        "schema_version": "Step11-AISecondCheck-1.0", "status": "AI_SECOND_CHECK_COMPLETE_HUMAN_NOT_CONFIRMED",
        "created_at_utc": utc_now(), "rows": len(proposal),
        "source_proposal_sha256": sha256(PROPOSAL_CSV), "reviewed_csv_sha256": sha256(AI_REVIEWED_CSV),
        "aggregate_csv_sha256": sha256(AI_AGGREGATE_CSV), "report_sha256": sha256(AI_REPORT),
        "script_sha256": sha256(Path(__file__).resolve()), "codebook_sha256": sha256(workflow.CODEBOOK_SOURCE),
        "warning": "歯科医師による最終確認前。2名独立評価、κ、Gold Standardを主張しない。",
    }
    write_json(AI_REVIEW_MANIFEST, record)
    secure_output()
    print(f"SOLO_AI_PROVISIONAL_OK: {len(proposal)}件 / レポート={AI_REPORT}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    subparsers.add_parser("prepare", help="AI仮分類と一人確認用Excelを新規生成")
    validate_parser = subparsers.add_parser("validate", help="確認途中の選択肢・矛盾・完了数を検査")
    validate_parser.add_argument("--require-complete", action="store_true")
    subparsers.add_parser("export-confirmed", help="全100件確認後に確定候補CSVを書き出す")
    subparsers.add_parser("export-ai-provisional", help="AI案を二次整合性検査し、暫定集計とレポートを作る")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    try:
        if args.command == "prepare":
            prepare()
        elif args.command == "validate":
            validate(args.require_complete)
        elif args.command == "export-confirmed":
            export_confirmed()
        elif args.command == "export-ai-provisional":
            export_ai_provisional()
    except SoloReviewError as error:
        raise SystemExit(f"STOP: {error}") from error


if __name__ == "__main__":
    main()
