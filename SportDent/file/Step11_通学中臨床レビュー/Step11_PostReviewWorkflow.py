"""Step11の回答提出後を、原本凍結→調停→合意値固定→最終評価へ進める。

このスクリプトは人の回答や承認を代行しない。前段成果物が未完了なら停止し、
AfterReview_DoNotOpenの自動回答・歯牙障害区分は一切読み込まない。
"""
from __future__ import annotations

import argparse
import fcntl
import hashlib
import json
import os
import secrets
import shutil
import tempfile
from datetime import datetime, timezone
from contextlib import contextmanager
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

import Step11_ReviewWorkflow as workflow


HERE = Path(__file__).resolve().parent
INTERNAL = workflow.INTERNAL
FROZEN_ROOT = INTERNAL / "FrozenOriginals"
AGREEMENT_ROOT = INTERNAL / "AgreementResults"
ADJUDICATION_ROOT = INTERNAL / "Adjudication"
CONSENSUS_ROOT = INTERNAL / "FrozenConsensus"
FINAL_RULES_ROOT = INTERNAL / "FinalRules"
FINAL_RULES_MANIFEST = FINAL_RULES_ROOT / "FINAL_RULES_FROZEN.json"
FINAL_CODEBOOK = FINAL_RULES_ROOT / "Step11-30_最終評価用コードブック.md"
FINAL_MACHINE_SPEC = FINAL_RULES_ROOT / "Step11-30C_最終評価用機械仕様.json"
FINAL_EXCELS_MANIFEST = INTERNAL / "FINAL_EXCELS_PREPARED.json"
PROCESS_LOCK = INTERNAL / ".step11-postreview.lock"

FROZEN_SCHEMA = "Step11-FrozenOriginals-1.0"
ADJUDICATION_SCHEMA = "Step11-Adjudication-1.0"
CONSENSUS_SCHEMA = "Step11-Consensus-1.0"
FINAL_RULES_SCHEMA = "Step11-FinalRules-1.0"
FINAL_EXCELS_SCHEMA = "Step11-FinalExcels-1.0"
AUDIT_CASES = 10
REASONS = ["原文の明示記載", "時系列", "分類境界", "項目間整合性", "情報不足・曖昧", "評価者入力ミス", "その他"]
ADJ_CONFIRM = "確認済み"
CONSENSUS_CONFIRM = "この内容で人手合意値を固定する"
FINAL_RULES_CONFIRM = "この内容で最終評価用規則を固定する"
FINAL_PREPARE_CONFIRM = "凍結規則で最終評価用Excelを生成する"
MODEL_RULE_STATUS = "未実装（人手一致度評価のみ）"
FINAL_CHECK_ITEMS = [
    "開発用調停と誤り分析を反映した",
    "選択肢・定義・矛盾規則を確定した",
    "最終評価中は変更しない",
    "旧自動暫定分類を最終性能評価へ流用しない",
    "選択肢と機械的矛盾規則は開発用版から変更しない",
]
CHANGE_LOG_COLUMNS = ["変更ID", "対象項目", "変更前", "変更後", "変更理由", "確認者", "確認日"]

FIELDS = [field for field in workflow.CHOICES if field != workflow.CONFIRM_COLUMN]
ADJ_COLUMNS = [
    "調停項目ID", workflow.ID_COLUMN, workflow.TEXT_COLUMN, "対象種別", "項目", "評価者A回答",
    "評価者B回答", "暫定合意値", "最終合意値", "理由区分", "理由メモ", "項目確認",
]


class PostReviewError(RuntimeError):
    """人手原本を変更せず後半工程を停止すべき問題。"""


@contextmanager
def process_lock():
    """同じ後半工程の同時実行を待たずに拒否する。"""
    INTERNAL.mkdir(parents=True, exist_ok=True)
    INTERNAL.chmod(0o700)
    with PROCESS_LOCK.open("a+", encoding="utf-8") as stream:
        PROCESS_LOCK.chmod(0o600)
        try:
            fcntl.flock(stream.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        except BlockingIOError as error:
            raise PostReviewError("別のStep11後半工程が実行中です") from error
        stream.seek(0)
        stream.truncate()
        stream.write(json.dumps({"pid": os.getpid(), "started_at_utc": utc_now()}, ensure_ascii=False))
        stream.flush()
        os.fsync(stream.fileno())
        try:
            yield
        finally:
            fcntl.flock(stream.fileno(), fcntl.LOCK_UN)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def json_write(path: Path, value: object) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    path.chmod(0o600)


def bytes_sha256(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def ordered_ids_sha256(ids: list[str]) -> str:
    return bytes_sha256(("\n".join(ids) + "\n").encode("utf-8"))


def phase_number(phase: str) -> str:
    return "08" if phase == "開発用" else "16"


def source_csv(phase: str, reviewer: str) -> Path:
    number = phase_number(phase)
    return INTERNAL / f"Step11-{number}{reviewer}_{phase}_評価者{reviewer}_盲検100件.csv"


def working_excel(phase: str, reviewer: str) -> Path:
    number = phase_number(phase)
    return INTERNAL / f"Step11-{number}{reviewer}_{phase}_評価者{reviewer}_盲検100件.xlsx"


def frozen_dir(phase: str) -> Path:
    return FROZEN_ROOT / phase


def _read_source(phase: str, reviewer: str) -> pd.DataFrame:
    path = source_csv(phase, reviewer)
    if not path.exists():
        raise PostReviewError(f"割付CSVがありません: {path}")
    frame = pd.read_csv(path, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    needed = [workflow.ID_COLUMN, workflow.TEXT_COLUMN]
    if not set(needed).issubset(frame.columns) or len(frame) != workflow.EXPECTED_ROWS:
        raise PostReviewError(f"割付CSVの構造または件数が不正です: {path}")
    frame = frame[needed].copy()
    if frame[workflow.ID_COLUMN].duplicated().any() or frame[workflow.TEXT_COLUMN].str.strip().eq("").any():
        raise PostReviewError(f"割付CSVのID重複または原文空欄があります: {path}")
    return frame


def _workbook_rows(path: Path) -> tuple[pd.DataFrame, dict[str, str], list[str]]:
    if not path.exists():
        raise PostReviewError(f"Excelがありません: {path}")
    book = load_workbook(path, data_only=False, keep_links=True)
    errors = workflow.workbook_has_forbidden_content(book)
    if workflow.INPUT_SHEET not in book.sheetnames or "メタデータ" not in book.sheetnames:
        book.close()
        raise PostReviewError(f"Excelの必須シートがありません: {path}")
    sheet = book[workflow.INPUT_SHEET]
    headers = [sheet.cell(workflow.HEADER_ROW, col).value for col in range(1, len(workflow.INPUT_COLUMNS) + 1)]
    if headers != workflow.INPUT_COLUMNS:
        errors.append("入力列または列順が不一致")
    rows: list[dict[str, str]] = []
    for row_number in range(workflow.FIRST_DATA_ROW, workflow.FIRST_DATA_ROW + workflow.EXPECTED_ROWS):
        rows.append({
            header: "" if sheet.cell(row_number, col).value is None else str(sheet.cell(row_number, col).value).strip()
            for col, header in enumerate(workflow.INPUT_COLUMNS, start=1)
        })
    meta = workflow.metadata(book)
    book.close()
    return pd.DataFrame(rows), meta, errors


def _generic_workbook_safety_errors(book) -> list[str]:
    errors: list[str] = []
    for sheet in book.worksheets:
        if sheet.sheet_state != "visible":
            errors.append(f"非表示シートがあります: {sheet.title}")
        for key, dimension in sheet.column_dimensions.items():
            if dimension.hidden:
                errors.append(f"非表示列があります: {sheet.title}!{key}")
        for key, dimension in sheet.row_dimensions.items():
            if dimension.hidden:
                errors.append(f"非表示行があります: {sheet.title}!{key}")
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    errors.append(f"数式セルがあります: {sheet.title}!{cell.coordinate}")
                if cell.hyperlink is not None:
                    errors.append(f"ハイパーリンクがあります: {sheet.title}!{cell.coordinate}")
                if cell.comment is not None:
                    errors.append(f"コメントオブジェクトがあります: {sheet.title}!{cell.coordinate}")
    if getattr(book, "_external_links", []):
        errors.append("外部リンクがあります")
    if getattr(book, "vba_archive", None) is not None:
        errors.append("VBAアーカイブがあります")
    return errors


def _workbook_fixed_fingerprint(
    path: Path,
    editable_sheet: str,
    editable_columns: set[int],
    first_editable_row: int,
) -> str:
    """回答セルの値だけを除外し、参照情報・構造・保護設定を固定する。"""
    book = load_workbook(path, data_only=False, keep_links=True)
    safety_errors = _generic_workbook_safety_errors(book)
    if safety_errors:
        book.close()
        raise PostReviewError("Excelの固定領域を記録できません: " + " / ".join(sorted(set(safety_errors))))
    payload: dict[str, object] = {
        "sheet_order": book.sheetnames,
        "defined_names": sorted(
            (name, str(getattr(definition, "attr_text", "")))
            for name, definition in book.defined_names.items()
        ),
        "sheets": {},
    }
    for sheet in book.worksheets:
        cells: list[list[object]] = []
        for row in sheet.iter_rows():
            for cell in row:
                editable = (
                    sheet.title == editable_sheet
                    and cell.row >= first_editable_row
                    and cell.column in editable_columns
                )
                fixed_value = "" if cell.value is None else str(cell.value)
                cells.append([
                    cell.coordinate,
                    "<EDITABLE>" if editable else fixed_value,
                    "<EDITABLE>" if editable else ("blank" if fixed_value == "" else str(cell.data_type)),
                    bool(cell.protection.locked),
                ])
        validations = sorted(
            (
                str(validation.type), str(validation.formula1), str(validation.formula2),
                str(validation.sqref), bool(validation.allow_blank), bool(validation.showErrorMessage),
            )
            for validation in sheet.data_validations.dataValidation
        )
        payload["sheets"][sheet.title] = {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "protected": bool(sheet.protection.sheet),
            "cells": cells,
            "validations": validations,
        }
    book.close()
    return bytes_sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8"))


def _safe_frozen_member(root: Path, filename: object, label: str) -> Path:
    name = str(filename or "")
    if not name or Path(name).name != name:
        raise PostReviewError(f"{label}のファイル名が不正です")
    path = root / name
    if path.is_symlink() or path.resolve().parent != root.resolve():
        raise PostReviewError(f"{label}が凍結領域外を参照しています")
    return path


def _require_mode(path: Path, expected: int, label: str) -> None:
    if path.stat().st_mode & 0o777 != expected:
        raise PostReviewError(f"{label}の権限が{expected:04o}ではありません: {path}")


def inspect_complete_workbook(
    phase: str,
    reviewer: str,
    path: Path | None = None,
    require_complete: bool = True,
) -> pd.DataFrame:
    path = path or working_excel(phase, reviewer)
    before = workflow.sha256(path)
    if phase == "開発用" and path == working_excel(phase, reviewer):
        frame, errors = workflow.inspect_workbook(path, reviewer, require_complete=require_complete)
    else:
        frame, meta, errors = _workbook_rows(path)
        source = _read_source(phase, reviewer)
        if not frame[[workflow.ID_COLUMN, workflow.TEXT_COLUMN]].equals(source):
            errors.append("割付CSVとレビューID／原文／順序が不一致")
        if meta.get("評価段階") != phase or meta.get("評価者") != reviewer:
            errors.append("評価段階または評価者メタデータが不一致")
        if phase == "最終評価用":
            final_manifest = _load_final_rules()
            prepared_manifest = _load_final_excel_manifest()
            if meta.get("コードブック版") != str(final_manifest.get("final_version", "")):
                errors.append("最終コードブック版が不一致")
            if meta.get("最終規則manifest SHA-256") != workflow.sha256(FINAL_RULES_MANIFEST):
                errors.append("最終規則manifestハッシュが不一致")
            expected_fingerprint = prepared_manifest.get("workbooks", {}).get(reviewer, {}).get(
                "fixed_workbook_fingerprint_sha256"
            )
            actual_fingerprint = _workbook_fixed_fingerprint(
                path,
                workflow.INPUT_SHEET,
                set(range(3, len(workflow.INPUT_COLUMNS) + 1)),
                workflow.FIRST_DATA_ROW,
            )
            if actual_fingerprint != expected_fingerprint:
                errors.append("最終評価Excelの固定領域・選択肢・判定基準または保護設定が変更されています")
        states: list[str] = []
        messages: list[str] = []
        for _, row in frame.iterrows():
            values = {column: str(row[column]).strip() for column in workflow.INPUT_COLUMNS}
            filled = sum(bool(values[column]) for column in workflow.REQUIRED_COLUMNS)
            problems = workflow.row_problems(values)
            if filled == 0 and not values[workflow.COMMENT_COLUMN]:
                state = "未入力"
            elif filled < len(workflow.REQUIRED_COLUMNS):
                state = "入力途中"
            elif problems:
                state = "要確認"
            else:
                state = "入力完了"
            states.append(state)
            messages.append(" / ".join(problems))
        frame["入力状態"] = states
        frame["検査メッセージ"] = messages
        if any(messages):
            errors.append("選択肢外入力または回答矛盾があります")
        if require_complete and any(state != "入力完了" for state in states):
            errors.append(f"未完了または要確認が{sum(state != '入力完了' for state in states)}件あります")
    if workflow.sha256(path) != before:
        errors.append("検査中にExcelが変更されました")
    if errors:
        raise PostReviewError(f"評価者{reviewer}Excelを凍結できません: {' / '.join(sorted(set(errors)))}")
    return frame


def validate_submissions(phase: str, require_complete: bool) -> None:
    results = {
        reviewer: inspect_complete_workbook(phase, reviewer, require_complete=require_complete)
        for reviewer in ["A", "B"]
    }
    if not results["A"][[workflow.ID_COLUMN, workflow.TEXT_COLUMN]].equals(
        results["B"][[workflow.ID_COLUMN, workflow.TEXT_COLUMN]]
    ):
        raise PostReviewError("評価者A/Bの標本・原文・順序が一致しません")
    counts = {
        reviewer: int(results[reviewer].get("入力状態", pd.Series(dtype=str)).eq("入力完了").sum())
        for reviewer in ["A", "B"]
    }
    kind = "完了検査" if require_complete else "構造・進捗検査"
    print(f"SUBMISSIONS_VALID: {phase} / {kind} / A完了{counts['A']}/100 / B完了{counts['B']}/100")


def _load_final_excel_manifest() -> dict[str, object]:
    if not FINAL_EXCELS_MANIFEST.exists() or FINAL_EXCELS_MANIFEST.is_symlink():
        raise PostReviewError("最終評価用Excel生成記録がありません")
    try:
        manifest = json.loads(FINAL_EXCELS_MANIFEST.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise PostReviewError("最終評価用Excel生成記録を読めません") from error
    if manifest.get("schema_version") != FINAL_EXCELS_SCHEMA or manifest.get("status") != "PREPARED":
        raise PostReviewError("最終評価用Excel生成記録の版または状態が不正です")
    if manifest.get("final_rules_manifest_sha256") != workflow.sha256(FINAL_RULES_MANIFEST):
        raise PostReviewError("最終評価用Excelと凍結規則が一致しません")
    for reviewer in ["A", "B"]:
        info = manifest.get("workbooks", {}).get(reviewer, {})
        if info.get("source_csv_sha256") != workflow.sha256(source_csv("最終評価用", reviewer)):
            raise PostReviewError(f"最終評価者{reviewer}の割付CSVが生成時から変更されています")
        if not info.get("fixed_workbook_fingerprint_sha256"):
            raise PostReviewError(f"最終評価者{reviewer}の固定領域fingerprintがありません")
    return manifest


def _frozen_manifest(phase: str) -> tuple[Path, dict[str, object]]:
    root = frozen_dir(phase)
    manifest_path = root / "manifest.json"
    if not manifest_path.exists():
        raise PostReviewError(f"凍結manifestがありません: {manifest_path}")
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise PostReviewError("凍結manifestを読めません") from error
    if manifest.get("schema_version") != FROZEN_SCHEMA or manifest.get("phase") != phase or manifest.get("status") != "FROZEN":
        raise PostReviewError("凍結manifestの版・段階・状態が不正です")
    if root.is_symlink() or manifest_path.is_symlink():
        raise PostReviewError("凍結原本にシンボリックリンクがあります")
    _require_mode(root, 0o500, "凍結原本ディレクトリ")
    _require_mode(manifest_path, 0o400, "凍結manifest")
    for reviewer in ["A", "B"]:
        info = manifest.get("workbooks", {}).get(reviewer, {})
        file = _safe_frozen_member(root, info.get("file"), f"凍結評価者{reviewer}原本")
        if not file.exists() or workflow.sha256(file) != info.get("sha256") or file.stat().st_size != info.get("size"):
            raise PostReviewError(f"凍結評価者{reviewer}原本がmanifestと一致しません")
        _require_mode(file, 0o400, f"凍結評価者{reviewer}原本")
    codebook = _safe_frozen_member(root, manifest.get("codebook_file"), "凍結コードブック")
    approval = _safe_frozen_member(root, manifest.get("approval_file"), "凍結承認・規則記録")
    if not codebook.exists() or workflow.sha256(codebook) != manifest.get("codebook_sha256"):
        raise PostReviewError("凍結コードブックがmanifestと一致しません")
    if not approval.exists() or workflow.sha256(approval) != manifest.get("approval_sha256"):
        raise PostReviewError("凍結承認・規則記録がmanifestと一致しません")
    _require_mode(codebook, 0o400, "凍結コードブック")
    _require_mode(approval, 0o400, "凍結承認・規則記録")
    return root, manifest


def freeze_submissions(phase: str, confirm_exact: str) -> None:
    expected_confirmation = f"{phase}A/B回答原本を凍結する"
    if confirm_exact != expected_confirmation:
        raise PostReviewError(f"--confirm-exactには『{expected_confirmation}』を指定してください")
    if phase == "開発用" and not workflow.APPROVAL.exists():
        raise PostReviewError("開発用コードブックが未承認です")
    if phase == "最終評価用" and not FINAL_RULES_MANIFEST.exists():
        raise PostReviewError("最終評価用規則が未凍結です")
    target = frozen_dir(phase)
    if target.exists():
        raise PostReviewError(f"凍結原本を上書きしません: {target}")
    frames = {reviewer: inspect_complete_workbook(phase, reviewer) for reviewer in ["A", "B"]}
    if not frames["A"][[workflow.ID_COLUMN, workflow.TEXT_COLUMN]].equals(frames["B"][[workflow.ID_COLUMN, workflow.TEXT_COLUMN]]):
        raise PostReviewError("評価者A/Bの標本・原文・順序が一致しません")
    source_hashes = {reviewer: workflow.sha256(working_excel(phase, reviewer)) for reviewer in ["A", "B"]}
    ids = frames["A"][workflow.ID_COLUMN].tolist()
    FROZEN_ROOT.mkdir(parents=True, exist_ok=True)
    FROZEN_ROOT.chmod(0o700)
    temporary = Path(tempfile.mkdtemp(prefix=f".freeze-{phase}-", dir=FROZEN_ROOT))
    temporary.chmod(0o700)
    try:
        workbook_info: dict[str, dict[str, object]] = {}
        for reviewer in ["A", "B"]:
            source = working_excel(phase, reviewer)
            copy = temporary / source.name
            shutil.copy2(source, copy)
            if workflow.sha256(source) != source_hashes[reviewer] or workflow.sha256(copy) != source_hashes[reviewer]:
                raise PostReviewError(f"評価者{reviewer}Excelが検査後またはコピー中に変更されました")
            workbook_info[reviewer] = {
                "file": copy.name, "sha256": workflow.sha256(copy), "size": copy.stat().st_size,
                "rows": len(frames[reviewer]), "source_csv_sha256": workflow.sha256(source_csv(phase, reviewer)),
            }
        if phase == "開発用":
            codebook_source = workflow.CODEBOOK_SOURCE
            approval_source = workflow.APPROVAL
            codebook_version = workflow.CODEBOOK_VERSION
            final_manifest_hash = ""
        else:
            rules = json.loads(FINAL_RULES_MANIFEST.read_text(encoding="utf-8"))
            codebook_source = FINAL_CODEBOOK
            approval_source = FINAL_RULES_MANIFEST
            codebook_version = str(rules.get("final_version", ""))
            final_manifest_hash = workflow.sha256(FINAL_RULES_MANIFEST)
        codebook_copy = temporary / codebook_source.name
        approval_copy = temporary / approval_source.name
        shutil.copy2(codebook_source, codebook_copy)
        shutil.copy2(approval_source, approval_copy)
        manifest = {
            "schema_version": FROZEN_SCHEMA, "phase": phase, "status": "FROZEN",
            "frozen_at_utc": utc_now(), "codebook_version": codebook_version,
            "codebook_file": codebook_copy.name, "codebook_sha256": workflow.sha256(codebook_copy),
            "approval_file": approval_copy.name, "approval_sha256": workflow.sha256(approval_copy),
            "final_rules_manifest_sha256": final_manifest_hash,
            "postreview_workflow_sha256": workflow.sha256(Path(__file__).resolve()),
            "review_workflow_sha256": workflow.sha256(Path(workflow.__file__).resolve()),
            "ordered_review_ids_sha256": ordered_ids_sha256(ids), "workbooks": workbook_info,
            "auto_answer_inputs_used": False,
        }
        json_write(temporary / "manifest.json", manifest)
        for file in temporary.iterdir():
            file.chmod(0o400)
        os.replace(temporary, target)
        target.chmod(0o500)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise
    _frozen_manifest(phase)
    print(f"FREEZE_OK: {phase} / A/B各100件 / 凍結原本={target}")


def _load_frozen_frames(phase: str) -> tuple[pd.DataFrame, pd.DataFrame, Path, dict[str, object]]:
    root, manifest = _frozen_manifest(phase)
    frames: dict[str, pd.DataFrame] = {}
    for reviewer in ["A", "B"]:
        info = manifest["workbooks"][reviewer]
        frame, _, errors = _workbook_rows(root / str(info["file"]))
        if errors or len(frame) != workflow.EXPECTED_ROWS:
            raise PostReviewError(f"凍結評価者{reviewer}Excelの構造が不正です: {' / '.join(errors)}")
        frames[reviewer] = frame
    if not frames["A"][[workflow.ID_COLUMN, workflow.TEXT_COLUMN]].equals(
        frames["B"][[workflow.ID_COLUMN, workflow.TEXT_COLUMN]]
    ):
        raise PostReviewError("凍結A/Bの標本・原文・順序が一致しません")
    ids = frames["A"][workflow.ID_COLUMN].tolist()
    if ordered_ids_sha256(ids) != manifest.get("ordered_review_ids_sha256"):
        raise PostReviewError("凍結A/Bの順序付きレビューIDハッシュが不一致です")
    return frames["A"], frames["B"], root, manifest


def _adjudication_dir(phase: str) -> Path:
    return ADJUDICATION_ROOT / phase


def _adjudication_paths(phase: str) -> tuple[Path, Path]:
    root = _adjudication_dir(phase)
    return root / f"Step11-28_{phase}100件_第三者調停票.xlsx", root / "Step11-28A_調停対象・無作為監査抽出manifest.json"


def _expected_adjudication_rows(phase: str, salt: str, audit_ids: list[str] | None = None) -> tuple[pd.DataFrame, list[str]]:
    first, second, _, _ = _load_frozen_frames(phase)
    if audit_ids is None:
        ranked = sorted(
            first[workflow.ID_COLUMN].tolist(),
            key=lambda review_id: hashlib.sha256(f"{salt}|{review_id}".encode("utf-8")).hexdigest(),
        )
        audit_ids = ranked[:AUDIT_CASES]
    audit_set = set(audit_ids)
    rows: list[dict[str, str]] = []
    for index in range(len(first)):
        review_id = str(first.iloc[index][workflow.ID_COLUMN])
        text = str(first.iloc[index][workflow.TEXT_COLUMN])
        for field_index, field in enumerate(FIELDS, start=1):
            a_value = str(first.iloc[index][field])
            b_value = str(second.iloc[index][field])
            if a_value != b_value:
                target_type = "不一致"
                provisional = ""
            elif review_id in audit_set:
                target_type = "一致無作為監査"
                provisional = a_value
            else:
                continue
            rows.append({
                "調停項目ID": f"{review_id}-{field_index:02d}", workflow.ID_COLUMN: review_id,
                workflow.TEXT_COLUMN: text, "対象種別": target_type, "項目": field,
                "評価者A回答": a_value, "評価者B回答": b_value, "暫定合意値": provisional,
                "最終合意値": "", "理由区分": "", "理由メモ": "", "項目確認": "",
            })
    return pd.DataFrame(rows, columns=ADJ_COLUMNS), audit_ids


def _build_adjudication_workbook(
    rows: pd.DataFrame,
    phase: str,
    destination: Path,
    codebook: Path,
    codebook_version: str,
) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = "調停入力"
    sheet.append(ADJ_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True)
    for row_index, (_, source) in enumerate(rows.iterrows(), start=2):
        for col_index, column in enumerate(ADJ_COLUMNS, start=1):
            cell = sheet.cell(row_index, col_index, str(source[column]))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.protection = Protection(locked=column not in {"最終合意値", "理由区分", "理由メモ", "項目確認"})
            if not cell.protection.locked:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
    choices = book.create_sheet("選択肢")
    final_names: dict[str, str] = {}
    for index, field in enumerate(FIELDS, start=1):
        choices.cell(1, index, field)
        for row, value in enumerate(workflow.CHOICES[field], start=2):
            choices.cell(row, index, value)
        name = f"adj_choices_{index:02d}"
        column = get_column_letter(index)
        book.defined_names.add(DefinedName(name, attr_text=f"'選択肢'!${column}$2:${column}${len(workflow.CHOICES[field]) + 1}"))
        final_names[field] = name
    reason_column = len(FIELDS) + 1
    choices.cell(1, reason_column, "理由区分")
    for row, value in enumerate(REASONS, start=2):
        choices.cell(row, reason_column, value)
    reason_letter = get_column_letter(reason_column)
    book.defined_names.add(DefinedName("adj_reasons", attr_text=f"'選択肢'!${reason_letter}$2:${reason_letter}${len(REASONS) + 1}"))
    choices.cell(1, reason_column + 1, "項目確認")
    choices.cell(2, reason_column + 1, ADJ_CONFIRM)
    confirm_letter = get_column_letter(reason_column + 1)
    book.defined_names.add(DefinedName("adj_confirm", attr_text=f"'選択肢'!${confirm_letter}$2:${confirm_letter}$2"))
    final_col = ADJ_COLUMNS.index("最終合意値") + 1
    reason_col = ADJ_COLUMNS.index("理由区分") + 1
    confirm_col = ADJ_COLUMNS.index("項目確認") + 1
    for row_index, (_, source) in enumerate(rows.iterrows(), start=2):
        validation = DataValidation(type="list", formula1=f"={final_names[str(source['項目'])]}")
        validation.showErrorMessage = True
        sheet.add_data_validation(validation)
        validation.add(sheet.cell(row_index, final_col))
    reason_validation = DataValidation(type="list", formula1="=adj_reasons")
    confirm_validation = DataValidation(type="list", formula1="=adj_confirm")
    sheet.add_data_validation(reason_validation)
    sheet.add_data_validation(confirm_validation)
    if len(rows):
        reason_validation.add(f"{get_column_letter(reason_col)}2:{get_column_letter(reason_col)}{len(rows) + 1}")
        confirm_validation.add(f"{get_column_letter(confirm_col)}2:{get_column_letter(confirm_col)}{len(rows) + 1}")
    all_answers = book.create_sheet("A_B全回答")
    first, second, _, _ = _load_frozen_frames(phase)
    reference_columns = [workflow.ID_COLUMN, workflow.TEXT_COLUMN]
    for field in FIELDS:
        reference_columns.extend([f"A:{field}", f"B:{field}"])
    reference_columns.extend(["A:コメント", "B:コメント"])
    all_answers.append(reference_columns)
    for index in range(len(first)):
        values = [first.iloc[index][workflow.ID_COLUMN], first.iloc[index][workflow.TEXT_COLUMN]]
        for field in FIELDS:
            values.extend([first.iloc[index][field], second.iloc[index][field]])
        values.extend([first.iloc[index][workflow.COMMENT_COLUMN], second.iloc[index][workflow.COMMENT_COLUMN]])
        all_answers.append(values)
    codebook_sheet = book.create_sheet("判定基準")
    for row, line in enumerate(codebook.read_text(encoding="utf-8").splitlines(), start=1):
        codebook_sheet.cell(row, 1, line).alignment = Alignment(wrap_text=True)
    metadata_sheet = book.create_sheet("メタデータ")
    metadata_rows = [
        ("schema_version", ADJUDICATION_SCHEMA), ("評価段階", phase),
        ("生成日時UTC", utc_now()), ("コードブック版", codebook_version),
        ("自動回答・歯牙障害区分の入力使用", "なし"), ("調停対象件数", str(len(rows))),
    ]
    for key, value in metadata_rows:
        metadata_sheet.append([key, value])
    for protected_sheet in book.worksheets:
        protected_sheet.protection.sheet = True
        protected_sheet.protection.password = "Step11Adjudication"
    sheet.freeze_panes = "I2"
    widths = [22, 18, 80, 18, 30, 24, 24, 24, 24, 24, 45, 16]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    book.save(destination)


def prepare_adjudication(phase: str) -> None:
    if not (AGREEMENT_ROOT / phase).exists():
        raise PostReviewError("調停前一致度解析が未完了です")
    frozen, frozen_manifest = _frozen_manifest(phase)
    target, manifest_path = _adjudication_paths(phase)
    if target.exists() or manifest_path.exists() or target.parent.exists():
        raise PostReviewError(f"既存の調停工程を上書きしません: {target.parent}")
    salt = secrets.token_hex(32)
    rows, audit_ids = _expected_adjudication_rows(phase, salt)
    if rows.empty:
        raise PostReviewError("調停・監査対象が0件です")
    ADJUDICATION_ROOT.mkdir(parents=True, exist_ok=True)
    ADJUDICATION_ROOT.chmod(0o700)
    temporary = Path(tempfile.mkdtemp(prefix=f".adjudication-{phase}-", dir=ADJUDICATION_ROOT))
    temporary.chmod(0o700)
    try:
        workbook_path = temporary / target.name
        codebook = frozen / str(frozen_manifest["codebook_file"])
        _build_adjudication_workbook(
            rows, phase, workbook_path, codebook, str(frozen_manifest["codebook_version"])
        )
        workbook_path.chmod(0o600)
        fixed_fingerprint = _workbook_fixed_fingerprint(
            workbook_path,
            "調停入力",
            {ADJ_COLUMNS.index(column) + 1 for column in {"最終合意値", "理由区分", "理由メモ", "項目確認"}},
            2,
        )
        manifest = {
            "schema_version": ADJUDICATION_SCHEMA, "phase": phase, "status": "PREPARED",
            "created_at_utc": utc_now(), "frozen_manifest_sha256": workflow.sha256(frozen / "manifest.json"),
            "agreement_result_directory": str((AGREEMENT_ROOT / phase).relative_to(INTERNAL)),
            "agreement_files": {
                file.name: workflow.sha256(file) for file in sorted((AGREEMENT_ROOT / phase).iterdir()) if file.is_file()
            },
            "audit_selection_method": "SHA-256(salt|レビューID)昇順上位10件",
            "audit_salt": salt, "audit_review_ids": audit_ids, "audit_case_count": len(audit_ids),
            "target_row_count": len(rows), "workbook": workbook_path.name,
            "initial_workbook_sha256": workflow.sha256(workbook_path),
            "fixed_workbook_fingerprint_sha256": fixed_fingerprint,
            "codebook_version": str(frozen_manifest["codebook_version"]),
            "postreview_workflow_sha256": workflow.sha256(Path(__file__).resolve()),
            "auto_answer_inputs_used": False,
        }
        json_write(temporary / manifest_path.name, manifest)
        os.replace(temporary, target.parent)
        target.parent.chmod(0o700)
        for file in target.parent.iterdir():
            file.chmod(0o600)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise
    print(f"ADJUDICATION_PREPARED: {phase} / 対象{len(rows)}項目 / 一致監査{len(audit_ids)}事例")


def inspect_adjudication(phase: str, require_complete: bool) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, list[str]]:
    workbook_path, manifest_path = _adjudication_paths(phase)
    if not workbook_path.exists() or not manifest_path.exists():
        raise PostReviewError("調停票または調停manifestがありません")
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise PostReviewError("調停manifestを読めません") from error
    if (
        manifest.get("schema_version") != ADJUDICATION_SCHEMA
        or manifest.get("phase") != phase
        or manifest.get("status") != "PREPARED"
    ):
        raise PostReviewError("調停manifestの版または段階が不一致です")
    frozen, _ = _frozen_manifest(phase)
    if manifest.get("frozen_manifest_sha256") != workflow.sha256(frozen / "manifest.json"):
        raise PostReviewError("調停manifestと凍結原本manifestが一致しません")
    agreement = AGREEMENT_ROOT / phase
    for filename, expected_hash in manifest.get("agreement_files", {}).items():
        file = agreement / filename
        if not file.exists() or workflow.sha256(file) != expected_hash:
            raise PostReviewError("調停前一致度結果が調停票生成時から変更されています")
    expected, expected_audit_ids = _expected_adjudication_rows(
        phase, str(manifest.get("audit_salt", "")), list(manifest.get("audit_review_ids", [])),
    )
    if expected_audit_ids != list(manifest.get("audit_review_ids", [])) or len(expected) != manifest.get("target_row_count"):
        raise PostReviewError("調停対象または一致監査標本がmanifestと一致しません")
    book = load_workbook(workbook_path, data_only=False, keep_links=True)
    errors = _generic_workbook_safety_errors(book)
    expected_sheets = {"調停入力", "選択肢", "A_B全回答", "判定基準", "メタデータ"}
    if set(book.sheetnames) != expected_sheets:
        book.close()
        raise PostReviewError("調停Excelのシート構成が不一致")
    sheet = book["調停入力"]
    headers = [sheet.cell(1, col).value for col in range(1, len(ADJ_COLUMNS) + 1)]
    if headers != ADJ_COLUMNS or sheet.max_column != len(ADJ_COLUMNS):
        errors.append("調停Excelの列または列順が不一致")
    rows: list[dict[str, str]] = []
    for row_number in range(2, 2 + len(expected)):
        rows.append({
            column: "" if sheet.cell(row_number, col).value is None else str(sheet.cell(row_number, col).value).strip()
            for col, column in enumerate(ADJ_COLUMNS, start=1)
        })
    actual = pd.DataFrame(rows, columns=ADJ_COLUMNS)
    extra = [
        sheet.cell(row, col).value
        for row in range(2 + len(expected), sheet.max_row + 1)
        for col in range(1, sheet.max_column + 1)
        if sheet.cell(row, col).value not in {None, ""}
    ]
    if extra:
        errors.append("調停Excelに余剰行があります")
    fixed_columns = ADJ_COLUMNS[:8]
    if not actual[fixed_columns].equals(expected[fixed_columns]):
        errors.append("調停対象、原文、A/B回答または暫定値が変更されています")
    book.close()
    if _workbook_fixed_fingerprint(
        workbook_path,
        "調停入力",
        {ADJ_COLUMNS.index(column) + 1 for column in {"最終合意値", "理由区分", "理由メモ", "項目確認"}},
        2,
    ) != manifest.get("fixed_workbook_fingerprint_sha256"):
        errors.append("調停Excelの固定領域・補助シート・選択肢または保護設定が変更されています")
    row_errors: list[str] = []
    for index, row in actual.iterrows():
        final = str(row["最終合意値"])
        reason = str(row["理由区分"])
        memo = str(row["理由メモ"])
        confirmation = str(row["項目確認"])
        field = str(row["項目"])
        if final and final not in workflow.CHOICES[field]:
            row_errors.append(f"{index + 2}行: 最終合意値が選択肢外")
        if reason and reason not in REASONS:
            row_errors.append(f"{index + 2}行: 理由区分が選択肢外")
        if confirmation and confirmation != ADJ_CONFIRM:
            row_errors.append(f"{index + 2}行: 項目確認が選択肢外")
        if require_complete:
            if not final or not reason or confirmation != ADJ_CONFIRM:
                row_errors.append(f"{index + 2}行: 最終合意値・理由・確認が未完了")
            changed_audit = row["対象種別"] == "一致無作為監査" and final != row["暫定合意値"]
            third_value = row["対象種別"] == "不一致" and final not in {row["評価者A回答"], row["評価者B回答"]}
            if (changed_audit or third_value or reason == "その他") and not memo:
                row_errors.append(f"{index + 2}行: 変更・第三値・その他理由には理由メモ必須")
    errors.extend(row_errors)
    first, second, _, _ = _load_frozen_frames(phase)
    consensus_rows: list[dict[str, str]] = []
    trace_rows: list[dict[str, str]] = []
    target_lookup = {(row[workflow.ID_COLUMN], row["項目"]): row for _, row in actual.iterrows()}
    for index in range(len(first)):
        review_id = str(first.iloc[index][workflow.ID_COLUMN])
        consensus: dict[str, str] = {workflow.ID_COLUMN: review_id}
        memos = [str(first.iloc[index][workflow.COMMENT_COLUMN]), str(second.iloc[index][workflow.COMMENT_COLUMN])]
        mismatch_count = audit_count = audit_changes = third_count = 0
        for field in FIELDS:
            a_value = str(first.iloc[index][field])
            b_value = str(second.iloc[index][field])
            target = target_lookup.get((review_id, field))
            if target is None:
                if a_value != b_value:
                    errors.append(f"{review_id}/{field}: 不一致項目が調停対象にありません")
                    value = ""
                else:
                    value = a_value
                formation = "A/B一致・非監査"
            else:
                value = str(target["最終合意値"])
                formation = str(target["対象種別"])
                if target["対象種別"] == "不一致":
                    mismatch_count += 1
                    if value not in {a_value, b_value, ""}:
                        third_count += 1
                else:
                    audit_count += 1
                    if value and value != a_value:
                        audit_changes += 1
                if str(target["理由メモ"]):
                    memos.append(str(target["理由メモ"]))
                trace_rows.append({
                    workflow.ID_COLUMN: review_id, "項目": field, "対象種別": str(target["対象種別"]),
                    "評価者A回答": a_value, "評価者B回答": b_value, "最終合意値": value,
                    "理由区分": str(target["理由区分"]), "理由メモ": str(target["理由メモ"]),
                })
            consensus[field] = value
        consensus[workflow.COMMENT_COLUMN] = " / ".join(value for value in memos if value)
        consensus["合意形成区分"] = "調停あり" if mismatch_count else "A/B一致"
        consensus["不一致項目数"] = str(mismatch_count)
        consensus["監査一致項目数"] = str(audit_count)
        consensus["一致監査変更項目数"] = str(audit_changes)
        consensus["第三値採用項目数"] = str(third_count)
        consensus_rows.append(consensus)
        if require_complete:
            values = {
                **{field: consensus[field] for field in FIELDS},
                workflow.COMMENT_COLUMN: consensus[workflow.COMMENT_COLUMN],
                workflow.CONFIRM_COLUMN: "確認済み",
            }
            problems = workflow.row_problems(values)
            if problems:
                errors.append(f"{review_id}: 合意値の項目間矛盾: {' / '.join(problems)}")
    if errors:
        raise PostReviewError("\n".join(sorted(set(errors))))
    return actual, pd.DataFrame(consensus_rows), pd.DataFrame(trace_rows), list(manifest.get("audit_review_ids", []))


def validate_adjudication(phase: str, require_complete: bool) -> None:
    actual, consensus, _, audit_ids = inspect_adjudication(phase, require_complete=require_complete)
    completed = int((
        actual["最終合意値"].ne("") & actual["理由区分"].ne("") & actual["項目確認"].eq(ADJ_CONFIRM)
    ).sum())
    status = "完了検査" if require_complete else "構造・進捗検査"
    print(f"ADJUDICATION_VALID: {phase} / {status} / 完了{completed}/{len(actual)}項目 / 一致監査{len(audit_ids)}事例 / 合意対象{len(consensus)}件")


def freeze_consensus(phase: str, confirmed_by: str, confirm_exact: str) -> None:
    if confirm_exact != CONSENSUS_CONFIRM:
        raise PostReviewError(f"--confirm-exactには『{CONSENSUS_CONFIRM}』を指定してください")
    if not confirmed_by.strip():
        raise PostReviewError("合意値の確定責任者を指定してください")
    actual, consensus, trace, audit_ids = inspect_adjudication(phase, require_complete=True)
    frozen, frozen_manifest = _frozen_manifest(phase)
    codebook_version = str(frozen_manifest["codebook_version"])
    target = CONSENSUS_ROOT / phase
    if target.exists():
        raise PostReviewError(f"凍結合意値を上書きしません: {target}")
    CONSENSUS_ROOT.mkdir(parents=True, exist_ok=True)
    CONSENSUS_ROOT.chmod(0o700)
    temporary = Path(tempfile.mkdtemp(prefix=f".consensus-{phase}-", dir=CONSENSUS_ROOT))
    temporary.chmod(0o700)
    try:
        consensus_output = consensus.drop(columns=[workflow.COMMENT_COLUMN]).copy()
        consensus_output["コードブック版"] = codebook_version
        consensus_output["合意状態"] = "固定済み"
        consensus_file = temporary / f"Step11-29_{phase}100件_人手合意値_固定.csv"
        trace_file = temporary / f"Step11-29A_{phase}100件_調停監査証跡.csv"
        consensus_output.to_csv(consensus_file, index=False, encoding="utf-8-sig")
        trace.to_csv(trace_file, index=False, encoding="utf-8-sig")
        adjudication_file, adjudication_manifest = _adjudication_paths(phase)
        frozen_adjudication = temporary / adjudication_file.name
        shutil.copy2(adjudication_file, frozen_adjudication)
        manifest = {
            "schema_version": CONSENSUS_SCHEMA, "phase": phase, "status": "CONSENSUS_FROZEN",
            "frozen_at_utc": utc_now(), "confirmed_by": confirmed_by.strip(),
            "adjudication_workbook_sha256": workflow.sha256(adjudication_file),
            "adjudication_manifest_sha256": workflow.sha256(adjudication_manifest),
            "consensus_file": consensus_file.name, "consensus_sha256": workflow.sha256(consensus_file),
            "audit_file": trace_file.name, "audit_sha256": workflow.sha256(trace_file),
            "frozen_adjudication_file": frozen_adjudication.name,
            "audit_review_ids": audit_ids, "review_count": len(consensus), "target_item_count": len(actual),
            "codebook_version": codebook_version,
            "frozen_submission_manifest_sha256": workflow.sha256(frozen / "manifest.json"),
            "postreview_workflow_sha256": workflow.sha256(Path(__file__).resolve()),
            "auto_answer_inputs_used": False,
        }
        json_write(temporary / "manifest.json", manifest)
        for file in temporary.iterdir():
            file.chmod(0o400)
        os.replace(temporary, target)
        target.chmod(0o500)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise
    print(f"CONSENSUS_FROZEN: {phase} / {len(consensus)}件 / 自動回答未参照")


def _consensus_manifest(phase: str) -> tuple[Path, dict[str, object]]:
    root = CONSENSUS_ROOT / phase
    manifest_path = root / "manifest.json"
    if not manifest_path.exists():
        raise PostReviewError(f"{phase}人手合意値が未凍結です")
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise PostReviewError(f"{phase}合意manifestを読めません") from error
    if (
        manifest.get("schema_version") != CONSENSUS_SCHEMA
        or manifest.get("phase") != phase
        or manifest.get("status") != "CONSENSUS_FROZEN"
        or manifest.get("review_count") != workflow.EXPECTED_ROWS
    ):
        raise PostReviewError(f"{phase}合意manifestの版・状態・件数が不正です")
    frozen, frozen_manifest = _frozen_manifest(phase)
    if manifest.get("frozen_submission_manifest_sha256") != workflow.sha256(frozen / "manifest.json"):
        raise PostReviewError(f"{phase}合意値と凍結A/B原本が一致しません")
    if manifest.get("codebook_version") != frozen_manifest.get("codebook_version"):
        raise PostReviewError(f"{phase}合意値のコードブック版が凍結原本と一致しません")
    for file_key, hash_key, label in [
        ("consensus_file", "consensus_sha256", "人手合意値"),
        ("audit_file", "audit_sha256", "調停監査証跡"),
        ("frozen_adjudication_file", "adjudication_workbook_sha256", "凍結調停票"),
    ]:
        file = _safe_frozen_member(root, manifest.get(file_key), label)
        if not file.exists() or workflow.sha256(file) != manifest.get(hash_key):
            raise PostReviewError(f"{phase}{label}がmanifestと一致しません")
        _require_mode(file, 0o400, f"{phase}{label}")
    if root.is_symlink() or manifest_path.is_symlink():
        raise PostReviewError(f"{phase}合意領域にシンボリックリンクがあります")
    _require_mode(root, 0o500, f"{phase}合意ディレクトリ")
    _require_mode(manifest_path, 0o400, f"{phase}合意manifest")
    return root, manifest


def prepare_final_rules_draft() -> None:
    _consensus_manifest("開発用")
    frozen, frozen_manifest = _frozen_manifest("開発用")
    draft_root = INTERNAL / "FinalRulesDraft"
    if draft_root.exists():
        raise PostReviewError(f"既存の最終規則案を上書きしません: {draft_root}")
    temporary = Path(tempfile.mkdtemp(prefix=".final-rules-draft-", dir=INTERNAL))
    temporary.chmod(0o700)
    try:
        draft_codebook = temporary / "Step11-30_最終評価用コードブック案.md"
        frozen_codebook = _safe_frozen_member(frozen, frozen_manifest.get("codebook_file"), "開発用凍結コードブック")
        source_text = frozen_codebook.read_text(encoding="utf-8")
        draft_codebook.write_text(
            "<!-- 開発用調停後の変更だけを履歴と対応させて編集する。最終凍結後は変更禁止。 -->\n"
            + source_text.replace(
                "固定コードブック候補（研究者確認待ち）",
                "最終評価用コードブック案（凍結前）",
            ),
            encoding="utf-8",
        )
        changes = pd.DataFrame(columns=CHANGE_LOG_COLUMNS)
        changes.to_csv(temporary / "Step11-30A_コードブック変更履歴.csv", index=False, encoding="utf-8-sig")
        checklist = pd.DataFrame([{"確認項目": item, "確認値": ""} for item in FINAL_CHECK_ITEMS])
        checklist.to_csv(temporary / "Step11-30B_最終規則凍結チェック.csv", index=False, encoding="utf-8-sig")
        for file in temporary.iterdir():
            file.chmod(0o600)
        os.replace(temporary, draft_root)
        draft_root.chmod(0o700)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise
    print(f"FINAL_RULES_DRAFT_OK: 編集・確認用={draft_root} / まだ凍結していません")


def freeze_final_rules(
    final_version: str,
    confirmed_by: str,
    change_summary: str,
    model_rule_status: str,
    confirm_exact: str,
) -> None:
    if confirm_exact != FINAL_RULES_CONFIRM:
        raise PostReviewError(f"--confirm-exactには『{FINAL_RULES_CONFIRM}』を指定してください")
    if not all(value.strip() for value in [final_version, confirmed_by, change_summary, model_rule_status]):
        raise PostReviewError("最終版、確認者、変更要約、自動分類規則の状態をすべて指定してください")
    if model_rule_status.strip() != MODEL_RULE_STATUS:
        raise PostReviewError(
            "現段階では互換する自動分類器を凍結していません。"
            f"--model-rule-statusには『{MODEL_RULE_STATUS}』を指定してください"
        )
    if final_version == workflow.CODEBOOK_VERSION or "final" not in final_version.lower():
        raise PostReviewError("最終版番号は開発用版と分け、finalを含めてください")
    _, development_consensus_manifest = _consensus_manifest("開発用")
    development_consensus = CONSENSUS_ROOT / "開発用" / "manifest.json"
    draft_root = INTERNAL / "FinalRulesDraft"
    draft_codebook = draft_root / "Step11-30_最終評価用コードブック案.md"
    change_log = draft_root / "Step11-30A_コードブック変更履歴.csv"
    checklist = draft_root / "Step11-30B_最終規則凍結チェック.csv"
    for path in [development_consensus, draft_codebook, change_log, checklist]:
        if not path.exists():
            raise PostReviewError(f"最終規則凍結の必須資料がありません: {path}")
    try:
        check = pd.read_csv(checklist, encoding="utf-8-sig", dtype=str, keep_default_na=False)
        changes = pd.read_csv(change_log, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    except (OSError, pd.errors.ParserError) as error:
        raise PostReviewError("最終規則の確認表または変更履歴を読めません") from error
    if list(check.columns) != ["確認項目", "確認値"] or check["確認項目"].tolist() != FINAL_CHECK_ITEMS:
        raise PostReviewError("最終規則凍結チェックの項目・順序・列が不正です")
    if not check["確認値"].eq("確認済み").all():
        raise PostReviewError("最終規則凍結チェックが全項目『確認済み』ではありません")
    if list(changes.columns) != CHANGE_LOG_COLUMNS:
        raise PostReviewError("コードブック変更履歴の列が不正です")
    if draft_codebook.stat().st_size == 0:
        raise PostReviewError("最終評価用コードブック案が空です")
    if FINAL_RULES_ROOT.exists():
        raise PostReviewError(f"最終規則を上書きしません: {FINAL_RULES_ROOT}")
    temporary = Path(tempfile.mkdtemp(prefix=".final-rules-", dir=INTERNAL))
    temporary.chmod(0o700)
    try:
        codebook_copy = temporary / FINAL_CODEBOOK.name
        shutil.copy2(draft_codebook, codebook_copy)
        change_copy = temporary / change_log.name
        checklist_copy = temporary / checklist.name
        shutil.copy2(change_log, change_copy)
        shutil.copy2(checklist, checklist_copy)
        machine_spec = temporary / FINAL_MACHINE_SPEC.name
        json_write(machine_spec, {
            "schema_version": "Step11-MachineReadableCodebook-1.0",
            "final_version": final_version.strip(),
            "input_columns": workflow.INPUT_COLUMNS,
            "required_columns": workflow.REQUIRED_COLUMNS,
            "site_columns": workflow.SITE_COLUMNS,
            "choices": workflow.CHOICES,
            "definitions": workflow.DEFINITIONS,
            "validation_implementation_sha256": workflow.sha256(Path(workflow.__file__).resolve()),
            "schema_change_from_development": False,
            "automatic_classifier_status": MODEL_RULE_STATUS,
        })
        manifest = {
            "schema_version": FINAL_RULES_SCHEMA, "status": "FINAL_RULES_FROZEN",
            "final_version": final_version.strip(), "frozen_at_utc": utc_now(),
            "confirmed_by": confirmed_by.strip(), "change_summary": change_summary.strip(),
            "model_rule_status": model_rule_status.strip(),
            "codebook_file": codebook_copy.name, "codebook_sha256": workflow.sha256(codebook_copy),
            "change_log_file": change_copy.name, "change_log_sha256": workflow.sha256(change_copy),
            "checklist_file": checklist_copy.name, "checklist_sha256": workflow.sha256(checklist_copy),
            "machine_spec_file": machine_spec.name, "machine_spec_sha256": workflow.sha256(machine_spec),
            "development_consensus_manifest_sha256": workflow.sha256(development_consensus),
            "development_codebook_version": development_consensus_manifest["codebook_version"],
            "postreview_workflow_sha256": workflow.sha256(Path(__file__).resolve()),
            "old_automatic_predictions_reused": False,
        }
        json_write(temporary / FINAL_RULES_MANIFEST.name, manifest)
        for file in temporary.iterdir():
            file.chmod(0o400)
        os.replace(temporary, FINAL_RULES_ROOT)
        FINAL_RULES_ROOT.chmod(0o500)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise
    print(f"FINAL_RULES_FROZEN: {final_version} / 最終評価中は変更禁止")


def _load_final_rules() -> dict[str, object]:
    if not FINAL_RULES_MANIFEST.exists() or not FINAL_CODEBOOK.exists():
        raise PostReviewError("最終規則が未凍結です")
    try:
        manifest = json.loads(FINAL_RULES_MANIFEST.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise PostReviewError("最終規則manifestを読めません") from error
    if manifest.get("schema_version") != FINAL_RULES_SCHEMA or manifest.get("status") != "FINAL_RULES_FROZEN":
        raise PostReviewError("最終規則manifestの版または状態が不正です")
    if FINAL_RULES_ROOT.is_symlink() or FINAL_RULES_MANIFEST.is_symlink():
        raise PostReviewError("最終規則領域にシンボリックリンクがあります")
    _require_mode(FINAL_RULES_ROOT, 0o500, "最終規則ディレクトリ")
    _require_mode(FINAL_RULES_MANIFEST, 0o400, "最終規則manifest")
    for file_key, hash_key, label in [
        ("codebook_file", "codebook_sha256", "最終コードブック"),
        ("change_log_file", "change_log_sha256", "変更履歴"),
        ("checklist_file", "checklist_sha256", "凍結チェック"),
        ("machine_spec_file", "machine_spec_sha256", "機械仕様"),
    ]:
        file = _safe_frozen_member(FINAL_RULES_ROOT, manifest.get(file_key), label)
        if not file.exists() or workflow.sha256(file) != manifest.get(hash_key):
            raise PostReviewError(f"{label}が凍結manifestと一致しません")
        _require_mode(file, 0o400, label)
    _, development_consensus = _consensus_manifest("開発用")
    if manifest.get("development_consensus_manifest_sha256") != workflow.sha256(
        CONSENSUS_ROOT / "開発用" / "manifest.json"
    ):
        raise PostReviewError("最終規則と開発用合意値が一致しません")
    if manifest.get("development_codebook_version") != development_consensus.get("codebook_version"):
        raise PostReviewError("最終規則の開発用コードブック版が不一致です")
    try:
        machine_spec = json.loads(FINAL_MACHINE_SPEC.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise PostReviewError("最終評価用機械仕様を読めません") from error
    expected_spec = {
        "input_columns": workflow.INPUT_COLUMNS,
        "required_columns": workflow.REQUIRED_COLUMNS,
        "site_columns": workflow.SITE_COLUMNS,
        "choices": workflow.CHOICES,
        "definitions": workflow.DEFINITIONS,
        "validation_implementation_sha256": workflow.sha256(Path(workflow.__file__).resolve()),
        "schema_change_from_development": False,
        "automatic_classifier_status": MODEL_RULE_STATUS,
    }
    if (
        machine_spec.get("schema_version") != "Step11-MachineReadableCodebook-1.0"
        or machine_spec.get("final_version") != manifest.get("final_version")
        or any(machine_spec.get(key) != value for key, value in expected_spec.items())
    ):
        raise PostReviewError("最終評価用機械仕様と固定された入力・矛盾規則が一致しません")
    if manifest.get("model_rule_status") != MODEL_RULE_STATUS:
        raise PostReviewError("互換する自動分類器が未凍結である状態記録が不一致です")
    return manifest


def _build_final_workbook(frame: pd.DataFrame, reviewer: str, destination: Path, rules: dict[str, object]) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = workflow.INPUT_SHEET
    choice_names = workflow.add_choice_sheet(book)
    codebook_sheet = book.create_sheet("判定基準")
    codebook_sheet["A1"] = f"最終評価用固定コードブック {rules['final_version']}"
    codebook_sheet["A1"].font = Font(bold=True, size=14, color="006100")
    for row, line in enumerate(FINAL_CODEBOOK.read_text(encoding="utf-8").splitlines(), start=3):
        codebook_sheet.cell(row, 1, line).alignment = Alignment(wrap_text=True, vertical="top")
    codebook_sheet.column_dimensions["A"].width = 120
    metadata_sheet = book.create_sheet("メタデータ")
    metadata_rows = [
        ("コードブック版", str(rules["final_version"])), ("承認状態", "最終規則凍結済み"),
        ("評価段階", "最終評価用"), ("評価者", reviewer), ("作成日時（UTC）", utc_now()),
        ("元CSV", source_csv("最終評価用", reviewer).name),
        ("元CSV SHA-256", workflow.sha256(source_csv("最終評価用", reviewer))),
        ("コードブックMarkdown SHA-256", workflow.sha256(FINAL_CODEBOOK)),
        ("最終規則manifest SHA-256", workflow.sha256(FINAL_RULES_MANIFEST)),
        ("生成コード SHA-256", workflow.sha256(Path(__file__).resolve())),
        ("対象件数", str(workflow.EXPECTED_ROWS)),
        ("禁止閲覧", "相手回答／自動回答／歯牙障害区分／内部ID"),
    ]
    for key, value in metadata_rows:
        metadata_sheet.append([key, value])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(workflow.INPUT_COLUMNS))
    sheet["A1"] = f"Step11 最終評価用100件・評価者{reviewer}（{rules['final_version']}）"
    sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(workflow.INPUT_COLUMNS))
    sheet["A2"] = "凍結規則による最終評価です。定義を変更せず、曖昧例は判定不能と逸脱記録で扱ってください。"
    sheet["A2"].font = Font(bold=True, color="006100")
    sheet["A2"].fill = PatternFill("solid", fgColor="C6EFCE")
    for col, header in enumerate(workflow.INPUT_COLUMNS, start=1):
        cell = sheet.cell(workflow.HEADER_ROW, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True)
    for row_number, (_, source) in enumerate(frame.iterrows(), start=workflow.FIRST_DATA_ROW):
        sheet.cell(row_number, 1, str(source[workflow.ID_COLUMN])).protection = Protection(locked=True)
        sheet.cell(row_number, 2, str(source[workflow.TEXT_COLUMN])).protection = Protection(locked=True)
        sheet.cell(row_number, 2).alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(3, len(workflow.INPUT_COLUMNS) + 1):
            cell = sheet.cell(row_number, col, "")
            cell.protection = Protection(locked=False)
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, field in enumerate(workflow.INPUT_COLUMNS, start=1):
        if field not in workflow.CHOICES:
            continue
        validation = DataValidation(type="list", formula1=f"={choice_names[field]}")
        validation.showErrorMessage = True
        sheet.add_data_validation(validation)
        validation.add(
            f"{get_column_letter(col)}{workflow.FIRST_DATA_ROW}:"
            f"{get_column_letter(col)}{workflow.FIRST_DATA_ROW + workflow.EXPECTED_ROWS - 1}"
        )
    sheet.freeze_panes = "C4"
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 80
    for col in range(3, len(workflow.INPUT_COLUMNS) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 44 if workflow.INPUT_COLUMNS[col - 1] == workflow.COMMENT_COLUMN else 22
    for protected_sheet in book.worksheets:
        protected_sheet.protection.sheet = True
        protected_sheet.protection.password = "Step11FinalReview"
    book.save(destination)


def prepare_final_excels(confirm_exact: str) -> None:
    if confirm_exact != FINAL_PREPARE_CONFIRM:
        raise PostReviewError(f"--confirm-exactには『{FINAL_PREPARE_CONFIRM}』を指定してください")
    rules = _load_final_rules()
    if frozen_dir("最終評価用").exists():
        raise PostReviewError("最終評価用原本が既に凍結済みです")
    targets = {reviewer: working_excel("最終評価用", reviewer) for reviewer in ["A", "B"]}
    if any(path.exists() for path in targets.values()) or FINAL_EXCELS_MANIFEST.exists():
        raise PostReviewError("既存の最終評価用Excelまたは生成記録を上書きしません")
    frames = {reviewer: _read_source("最終評価用", reviewer) for reviewer in ["A", "B"]}
    if not frames["A"].equals(frames["B"]):
        raise PostReviewError("最終評価用A/Bの標本・原文・順序が一致しません")
    development = _read_source("開発用", "A")
    if set(development[workflow.ID_COLUMN]) & set(frames["A"][workflow.ID_COLUMN]):
        raise PostReviewError("開発用と最終評価用のレビューIDが重複しています")
    if set(development[workflow.TEXT_COLUMN]) & set(frames["A"][workflow.TEXT_COLUMN]):
        raise PostReviewError("開発用と最終評価用の原文が重複しています")
    generated: dict[str, Path] = {}
    installed: list[Path] = []
    with tempfile.TemporaryDirectory(prefix=".final-excels-", dir=INTERNAL) as temp:
        temporary = Path(temp)
        workbook_records: dict[str, dict[str, str]] = {}
        for reviewer in ["A", "B"]:
            path = temporary / targets[reviewer].name
            _build_final_workbook(frames[reviewer], reviewer, path, rules)
            check, meta, errors = _workbook_rows(path)
            if errors or not check[[workflow.ID_COLUMN, workflow.TEXT_COLUMN]].equals(frames[reviewer]):
                raise PostReviewError(f"生成した最終評価者{reviewer}Excelの事前検査に失敗しました")
            if check[workflow.REQUIRED_COLUMNS + [workflow.COMMENT_COLUMN]].apply(
                lambda column: column.str.strip().ne("").any()
            ).any():
                raise PostReviewError("生成した最終評価Excelに回答が入っています")
            if meta.get("最終規則manifest SHA-256") != workflow.sha256(FINAL_RULES_MANIFEST):
                raise PostReviewError("生成した最終評価Excelの規則ハッシュが不一致です")
            path.chmod(0o600)
            generated[reviewer] = path
            workbook_records[reviewer] = {
                "file": targets[reviewer].name,
                "source_csv_sha256": workflow.sha256(source_csv("最終評価用", reviewer)),
                "initial_workbook_sha256": workflow.sha256(path),
                "fixed_workbook_fingerprint_sha256": _workbook_fixed_fingerprint(
                    path,
                    workflow.INPUT_SHEET,
                    set(range(3, len(workflow.INPUT_COLUMNS) + 1)),
                    workflow.FIRST_DATA_ROW,
                ),
            }
        prepared_manifest_path = temporary / FINAL_EXCELS_MANIFEST.name
        json_write(prepared_manifest_path, {
            "schema_version": FINAL_EXCELS_SCHEMA,
            "status": "PREPARED",
            "created_at_utc": utc_now(),
            "final_version": str(rules["final_version"]),
            "final_rules_manifest_sha256": workflow.sha256(FINAL_RULES_MANIFEST),
            "ordered_review_ids_sha256": ordered_ids_sha256(frames["A"][workflow.ID_COLUMN].tolist()),
            "workbooks": workbook_records,
            "auto_answer_inputs_used": False,
        })
        try:
            for reviewer in ["A", "B"]:
                os.replace(generated[reviewer], targets[reviewer])
                targets[reviewer].chmod(0o600)
                installed.append(targets[reviewer])
            os.replace(prepared_manifest_path, FINAL_EXCELS_MANIFEST)
            FINAL_EXCELS_MANIFEST.chmod(0o600)
            installed.append(FINAL_EXCELS_MANIFEST)
        except Exception:
            for path in reversed(installed):
                path.unlink(missing_ok=True)
            raise
    _load_final_excel_manifest()
    print("FINAL_EXCELS_PREPARED: A/B各100件 / 開発用と非重複 / 自動回答・歯牙障害区分なし")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    commands = parser.add_subparsers(dest="command", required=True)
    freeze = commands.add_parser("freeze-submissions", help="完了したA/B回答原本を解析前に凍結する")
    freeze.add_argument("--phase", choices=["開発用", "最終評価用"], required=True)
    freeze.add_argument("--confirm-exact", required=True)
    validate_submissions_parser = commands.add_parser(
        "validate-submissions", help="A/B回答を変更せず構造・進捗・矛盾を検査する"
    )
    validate_submissions_parser.add_argument("--phase", choices=["開発用", "最終評価用"], required=True)
    validate_submissions_parser.add_argument("--require-complete", action="store_true")
    adjudication = commands.add_parser("prepare-adjudication", help="凍結原本から第三者調停票を一度だけ生成する")
    adjudication.add_argument("--phase", choices=["開発用", "最終評価用"], required=True)
    for name in ["validate-adjudication", "check-adjudication"]:
        validate_parser = commands.add_parser(name, help="調停票の構造・進捗・合意値矛盾を検査する")
        validate_parser.add_argument("--phase", choices=["開発用", "最終評価用"], required=True)
        validate_parser.add_argument("--require-complete", action="store_true")
    consensus = commands.add_parser("freeze-consensus", help="完成した第三者調停と人手合意値を固定する")
    consensus.add_argument("--phase", choices=["開発用", "最終評価用"], required=True)
    consensus.add_argument("--confirmed-by", required=True)
    consensus.add_argument("--confirm-exact", required=True)
    commands.add_parser("prepare-final-rules-draft", help="開発用合意後に最終規則の編集・確認用案を作る")
    final_rules = commands.add_parser("freeze-final-rules", help="最終コードブックと規則を固定する")
    final_rules.add_argument("--final-version", required=True)
    final_rules.add_argument("--confirmed-by", required=True)
    final_rules.add_argument("--change-summary", required=True)
    final_rules.add_argument("--model-rule-status", required=True)
    final_rules.add_argument("--confirm-exact", required=True)
    final_excels = commands.add_parser("prepare-final-excels", help="凍結規則で最終評価用A/B Excelを生成する")
    final_excels.add_argument("--confirm-exact", required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    try:
        with process_lock():
            if args.command == "freeze-submissions":
                freeze_submissions(args.phase, args.confirm_exact)
            elif args.command == "validate-submissions":
                validate_submissions(args.phase, args.require_complete)
            elif args.command == "prepare-adjudication":
                prepare_adjudication(args.phase)
            elif args.command in {"validate-adjudication", "check-adjudication"}:
                validate_adjudication(args.phase, args.require_complete)
            elif args.command == "freeze-consensus":
                freeze_consensus(args.phase, args.confirmed_by, args.confirm_exact)
            elif args.command == "prepare-final-rules-draft":
                prepare_final_rules_draft()
            elif args.command == "freeze-final-rules":
                freeze_final_rules(
                    args.final_version, args.confirmed_by, args.change_summary,
                    args.model_rule_status, args.confirm_exact,
                )
            elif args.command == "prepare-final-excels":
                prepare_final_excels(args.confirm_exact)
    except (PostReviewError, workflow.WorkflowError) as error:
        raise SystemExit(f"POST_REVIEW_STOP: {error}") from error


if __name__ == "__main__":
    main()
