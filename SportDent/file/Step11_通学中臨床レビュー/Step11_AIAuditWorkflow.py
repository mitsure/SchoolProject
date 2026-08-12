"""通学中590件AI暫定分類の、一人研究者による盲検標本監査を準備・検査・集計する。"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

import Step11_ReviewWorkflow as workflow
import Step11_SoloAssistedReview as solo
import Step11_FullAIAnalysis as full_analysis


ROOT = solo.OUTPUT / "Audit100"
SEALED = ROOT / "AfterAudit_DoNotOpen"
RESULTS = ROOT / "Results"
ASSIGNMENT = ROOT / "Step11-46_AI標本監査100件_盲検割付.csv"
WORKBOOK = ROOT / "Step11-47_一人確認用_AI標本監査100件.xlsx"
DESIGN = ROOT / "Step11-48_AI標本監査_設計.md"
MANIFEST = ROOT / "Step11-49_AI標本監査_マニフェスト.json"
FROZEN_AI = SEALED / "Step11-50_AI標本監査100件_AI回答・区分_確認完了まで非表示.csv"

AGREEMENT = RESULTS / "Step11-51_AI標本監査_項目別一致率.csv"
BY_STRATUM = RESULTS / "Step11-52_AI標本監査_層別一致率.csv"
CONFUSION = RESULTS / "Step11-53_AI標本監査_混同行列.csv"
REPORT = RESULTS / "Step11-54_AI標本監査_結果レポート.md"
RESULT_MANIFEST = RESULTS / "Step11-55_AI標本監査_結果マニフェスト.json"

SCHEMA = "Step11-AIAudit-1.0"
SEED = 20260812
EXPECTED = 100
STRATA = {
    "起点機転判定不能": 30,
    "最優先・起点判定可能": 30,
    "要確認": 20,
    "通常": 20,
}
INPUT_COLUMNS = workflow.INPUT_COLUMNS
FIELDS = [field for field in workflow.CHOICES if field != workflow.CONFIRM_COLUMN]
SHEET = "監査入力"
HEADER_ROW = 3
FIRST_ROW = 4


class AuditError(RuntimeError):
    """監査原本を変更せず停止すべき問題。"""


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def secure_tree(path: Path) -> None:
    path.chmod(0o700)
    for child in path.rglob("*"):
        child.chmod(0o700 if child.is_dir() else 0o600)


def json_write(path: Path, value: object) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    path.chmod(0o600)


def full_frame() -> pd.DataFrame:
    if not full_analysis.FULL_CSV.exists() or not full_analysis.MANIFEST.exists():
        raise AuditError("590件AI暫定分類またはマニフェストがありません")
    record = json.loads(full_analysis.MANIFEST.read_text(encoding="utf-8"))
    if record.get("blind_classification_sha256_before_unblinding") != sha256(full_analysis.FULL_CSV):
        raise AuditError("590件AI暫定分類のハッシュが一致しません")
    frame = pd.read_csv(full_analysis.FULL_CSV, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    if len(frame) != 590 or frame[workflow.ID_COLUMN].duplicated().any():
        raise AuditError("590件AI暫定分類の件数またはIDが不正です")
    return frame


def select_sample() -> tuple[pd.DataFrame, pd.DataFrame, dict[str, int]]:
    frame = full_frame()
    development = pd.read_csv(solo.AI_REVIEWED_CSV, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    pool = frame.loc[~frame[workflow.ID_COLUMN].isin(development[workflow.ID_COLUMN])].copy()
    if len(pool) != 490:
        raise AuditError(f"開発用100件を除いた件数が490件ではありません: {len(pool)}")

    pool["監査層"] = ""
    pool.loc[pool["起点機転"].eq("判定不能"), "監査層"] = "起点機転判定不能"
    pool.loc[pool["監査層"].eq("") & pool["AI確認優先度"].eq("最優先"), "監査層"] = "最優先・起点判定可能"
    pool.loc[pool["監査層"].eq("") & pool["AI確認優先度"].eq("要確認"), "監査層"] = "要確認"
    pool.loc[pool["監査層"].eq("") & pool["AI確認優先度"].eq("通常"), "監査層"] = "通常"
    if pool["監査層"].eq("").any():
        raise AuditError("監査層へ割り付けできない事例があります")

    selected_parts: list[pd.DataFrame] = []
    population_sizes: dict[str, int] = {}
    for offset, (stratum, count) in enumerate(STRATA.items()):
        candidates = pool[pool["監査層"].eq(stratum)]
        population_sizes[stratum] = len(candidates)
        if len(candidates) < count:
            raise AuditError(f"監査層の候補数が不足しています: {stratum} {len(candidates)}/{count}")
        selected_parts.append(candidates.sample(n=count, random_state=SEED + offset))
    selected = pd.concat(selected_parts, ignore_index=True)
    selected = selected.sample(frac=1, random_state=SEED + 99).reset_index(drop=True)
    selected.insert(0, "監査番号", [f"AUD-{index:03d}" for index in range(1, EXPECTED + 1)])
    if len(selected) != EXPECTED or selected[workflow.ID_COLUMN].duplicated().any():
        raise AuditError("監査標本の件数またはIDが不正です")

    mapping = pd.read_csv(full_analysis.MAPPING, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    outcome = mapping[[workflow.ID_COLUMN, "歯牙障害"]]
    frozen = selected.merge(outcome, on=workflow.ID_COLUMN, how="left", validate="one_to_one")
    if frozen["歯牙障害"].eq("").any():
        raise AuditError("監査標本へ歯牙障害区分を結合できません")
    blind = selected[["監査番号", workflow.ID_COLUMN, workflow.TEXT_COLUMN]].copy()
    return blind, frozen, population_sizes


def add_choice_sheet(book: Workbook) -> dict[str, str]:
    sheet = book.create_sheet("選択肢")
    result: dict[str, str] = {}
    for index, (field, choices) in enumerate(workflow.CHOICES.items(), start=1):
        column = get_column_letter(index)
        sheet.cell(1, index, field)
        for row, choice in enumerate(choices, start=2):
            sheet.cell(row, index, choice)
        name = f"audit_choices_{index:02d}"
        book.defined_names.add(DefinedName(name, attr_text=f"'選択肢'!${column}$2:${column}${len(choices)+1}"))
        result[field] = name
    return result


def build_workbook(blind: pd.DataFrame, destination: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    choice_names = add_choice_sheet(book)

    codebook = book.create_sheet("判定基準")
    codebook["A1"] = f"AI標本監査・承認済みコードブック {workflow.CODEBOOK_VERSION}"
    codebook["A2"] = "AI回答と歯牙障害区分は伏せています。原文だけから判定してください。"
    for row, line in enumerate(workflow.CODEBOOK_SOURCE.read_text(encoding="utf-8").splitlines(), start=4):
        codebook.cell(row, 1, line)
    codebook.column_dimensions["A"].width = 120

    metadata = book.create_sheet("メタデータ")
    rows = [
        ("方式", "一人研究者によるAI暫定分類の盲検標本監査"), ("対象件数", "100"),
        ("コードブック版", workflow.CODEBOOK_VERSION), ("抽出seed", str(SEED)),
        ("作成日時UTC", utc_now()), ("AI回答", "確認完了まで非表示"),
        ("歯牙障害区分", "確認完了まで非表示"), ("割付CSV SHA-256", "生成後マニフェスト参照"),
    ]
    for row, (key, value) in enumerate(rows, start=1):
        metadata.cell(row, 1, key).font = Font(bold=True)
        metadata.cell(row, 2, value)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(INPUT_COLUMNS))
    sheet["A1"] = "Step11 AI標本監査100件：AI回答を見ず原文だけから判定"
    sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(INPUT_COLUMNS))
    sheet["A2"] = "空欄は未入力です。確認後は行確認を『確認済み』にしてください。"
    sheet["A2"].font = Font(bold=True, color="9C0006")
    sheet["A2"].fill = PatternFill("solid", fgColor="FFF2CC")

    for column, header in enumerate(INPUT_COLUMNS, start=1):
        cell = sheet.cell(HEADER_ROW, column, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for offset, (_, row) in enumerate(blind.iterrows(), start=FIRST_ROW):
        sheet.cell(offset, 1, row[workflow.ID_COLUMN])
        sheet.cell(offset, 2, row[workflow.TEXT_COLUMN])
        sheet.cell(offset, 1).protection = Protection(locked=True)
        sheet.cell(offset, 2).protection = Protection(locked=True)
        sheet.cell(offset, 2).alignment = Alignment(wrap_text=True, vertical="top")
        for column in range(3, len(INPUT_COLUMNS) + 1):
            cell = sheet.cell(offset, column, "")
            cell.protection = Protection(locked=False)
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for field, name in choice_names.items():
        column = INPUT_COLUMNS.index(field) + 1
        validation = DataValidation(type="list", formula1=f"={name}", allow_blank=True)
        validation.showErrorMessage = True
        validation.error = "固定選択肢から選んでください。"
        sheet.add_data_validation(validation)
        validation.add(f"{get_column_letter(column)}{FIRST_ROW}:{get_column_letter(column)}{FIRST_ROW + EXPECTED - 1}")

    sheet.freeze_panes = "C4"
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(INPUT_COLUMNS))}{FIRST_ROW + EXPECTED - 1}"
    for index, header in enumerate(INPUT_COLUMNS, start=1):
        width = 80 if header == workflow.TEXT_COLUMN else 22
        if header == workflow.COMMENT_COLUMN:
            width = 44
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(FIRST_ROW, FIRST_ROW + EXPECTED):
        sheet.row_dimensions[row].height = 70
    sheet.protection.sheet = True
    sheet.protection.password = "Step11Audit"
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = True
    for other in book.worksheets[1:]:
        other.protection.sheet = True
        other.protection.password = "Step11Audit"
    book.save(destination)


def prepare() -> None:
    expected = [ASSIGNMENT, WORKBOOK, DESIGN, MANIFEST, FROZEN_AI]
    existing = [path for path in expected if path.exists()]
    if existing:
        raise AuditError("既存の監査資料を上書きしません: " + ", ".join(path.name for path in existing))
    blind, frozen, population_sizes = select_sample()
    ROOT.parent.mkdir(parents=True, exist_ok=True)
    ROOT.parent.chmod(0o700)
    with tempfile.TemporaryDirectory(prefix="step11-ai-audit-", dir=ROOT.parent) as temp_name:
        temp = Path(temp_name)
        blind.to_csv(temp / ASSIGNMENT.name, index=False, encoding="utf-8-sig")
        build_workbook(blind, temp / WORKBOOK.name)
        design = f"""# Step11 AI暫定分類 標本監査設計

- 母集団：通学中590件のうち、分類規則の確認に用いた開発用100件を除く490件
- 監査標本：100件
- 抽出seed：{SEED}
- 起点機転判定不能：30件（母集団{population_sizes['起点機転判定不能']}件）
- 最優先・起点判定可能：30件（母集団{population_sizes['最優先・起点判定可能']}件）
- 要確認：20件（母集団{population_sizes['要確認']}件）
- 通常：20件（母集団{population_sizes['通常']}件）

AI回答、AI確信度、監査層および歯牙障害区分はExcelに表示しない。研究者は原文だけで全項目を入力する。完了後にAI回答と比較し、項目別完全一致率、層別一致率および混同行列を作成する。

この監査はAIと研究者1名の一致を調べるもので、2名の歯科医師間一致ではない。判定不能・最優先を意図的に多く含むため、標本全体の単純一致率を590件全体の一致率とみなさない。層別結果と母集団構成を併記する。
"""
        (temp / DESIGN.name).write_text(design, encoding="utf-8")
        sealed_temp = temp / "sealed"
        sealed_temp.mkdir()
        frozen.to_csv(sealed_temp / FROZEN_AI.name, index=False, encoding="utf-8-sig")

        assignment_hash = sha256(temp / ASSIGNMENT.name)
        record = {
            "schema_version": SCHEMA, "status": "PREPARED_HUMAN_REVIEW_PENDING", "created_at_utc": utc_now(),
            "seed": SEED, "rows": EXPECTED, "stratum_sample_sizes": STRATA,
            "stratum_population_sizes": population_sizes, "assignment_sha256": assignment_hash,
            "full_ai_sha256": sha256(full_analysis.FULL_CSV), "development_ai_sha256": sha256(solo.AI_REVIEWED_CSV),
            "frozen_ai_answers_sha256": sha256(sealed_temp / FROZEN_AI.name),
            "workflow_sha256": sha256(Path(__file__).resolve()), "codebook_sha256": sha256(workflow.CODEBOOK_SOURCE),
        }
        (temp / MANIFEST.name).write_text(json.dumps(record, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")

        ROOT.mkdir(parents=True, exist_ok=False)
        SEALED.mkdir(parents=True, exist_ok=False)
        ROOT.chmod(0o700); SEALED.chmod(0o700)
        for name in [ASSIGNMENT.name, WORKBOOK.name, DESIGN.name, MANIFEST.name]:
            os.replace(temp / name, ROOT / name)
        os.replace(sealed_temp / FROZEN_AI.name, FROZEN_AI)
    secure_tree(ROOT)
    print(f"AI_AUDIT_PREPARE_OK: {EXPECTED}件 / Excel={WORKBOOK}")


def read_answers() -> pd.DataFrame:
    if not WORKBOOK.exists():
        raise AuditError("監査Excelがありません")
    book = load_workbook(WORKBOOK, data_only=False, keep_links=True)
    if set(book.sheetnames) != {SHEET, "選択肢", "判定基準", "メタデータ"}:
        book.close(); raise AuditError("監査Excelのシート構成が不正です")
    for current in book.worksheets:
        if current.sheet_state != "visible":
            book.close(); raise AuditError("監査Excelに非表示シートがあります")
        for row in current.iter_rows():
            for cell in row:
                if cell.data_type == "f" or cell.hyperlink is not None:
                    book.close(); raise AuditError("監査Excelに数式またはリンクがあります")
    sheet = book[SHEET]
    headers = [sheet.cell(HEADER_ROW, column).value for column in range(1, len(INPUT_COLUMNS) + 1)]
    if headers != INPUT_COLUMNS:
        book.close(); raise AuditError("監査Excelの列または列順が不正です")
    rows = []
    for row_number in range(FIRST_ROW, FIRST_ROW + EXPECTED):
        rows.append({
            header: "" if sheet.cell(row_number, column).value is None else str(sheet.cell(row_number, column).value).strip()
            for column, header in enumerate(INPUT_COLUMNS, start=1)
        })
    book.close()
    answers = pd.DataFrame(rows)
    assignment = pd.read_csv(ASSIGNMENT, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    if not answers[[workflow.ID_COLUMN, workflow.TEXT_COLUMN]].equals(assignment[[workflow.ID_COLUMN, workflow.TEXT_COLUMN]]):
        raise AuditError("監査ExcelのレビューIDまたは原文が割付CSVと一致しません")
    return answers


def validate(require_complete: bool = False) -> pd.DataFrame:
    record = json.loads(MANIFEST.read_text(encoding="utf-8"))
    expected_hashes = {
        "assignment_sha256": sha256(ASSIGNMENT), "full_ai_sha256": sha256(full_analysis.FULL_CSV),
        "development_ai_sha256": sha256(solo.AI_REVIEWED_CSV), "frozen_ai_answers_sha256": sha256(FROZEN_AI),
        "workflow_sha256": sha256(Path(__file__).resolve()), "codebook_sha256": sha256(workflow.CODEBOOK_SOURCE),
    }
    mismatch = [key for key, value in expected_hashes.items() if record.get(key) != value]
    if mismatch:
        raise AuditError("監査資料のハッシュが一致しません: " + "/".join(mismatch))
    answers = read_answers()
    errors: list[str] = []
    completed = 0
    for index, row in answers.iterrows():
        values = {column: str(row.get(column, "")).strip() for column in INPUT_COLUMNS}
        problems = workflow.row_problems(values)
        if values[workflow.CONFIRM_COLUMN] == "確認済み":
            completed += 1
            missing = [column for column in workflow.REQUIRED_COLUMNS if not values.get(column)]
            if missing:
                problems.append("確認済みだが必須欄が空: " + "/".join(missing))
        elif values[workflow.CONFIRM_COLUMN]:
            problems.append("行確認が選択肢外")
        if problems:
            errors.append(f"行{index + FIRST_ROW}: " + " / ".join(problems))
    if require_complete and completed != EXPECTED:
        errors.append(f"確認済み件数が100件ではありません: {completed}/100")
    if errors:
        raise AuditError("\n".join(errors[:30]))
    print(f"AI_AUDIT_VALIDATE_OK: 確認済み {completed}/{EXPECTED}件")
    return answers


def score() -> None:
    outputs = [AGREEMENT, BY_STRATUM, CONFUSION, REPORT, RESULT_MANIFEST]
    if any(path.exists() for path in outputs):
        raise AuditError("既存の監査結果を上書きしません")
    answers = validate(require_complete=True)
    frozen = pd.read_csv(FROZEN_AI, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    merged = answers.merge(frozen, on=[workflow.ID_COLUMN, workflow.TEXT_COLUMN], suffixes=("_研究者", "_AI"), validate="one_to_one")
    if len(merged) != EXPECTED:
        raise AuditError("研究者回答とAI回答を100件結合できません")

    agreement_rows: list[dict[str, object]] = []
    stratum_rows: list[dict[str, object]] = []
    confusion_rows: list[dict[str, object]] = []
    population_sizes = json.loads(MANIFEST.read_text(encoding="utf-8"))["stratum_population_sizes"]
    for field in FIELDS:
        equal = merged[f"{field}_研究者"] == merged[f"{field}_AI"]
        weighted_matches = 0.0
        weighted_total = 0
        for stratum, group in merged.groupby("監査層"):
            group_equal = group[f"{field}_研究者"] == group[f"{field}_AI"]
            matches = int(group_equal.sum())
            rate = matches / len(group) * 100
            population = int(population_sizes[stratum])
            weighted_matches += population * rate / 100
            weighted_total += population
            stratum_rows.append({"項目": field, "監査層": stratum, "監査件数": len(group), "一致件数": matches, "完全一致率（%）": round(rate, 1), "層の母集団件数": population})
        agreement_rows.append({
            "項目": field, "監査件数": EXPECTED, "一致件数": int(equal.sum()),
            "標本内完全一致率（%）": round(float(equal.mean() * 100), 1),
            "層構成補正後一致率（%・記述的）": round(weighted_matches / weighted_total * 100, 1),
        })
        table = pd.crosstab(merged[f"{field}_研究者"], merged[f"{field}_AI"], dropna=False)
        for human_value in table.index:
            for ai_value in table.columns:
                confusion_rows.append({"項目": field, "研究者回答": human_value, "AI回答": ai_value, "件数": int(table.loc[human_value, ai_value])})

    agreement = pd.DataFrame(agreement_rows)
    by_stratum = pd.DataFrame(stratum_rows)
    confusion = pd.DataFrame(confusion_rows)
    RESULTS.mkdir(parents=True, exist_ok=False); RESULTS.chmod(0o700)
    agreement.to_csv(AGREEMENT, index=False, encoding="utf-8-sig")
    by_stratum.to_csv(BY_STRATUM, index=False, encoding="utf-8-sig")
    confusion.to_csv(CONFUSION, index=False, encoding="utf-8-sig")
    origin = agreement.loc[agreement["項目"].eq("起点機転")].iloc[0]
    direct = agreement.loc[agreement["項目"].eq("口腔・顔面への直接外力")].iloc[0]
    report = f"""# Step11 AI標本監査 結果レポート

- 監査対象：開発用100件と重複しない100件
- 方式：AI回答と歯牙障害区分を伏せ、研究者1名が原文だけから判定
- 起点機転の標本内完全一致率：{origin['標本内完全一致率（%）']}%
- 起点機転の層構成補正後一致率：{origin['層構成補正後一致率（%・記述的）']}%
- 直接外力の標本内完全一致率：{direct['標本内完全一致率（%）']}%
- 直接外力の層構成補正後一致率：{direct['層構成補正後一致率（%・記述的）']}%

判定不能・最優先を意図的に多く含むため、標本内一致率だけを590件全体へ一般化しない。層構成補正後一致率も記述的な推定であり、2名の歯科医師間一致やAIの外部検証性能ではない。不一致内容を確認し、必要なら分類規則を版更新して590件を再分類する。
"""
    REPORT.write_text(report, encoding="utf-8")
    record = {
        "schema_version": "Step11-AIAuditResults-1.0", "status": "SINGLE_RESEARCHER_AUDIT_COMPLETE",
        "created_at_utc": utc_now(), "rows": EXPECTED, "workbook_sha256": sha256(WORKBOOK),
        "frozen_ai_sha256": sha256(FROZEN_AI), "agreement_sha256": sha256(AGREEMENT),
        "by_stratum_sha256": sha256(BY_STRATUM), "confusion_sha256": sha256(CONFUSION),
        "report_sha256": sha256(REPORT), "warning": "AI対研究者1名の一致。2名の歯科医師間一致ではない。",
    }
    json_write(RESULT_MANIFEST, record)
    secure_tree(ROOT)
    print(f"AI_AUDIT_SCORE_OK: {REPORT}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    subs = parser.add_subparsers(dest="command", required=True)
    subs.add_parser("prepare")
    validate_parser = subs.add_parser("validate")
    validate_parser.add_argument("--require-complete", action="store_true")
    subs.add_parser("score")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    try:
        if args.command == "prepare": prepare()
        elif args.command == "validate": validate(args.require_complete)
        elif args.command == "score": score()
    except AuditError as error:
        raise SystemExit(f"STOP: {error}") from error


if __name__ == "__main__":
    main()
