"""Step11の2名独立レビュー用Excelを安全に準備・検査する。

Excelは人が入力する原本、既存CSVは割付と照合だけに使う。prepareは既存の
Excelを上書きせず、現時点では開発用100件だけを生成する。
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


HERE = Path(__file__).resolve().parent
PROJECT = HERE.parent.parent
OUTPUT = PROJECT / "CreateData" / "Step11_通学中臨床レビュー"
INTERNAL = OUTPUT / "InternalReview"
CODEBOOK_SOURCE = HERE / "Step11_ReviewCodebook_candidate.md"
CODEBOOK_OUTPUT = OUTPUT / "Step11-18_固定コードブック候補.csv"
SENTINEL = INTERNAL / "REVIEW_WORKFLOW_INITIALIZED.lock"
APPROVAL = INTERNAL / "CODEBOOK_APPROVED.json"
APPROVAL_SHEET = INTERNAL / "Step11-19_コードブック承認チェックシート.csv"

CODEBOOK_VERSION = "Step11-CB-1.0.0-rc1"
APPROVAL_SHEET_VERSION = "Step11-ApprovalSheet-1.0"
APPROVAL_PENDING = "研究者確認待ち（入力禁止）"
APPROVAL_COMPLETE = "承認済み"
READ_CONFIRMATION = "Step11-02を読んで確認済み"
APPROVAL_INTENT = "この内容で開発用コードブックを承認する"
INPUT_SHEET = "入力"
HEADER_ROW = 3
FIRST_DATA_ROW = 4
EXPECTED_ROWS = 100
EXPECTED_SHEETS = {"入力", "判定基準", "選択肢", "メタデータ"}

ID_COLUMN = "レビューID"
TEXT_COLUMN = "原文（内部確認用）"
COMMENT_COLUMN = "コメント"
CONFIRM_COLUMN = "行確認"

CHOICES: dict[str, list[str]] = {
    "起点機転": ["転倒・つまずき", "衝突・接触", "転落・落下", "その他", "判定不能"],
    "出来事の順序": ["単一機転", "複数機転・順序明記", "複数機転・順序不明", "判定不能"],
    "最終接触対象": [
        "路面・地面・床", "車両", "自転車部品", "人", "構造物", "その他", "該当なし", "判定不能",
    ],
    "口腔・顔面への直接外力": ["あり（明記）", "なし（明記）", "記載なし", "判定不能"],
    "前歯部の記載": ["記載あり", "記載なし", "判定不能"],
    "その他歯牙の記載": ["記載あり", "記載なし", "判定不能"],
    "歯牙・部位不明の記載": ["記載あり", "記載なし", "判定不能"],
    "口腔・口唇の記載": ["記載あり", "記載なし", "判定不能"],
    "顔面・顎の記載": ["記載あり", "記載なし", "判定不能"],
    "その他部位の記載": ["記載あり", "記載なし", "判定不能"],
    "ヘルメット": ["使用あり（明記）", "使用なし（明記）", "言及あり・使用状況不明", "記載なし", "該当なし"],
    "マウスガード": ["使用あり（明記）", "使用なし（明記）", "言及あり・使用状況不明", "記載なし", "該当なし"],
    "予防可能性": [
        "修正可能要因の明記あり", "修正可能要因の可能性あり", "修正可能要因の記載なし", "判定不能",
    ],
    "予防可能性の根拠区分": [
        "行動", "路面・天候", "車両・操作", "施設・構造物", "保護具", "複数", "その他", "該当なし",
    ],
    "判定不能理由": ["該当なし", "原文情報不足", "時系列不明", "複数候補", "用語が曖昧", "対象外疑い", "その他"],
    CONFIRM_COLUMN: ["確認済み"],
}

INPUT_COLUMNS = [ID_COLUMN, TEXT_COLUMN, *CHOICES.keys()]
INPUT_COLUMNS.insert(-1, COMMENT_COLUMN)
REQUIRED_COLUMNS = [column for column in INPUT_COLUMNS if column not in {ID_COLUMN, TEXT_COLUMN, COMMENT_COLUMN}]
SITE_COLUMNS = [
    "前歯部の記載", "その他歯牙の記載", "歯牙・部位不明の記載",
    "口腔・口唇の記載", "顔面・顎の記載", "その他部位の記載",
]

DEFINITIONS: dict[str, dict[str, str]] = {
    "起点機転": {
        "転倒・つまずき": "事故経過の起点が、バランス喪失、つまずき、滑り等による転倒である。",
        "衝突・接触": "人・車両・構造物等との接触が転倒等より先に生じた。",
        "転落・落下": "高低差を伴う落下が事故の起点である。",
        "その他": "起点は特定できるが既定3分類外。コメント必須。",
        "判定不能": "起点の情報不足、順序不明または矛盾により特定できない。",
    },
    "出来事の順序": {
        "単一機転": "判定に必要な出来事が1つだけ記載される。",
        "複数機転・順序明記": "複数の出来事と前後関係が明記される。",
        "複数機転・順序不明": "複数の出来事があるが前後関係を決められない。",
        "判定不能": "出来事の数または記載自体を判断できない。",
    },
    "最終接触対象": {
        "路面・地面・床": "道路、歩道、アスファルト、地面、校庭、床等。",
        "車両": "自動車、二輪車、自転車等の車体。",
        "自転車部品": "ハンドル、かご、フレーム、ペダル等。",
        "人": "歩行者、同乗者、他の児童生徒等。",
        "構造物": "壁、柱、塀、ガードレール、縁石、門扉等。",
        "その他": "接触対象は特定できるが既定分類外。コメント必須。",
        "該当なし": "口腔・顔面への接触対象が存在しないことが明確。",
        "判定不能": "対象の記載不足、複数候補または曖昧さにより決められない。",
    },
    "口腔・顔面への直接外力": {
        "あり（明記）": "口、歯、顔面、顎等の接触・強打が明記される。",
        "なし（明記）": "口腔・顔面への直接接触がなかったことが明記される。",
        "記載なし": "直接外力の有無への言及がない。事実として『なし』ではない。",
        "判定不能": "関連記載はあるが直接外力かを決められない。",
    },
    "予防可能性": {
        "修正可能要因の明記あり": "具体的な修正可能要因が明記される。",
        "修正可能要因の可能性あり": "修正可能要因が示唆されるが確定できない。",
        "修正可能要因の記載なし": "修正可能要因への言及がない。予防不能という意味ではない。",
        "判定不能": "記載が曖昧または矛盾し分類できない。",
    },
}


class WorkflowError(RuntimeError):
    """レビュー原本を変更せず処理を停止すべき問題。"""


def approval_sheet_rows() -> list[dict[str, str]]:
    """人が入力する欄と推奨案を分離した、非公開の承認チェック項目。"""
    agreement_proposal = (
        "起点機転は調停前A/B全100件で、判定不能を独立カテゴリとして含める。"
        "完全一致率80%以上、Wilson 95%CI下限70%以上、非加重Cohenκ 0.60以上、"
        "10,000回ペアbootstrapのκ95%CI下限0.40以上を全て満たす。"
        "有効bootstrapが95%未満ならκCI条件は未達とする。"
    )
    sparse_proposal = (
        "少なくとも一方が選んだ事例の和集合5件未満を希少カテゴリとする。"
        "結果確認後の統合・削除・置換はせず、件数と混同行列を記述する。"
        "κまたはCI算出不能は0/1へ置換せず算出不能とし、分布・一致率・Wilson 95%CI・"
        "判定不能件数を併記する。判定不能は主解析の独立カテゴリに含める。"
    )
    protection_proposal = (
        "ヘルメット：徒歩・鉄道・バス等の非乗車移動が原文に明記された場合のみ該当なし。"
        "自転車・二輪車・移動方法不明で言及がなければ記載なし。"
        "マウスガード：開発用では該当なしを使用せず、言及がなければ記載なし。"
    )
    rows = [
        ("sheet_version", "承認チェックシート版", APPROVAL_SHEET_VERSION, "はい", APPROVAL_SHEET_VERSION, "固定値。変更しない。"),
        ("codebook_version", "確認したコードブック版", CODEBOOK_VERSION, "はい", CODEBOOK_VERSION, "固定値。変更しない。"),
        ("codebook_read_confirmed", "判定基準を読んだ確認", "", "はい", READ_CONFIRMATION, "入力値へ例文を完全一致で入力する。"),
        ("approved_by", "研究責任者", "", "はい", "研究責任者名または管理コード", "この承認を行う責任者。"),
        ("reviewer_a_confirmed", "評価者A確認", "", "はい", "評価者A名または管理コード", "コードブックを確認した評価者A。"),
        ("reviewer_b_confirmed", "評価者B確認", "", "はい", "評価者B名または管理コード", "コードブックを確認した評価者B。"),
        ("adjudicator", "第三者調停担当者", "", "はい", "調停担当者名または管理コード", "A/B不一致項目の調停担当者。"),
        ("protection_na_rule", "保護具『該当なし』適用条件", "", "はい", protection_proposal, "推奨案を確認し、採用文または修正文を入力する。"),
        ("agreement_threshold", "起点機転の一致度許容水準", "", "はい", agreement_proposal, "結果を見る前に指標・数値・CI判定を固定する。"),
        ("sparse_category_rule", "κ算出不能・希少カテゴリの扱い", "", "はい", sparse_proposal, "結果を見る前に扱いを固定する。"),
        ("approval_date", "承認日", "", "はい", "YYYY-MM-DD", "YYYY-MM-DD形式。"),
        ("approval_intent", "明示的な承認意思", "", "はい", APPROVAL_INTENT, "入力値へ例文を完全一致で入力する。"),
    ]
    return [
        {"項目ID": item_id, "確認項目": label, "入力値": value, "必須": required, "入力例・推奨案": proposal, "説明": explanation}
        for item_id, label, value, required, proposal, explanation in rows
    ]


def create_approval_sheet() -> None:
    """承認は実行せず、研究者が確認・記入する別紙だけを新規作成する。"""
    INTERNAL.mkdir(parents=True, exist_ok=True)
    INTERNAL.chmod(0o700)
    if APPROVAL.exists():
        raise WorkflowError("コードブックは既に承認済みです")
    if APPROVAL_SHEET.exists():
        raise WorkflowError(f"既存の承認チェックシートを上書きしません: {APPROVAL_SHEET}")
    pd.DataFrame(approval_sheet_rows()).to_csv(APPROVAL_SHEET, index=False, encoding="utf-8-sig")
    APPROVAL_SHEET.chmod(0o600)
    print(f"APPROVAL_SHEET_OK: 承認は未実行 / 入力先={APPROVAL_SHEET}")


def inspect_approval_sheet(require_complete: bool = False) -> dict[str, str]:
    """別紙の構造と値を検査する。存在や保存だけでは承認しない。"""
    if not APPROVAL_SHEET.exists():
        raise WorkflowError(f"承認チェックシートがありません: {APPROVAL_SHEET}")
    if APPROVAL_SHEET.is_symlink() or APPROVAL_SHEET.stat().st_mode & 0o777 != 0o600:
        raise WorkflowError("承認チェックシートのリンクまたは権限が不正です")
    expected = pd.DataFrame(approval_sheet_rows())
    actual = pd.read_csv(APPROVAL_SHEET, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    if list(actual.columns) != list(expected.columns):
        raise WorkflowError("承認チェックシートの列または列順が不一致です")
    if actual["項目ID"].duplicated().any() or actual["項目ID"].tolist() != expected["項目ID"].tolist():
        raise WorkflowError("承認チェックシートの項目ID、行順または重複が不正です")
    for column in ["確認項目", "必須", "入力例・推奨案", "説明"]:
        if not actual[column].equals(expected[column]):
            raise WorkflowError(f"承認チェックシートの固定列が変更されています: {column}")
    values = dict(zip(actual["項目ID"], actual["入力値"], strict=True))
    if values["sheet_version"] != APPROVAL_SHEET_VERSION or values["codebook_version"] != CODEBOOK_VERSION:
        raise WorkflowError("承認チェックシートまたはコードブック版が不一致です")
    dangerous = [item_id for item_id, value in values.items() if value.lstrip().startswith(("=", "+", "-", "@"))]
    if dangerous:
        raise WorkflowError(f"数式として解釈され得る入力があります: {dangerous}")
    pending = [item_id for item_id, value in values.items() if not value.strip()]
    if values.get("codebook_read_confirmed") not in {"", READ_CONFIRMATION}:
        raise WorkflowError("判定基準を読んだ確認文が指定文と一致しません")
    if values.get("approval_intent") not in {"", APPROVAL_INTENT}:
        raise WorkflowError("明示的な承認文が指定文と一致しません")
    if values.get("approval_date"):
        try:
            datetime.strptime(values["approval_date"], "%Y-%m-%d")
        except ValueError as error:
            raise WorkflowError("承認日はYYYY-MM-DDで入力してください") from error
    if require_complete and pending:
        raise WorkflowError(f"未入力の承認項目があります: {pending}")
    status = "入力完了（承認処理は未実行）" if not pending else f"未入力{len(pending)}項目（承認処理は未実行）"
    print(f"APPROVAL_SHEET_VALID: {status}")
    return values


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def secure_tree(path: Path) -> None:
    path.chmod(0o700)
    for child in path.rglob("*"):
        child.chmod(0o700 if child.is_dir() else 0o600)


def private_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.parent.chmod(0o700)
    path.write_text(text, encoding="utf-8")
    path.chmod(0o600)


def source_path(reviewer: str) -> Path:
    return INTERNAL / f"Step11-08{reviewer}_開発用_評価者{reviewer}_盲検100件.csv"


def workbook_path(reviewer: str) -> Path:
    return INTERNAL / f"Step11-08{reviewer}_開発用_評価者{reviewer}_盲検100件.xlsx"


def approved_snapshot_path() -> Path:
    return INTERNAL / f"Step11-18_固定コードブック_{CODEBOOK_VERSION}_承認済み.csv"


def read_source(reviewer: str) -> pd.DataFrame:
    path = source_path(reviewer)
    if not path.exists():
        raise WorkflowError(f"開発用の元CSVがありません: {path}")
    frame = pd.read_csv(path, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    missing = {ID_COLUMN, TEXT_COLUMN} - set(frame.columns)
    if missing:
        raise WorkflowError(f"元CSVの必須列がありません: {sorted(missing)}")
    frame = frame[[ID_COLUMN, TEXT_COLUMN]].copy()
    if len(frame) != EXPECTED_ROWS:
        raise WorkflowError(f"元CSVが{EXPECTED_ROWS}件ではありません: {path} / {len(frame)}件")
    if frame[ID_COLUMN].duplicated().any() or not frame[ID_COLUMN].str.fullmatch(r"R-[0-9A-F]{10}").all():
        raise WorkflowError(f"元CSVのレビューIDが不正です: {path}")
    if frame[TEXT_COLUMN].str.strip().eq("").any():
        raise WorkflowError(f"元CSVに空の原文があります: {path}")
    return frame


def codebook_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for field, values in CHOICES.items():
        for value in values:
            definition = DEFINITIONS.get(field, {}).get(value, "固定選択肢。詳細は判定基準Markdownを参照する。")
            comment_required = "はい" if (
                value == "その他"
                or (field == "その他部位の記載" and value == "記載あり")
                or (field == "判定不能理由" and value == "対象外疑い")
            ) else "いいえ"
            rows.append({
                "コードブック版": CODEBOOK_VERSION,
                "項目": field,
                "選択値": value,
                "定義": definition,
                "コメント必須": comment_required,
                "状態": "研究者確認待ち",
            })
    return rows


def approved_codebook_frame(record: dict[str, object]) -> pd.DataFrame:
    """候補表と承認記録から、承認済みsnapshotの唯一の期待値を作る。"""
    frame = pd.DataFrame(codebook_rows())
    frame["状態"] = APPROVAL_COMPLETE
    frame["承認者"] = str(record.get("approved_by", ""))
    frame["承認日"] = str(record.get("approval_date", ""))
    frame["保護具の該当なし適用条件"] = str(record.get("protection_na_rule", ""))
    frame["起点機転の一致度許容水準"] = str(record.get("agreement_threshold", ""))
    frame["κ算出不能・希少カテゴリの扱い"] = str(record.get("sparse_category_rule", ""))
    return frame


def write_public_codebook() -> None:
    pd.DataFrame(codebook_rows()).to_csv(CODEBOOK_OUTPUT, index=False, encoding="utf-8-sig")


def safe_defined_name(field: str, index: int) -> str:
    return f"choices_{index:02d}"


def add_choice_sheet(workbook: Workbook) -> dict[str, str]:
    sheet = workbook.create_sheet("選択肢")
    names: dict[str, str] = {}
    for index, (field, values) in enumerate(CHOICES.items(), start=1):
        column = get_column_letter(index)
        sheet.cell(1, index, field)
        for row, value in enumerate(values, start=2):
            sheet.cell(row, index, value)
        name = safe_defined_name(field, index)
        reference = f"'選択肢'!${column}$2:${column}${len(values) + 1}"
        workbook.defined_names.add(DefinedName(name, attr_text=reference))
        names[field] = name
        sheet.column_dimensions[column].width = max(18, min(36, max(map(len, [field, *values])) * 1.8))
    sheet.freeze_panes = "A2"
    return names


def add_codebook_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("判定基準")
    sheet["A1"] = f"固定コードブック候補 {CODEBOOK_VERSION}（研究者承認前は入力禁止）"
    sheet["A1"].font = Font(bold=True, size=14, color="9C0006")
    sheet["A2"] = "このシートは閲覧用です。正式な定義は同梱MarkdownとStep11-18 CSVを確認してください。"
    for row, line in enumerate(CODEBOOK_SOURCE.read_text(encoding="utf-8").splitlines(), start=4):
        cell = sheet.cell(row, 1, line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 120
    sheet.freeze_panes = "A4"


def add_metadata_sheet(workbook: Workbook, reviewer: str, csv_path: Path) -> None:
    sheet = workbook.create_sheet("メタデータ")
    rows = [
        ("コードブック版", CODEBOOK_VERSION),
        ("承認状態", APPROVAL_PENDING),
        ("評価段階", "開発用"),
        ("評価者", reviewer),
        ("作成日時（UTC）", datetime.now(timezone.utc).isoformat()),
        ("元CSV", csv_path.name),
        ("元CSV SHA-256", sha256(csv_path)),
        ("コードブックMarkdown SHA-256", sha256(CODEBOOK_SOURCE)),
        ("ワークフローコード SHA-256", sha256(Path(__file__).resolve())),
        ("対象件数", str(EXPECTED_ROWS)),
        ("回答正本", "このExcel。CSVを並行編集しない。"),
        ("禁止閲覧", "相手回答／自動回答／歯牙障害区分／内部ID"),
        ("承認者", ""),
        ("承認日", ""),
    ]
    for row, (key, value) in enumerate(rows, start=1):
        sheet.cell(row, 1, key).font = Font(bold=True)
        sheet.cell(row, 2, value)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 90


def build_workbook(frame: pd.DataFrame, reviewer: str, destination: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = INPUT_SHEET
    choice_names = add_choice_sheet(workbook)
    add_codebook_sheet(workbook)
    add_metadata_sheet(workbook, reviewer, source_path(reviewer))

    last_column = get_column_letter(len(INPUT_COLUMNS))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(INPUT_COLUMNS))
    sheet["A1"] = f"Step11 開発用100件・評価者{reviewer}（{CODEBOOK_VERSION}）"
    sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(INPUT_COLUMNS))
    sheet["A2"] = "研究者確認待ち：承認記録が作成されるまで回答を入力しないでください。"
    sheet["A2"].font = Font(bold=True, color="9C0006")
    sheet["A2"].fill = PatternFill("solid", fgColor="FFC7CE")
    sheet["A2"].alignment = Alignment(horizontal="center")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column, header in enumerate(INPUT_COLUMNS, start=1):
        cell = sheet.cell(HEADER_ROW, column, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if header == ID_COLUMN:
            cell.comment = Comment("不透明な管理IDです。変更しないでください。", "SportDent")
        elif header == TEXT_COLUMN:
            cell.comment = Comment("この原文だけを読み、定型項目や自動回答を推測しないでください。", "SportDent")

    for offset, (_, source_row) in enumerate(frame.iterrows(), start=FIRST_DATA_ROW):
        id_cell = sheet.cell(offset, 1, str(source_row[ID_COLUMN]))
        text_cell = sheet.cell(offset, 2, str(source_row[TEXT_COLUMN]))
        id_cell.data_type = "s"
        text_cell.data_type = "s"
        text_cell.alignment = Alignment(wrap_text=True, vertical="top")
        id_cell.protection = Protection(locked=True)
        text_cell.protection = Protection(locked=True)
        for column in range(3, len(INPUT_COLUMNS) + 1):
            cell = sheet.cell(offset, column, "")
            # 研究者承認までは物理的に編集不可とし、approve時にのみ解除する。
            cell.protection = Protection(locked=True)
            cell.fill = PatternFill("solid", fgColor="E7E6E6")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column, field in enumerate(INPUT_COLUMNS, start=1):
        if field not in CHOICES:
            continue
        validation = DataValidation(type="list", formula1=f"={choice_names[field]}", allow_blank=True)
        validation.error = "固定選択肢から選んでください。空欄は未入力として扱います。"
        validation.errorTitle = "選択肢外の入力"
        validation.prompt = "コードブックの固定選択肢から選択してください。"
        validation.promptTitle = field
        validation.showErrorMessage = True
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
        validation.add(f"{get_column_letter(column)}{FIRST_DATA_ROW}:{get_column_letter(column)}{FIRST_DATA_ROW + EXPECTED_ROWS - 1}")

    sheet.freeze_panes = "C4"
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{last_column}{FIRST_DATA_ROW + EXPECTED_ROWS - 1}"
    widths = {1: 18, 2: 80}
    for index in range(3, len(INPUT_COLUMNS) + 1):
        widths[index] = 22 if INPUT_COLUMNS[index - 1] != COMMENT_COLUMN else 44
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + EXPECTED_ROWS):
        sheet.row_dimensions[row].height = 54

    sheet.protection.sheet = True
    sheet.protection.password = "Step11Review"
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = False
    for other in workbook.worksheets[1:]:
        other.protection.sheet = True
        other.protection.password = "Step11Review"
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"
    workbook.save(destination)


def metadata(workbook) -> dict[str, str]:
    sheet = workbook["メタデータ"]
    result: dict[str, str] = {}
    for row in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
        if row[0] is not None:
            result[str(row[0])] = "" if row[1] is None else str(row[1])
    return result


def approval_summary_text(record: dict[str, object]) -> str:
    return (
        f"承認記録：研究責任者={str(record.get('approved_by', ''))}／"
        f"評価者A={str(record.get('reviewer_a_confirmed', ''))}／"
        f"評価者B={str(record.get('reviewer_b_confirmed', ''))}／"
        f"調停担当者={str(record.get('adjudicator', ''))}。"
        f"保護具『該当なし』条件：{str(record.get('protection_na_rule', ''))}。"
        f"一致度許容水準：{str(record.get('agreement_threshold', ''))}。"
        f"希少カテゴリ規則：{str(record.get('sparse_category_rule', ''))}"
    )


def approval_record_errors(meta: dict[str, str], approval_summary: str) -> list[str]:
    """承認JSONを単なる存在確認でなく内容・版・Excelメタデータまで照合する。"""
    errors: list[str] = []
    if not APPROVAL.exists():
        return ["コードブック承認記録がありません"]
    try:
        record = json.loads(APPROVAL.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return ["コードブック承認記録を正しく読めません"]
    if record.get("codebook_version") != CODEBOOK_VERSION or record.get("status") != APPROVAL_COMPLETE:
        errors.append("承認記録の版または状態が不一致です")
    expected_hashes = {
        "codebook_markdown_sha256": sha256(CODEBOOK_SOURCE),
        "codebook_candidate_csv_sha256": sha256(CODEBOOK_OUTPUT),
        "workflow_code_sha256": sha256(Path(__file__).resolve()),
    }
    for key, expected_hash in expected_hashes.items():
        if record.get(key) != expected_hash:
            errors.append(f"承認記録の{key}が現行承認対象と一致しません")
    sheet_hash = str(record.get("approval_sheet_sha256", ""))
    if sheet_hash:
        if not APPROVAL_SHEET.exists() or sha256(APPROVAL_SHEET) != sheet_hash:
            errors.append("承認チェックシートが承認時のハッシュと一致しません")
    if not str(record.get("approved_by", "")).strip():
        errors.append("承認記録の承認者が空欄です")
    for key, label in [
        ("reviewer_a_confirmed", "評価者A確認"),
        ("reviewer_b_confirmed", "評価者B確認"),
        ("adjudicator", "調停担当者"),
        ("protection_na_rule", "保護具『該当なし』適用条件"),
        ("agreement_threshold", "起点機転の一致度許容水準"),
        ("sparse_category_rule", "κ算出不能・希少カテゴリの扱い"),
    ]:
        if not str(record.get(key, "")).strip():
            errors.append(f"承認記録の{label}が空欄です")
    try:
        datetime.strptime(str(record.get("approval_date", "")), "%Y-%m-%d")
    except ValueError:
        errors.append("承認記録の日付が不正です")
    if meta.get("承認状態") != APPROVAL_COMPLETE:
        errors.append("Excelメタデータが承認済みではありません")
    if meta.get("コードブック版") != record.get("codebook_version"):
        errors.append("Excelと承認記録のコードブック版が一致しません")
    if meta.get("承認者") != str(record.get("approved_by", "")):
        errors.append("Excelと承認記録の承認者が一致しません")
    if meta.get("承認日") != str(record.get("approval_date", "")):
        errors.append("Excelと承認記録の承認日が一致しません")
    if approval_summary != approval_summary_text(record):
        errors.append("Excel判定基準シートの承認条件表示が承認記録と一致しません")
    try:
        sentinel = json.loads(SENTINEL.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        errors.append("初期化・承認状態記録を正しく読めません")
    else:
        if sentinel.get("approval_status") != APPROVAL_COMPLETE:
            errors.append("初期化・承認状態記録が承認済みではありません")
        if sentinel.get("approval_record_sha256") != sha256(APPROVAL):
            errors.append("初期化記録と承認記録のハッシュが一致しません")
    snapshot = approved_snapshot_path()
    if not snapshot.exists():
        errors.append("承認済みコードブックsnapshotがありません")
    else:
        try:
            snapshot_frame = pd.read_csv(snapshot, encoding="utf-8-sig", dtype=str, keep_default_na=False)
        except (OSError, pd.errors.ParserError):
            errors.append("承認済みコードブックsnapshotを正しく読めません")
        else:
            expected_snapshot = approved_codebook_frame(record).astype(str)
            if list(snapshot_frame.columns) != list(expected_snapshot.columns) or not snapshot_frame.equals(expected_snapshot):
                errors.append("承認済みコードブックsnapshotの列・行・値が固定原本と一致しません")
    return errors


def approved_codebook_text(approval_date: str) -> str:
    """候補Markdownから、承認後Excelへ固定表示する本文を決定的に作る。"""
    return CODEBOOK_SOURCE.read_text(encoding="utf-8").rstrip("\n").replace(
        "固定コードブック候補（研究者確認待ち）",
        f"固定コードブック（{approval_date}承認済み）",
    ).replace(
        "承認するまでは、開発用・最終評価用のいずれにも入力を開始しない。",
        f"{approval_date}に開発用コードブックとして承認した。",
    )


def workbook_has_forbidden_content(workbook) -> list[str]:
    errors: list[str] = []
    if set(workbook.sheetnames) != EXPECTED_SHEETS or len(workbook.sheetnames) != len(EXPECTED_SHEETS):
        errors.append(f"シート構成が不一致: {workbook.sheetnames}")
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            errors.append(f"非表示シートがあります: {sheet.title}")
        for dimension_key, dimension in sheet.column_dimensions.items():
            if dimension.hidden:
                errors.append(f"非表示列があります: {sheet.title}!{dimension_key}")
        for row_number, dimension in sheet.row_dimensions.items():
            if dimension.hidden:
                errors.append(f"非表示行があります: {sheet.title}!{row_number}")
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    errors.append(f"数式セルがあります: {sheet.title}!{cell.coordinate}")
                if cell.hyperlink is not None:
                    errors.append(f"ハイパーリンクがあります: {sheet.title}!{cell.coordinate}")
    if getattr(workbook, "_external_links", []):
        errors.append("外部リンクがあります")
    if getattr(workbook, "vba_archive", None) is not None:
        errors.append("VBAアーカイブがあります")
    return errors


def row_problems(values: dict[str, str]) -> list[str]:
    problems: list[str] = []
    for field, choices in CHOICES.items():
        value = values.get(field, "").strip()
        if value and value not in choices:
            problems.append(f"{field}: 選択肢外『{value}』")
    comment = values.get(COMMENT_COLUMN, "").strip()
    if values.get("起点機転") == "その他" and not comment:
        problems.append("起点機転がその他のためコメント必須")
    if values.get("最終接触対象") == "その他" and not comment:
        problems.append("最終接触対象がその他のためコメント必須")
    if values.get("その他部位の記載") == "記載あり" and not comment:
        problems.append("その他部位の具体名をコメントへ記載")
    specific_dental = values.get("前歯部の記載") == "記載あり" or values.get("その他歯牙の記載") == "記載あり"
    if values.get("歯牙・部位不明の記載") == "記載あり" and specific_dental and not comment:
        problems.append("歯牙・部位不明と具体的歯牙部位を併記する場合は説明コメント必須")
    if values.get("予防可能性の根拠区分") == "その他" and not comment:
        problems.append("予防可能性の根拠区分がその他のためコメント必須")
    if values.get("判定不能理由") in {"その他", "対象外疑い"} and not comment:
        problems.append("判定不能理由の説明コメント必須")
    if values.get("出来事の順序") == "複数機転・順序不明" and values.get("起点機転") != "判定不能" and not comment:
        problems.append("順序不明なのに起点を確定する場合は説明コメント必須")
    if values.get("口腔・顔面への直接外力") == "あり（明記）" and values.get("最終接触対象") == "該当なし":
        problems.append("直接外力ありと最終接触対象の該当なしが矛盾")
    if values.get("口腔・顔面への直接外力") == "なし（明記）" and values.get("最終接触対象") not in {"", "該当なし", "判定不能"}:
        problems.append("直接外力なしと具体的な最終接触対象が矛盾")
    if values.get("口腔・顔面への直接外力") == "記載なし" and values.get("最終接触対象") not in {"", "該当なし", "判定不能"}:
        problems.append("直接外力の記載なしと具体的な最終接触対象が矛盾")

    preventability = values.get("予防可能性", "")
    basis = values.get("予防可能性の根拠区分", "")
    positive = preventability in {"修正可能要因の明記あり", "修正可能要因の可能性あり"}
    if positive and basis == "該当なし":
        problems.append("修正可能要因あり／可能性ありでは根拠区分が必要")
    if preventability and not positive and basis and basis != "該当なし":
        problems.append("修正可能要因を選ばない場合の根拠区分は該当なし")

    uncertainty_fields = [field for field in CHOICES if field not in {"判定不能理由", CONFIRM_COLUMN}]
    has_uncertain = any(values.get(field) == "判定不能" for field in uncertainty_fields)
    reason = values.get("判定不能理由", "")
    if has_uncertain and reason == "該当なし":
        problems.append("判定不能項目があるため判定不能理由が必要")
    if not has_uncertain and reason and reason != "該当なし":
        problems.append("判定不能項目がないため判定不能理由は該当なし")
    return problems


def inspect_workbook(
    path: Path,
    reviewer: str,
    require_complete: bool = False,
    verify_approval_record: bool = True,
) -> tuple[pd.DataFrame, list[str]]:
    if not path.exists():
        raise WorkflowError(f"Excel入力票がありません: {path}")
    if path.is_symlink() or path.parent.is_symlink():
        raise WorkflowError(f"非公開レビュー原本にシンボリックリンクは使用できません: {path}")
    if path.stat().st_mode & 0o777 != 0o600 or path.parent.stat().st_mode & 0o777 != 0o700:
        raise WorkflowError(f"非公開レビュー原本または親ディレクトリの権限が不正です: {path}")
    workbook = load_workbook(path, data_only=False, keep_links=True)
    errors = workbook_has_forbidden_content(workbook)
    if INPUT_SHEET not in workbook.sheetnames:
        workbook.close()
        raise WorkflowError("入力シートがありません")
    sheet = workbook[INPUT_SHEET]
    meta = metadata(workbook) if "メタデータ" in workbook.sheetnames else {}
    approval_status = meta.get("承認状態", "")
    if approval_status not in {APPROVAL_PENDING, APPROVAL_COMPLETE}:
        errors.append("承認状態のメタデータが不正です")
    expected_answer_locked = approval_status != APPROVAL_COMPLETE
    headers = [sheet.cell(HEADER_ROW, column).value for column in range(1, len(INPUT_COLUMNS) + 1)]
    if headers != INPUT_COLUMNS:
        errors.append("入力列または列順が変更されています")
    if not sheet.protection.sheet:
        errors.append("入力シート保護が解除されています")
    for required_sheet in ["選択肢", "判定基準", "メタデータ"]:
        if required_sheet in workbook.sheetnames and not workbook[required_sheet].protection.sheet:
            errors.append(f"補助シート保護が解除されています: {required_sheet}")
    if sheet.max_column != len(INPUT_COLUMNS):
        errors.append(f"入力シートに余剰または不足列があります: {sheet.max_column}列")
    for dimension_key, dimension in sheet.column_dimensions.items():
        if dimension.hidden:
            errors.append(f"入力シートに非表示列があります: {dimension_key}列")
    for row in range(1, sheet.max_row + 1):
        if sheet.row_dimensions[row].hidden:
            errors.append(f"入力シートに非表示行があります: {row}行")
    expected_names = {safe_defined_name(field, index) for index, field in enumerate(CHOICES, start=1)}
    if set(workbook.defined_names) != expected_names:
        errors.append("選択肢の名前付き範囲が不一致です")
    if "選択肢" in workbook.sheetnames:
        choices_sheet = workbook["選択肢"]
        for index, (field, expected_values) in enumerate(CHOICES.items(), start=1):
            actual_values = [choices_sheet.cell(row, index).value for row in range(2, len(expected_values) + 2)]
            if choices_sheet.cell(1, index).value != field or actual_values != expected_values:
                errors.append(f"選択肢シートの内容が不一致です: {field}")
            name = safe_defined_name(field, index)
            expected_reference = f"'選択肢'!${get_column_letter(index)}$2:${get_column_letter(index)}${len(expected_values) + 1}"
            defined = workbook.defined_names.get(name)
            if defined is None or defined.attr_text != expected_reference:
                errors.append(f"選択肢の名前付き範囲参照が不一致です: {field}")
    validations = [validation for validation in sheet.data_validations.dataValidation if validation.type == "list"]
    validation_map = {str(validation.formula1 or "").lstrip("="): str(validation.sqref) for validation in validations}
    expected_validation_map = {
        safe_defined_name(field, index): (
            f"{get_column_letter(INPUT_COLUMNS.index(field) + 1)}{FIRST_DATA_ROW}:"
            f"{get_column_letter(INPUT_COLUMNS.index(field) + 1)}{FIRST_DATA_ROW + EXPECTED_ROWS - 1}"
        )
        for index, field in enumerate(CHOICES, start=1)
    }
    if validation_map != expected_validation_map or len(validations) != len(CHOICES):
        errors.append("ドロップダウン検証規則または適用範囲が不一致です")
    if "判定基準" in workbook.sheetnames:
        codebook_sheet = workbook["判定基準"]
        source_lines = CODEBOOK_SOURCE.read_text(encoding="utf-8").splitlines()
        embedded_lines = [
            "" if codebook_sheet.cell(row, 1).value is None else str(codebook_sheet.cell(row, 1).value)
            for row in range(4, 4 + len(source_lines))
        ]
        embedded_text = "\n".join(embedded_lines)
        source_text = CODEBOOK_SOURCE.read_text(encoding="utf-8").rstrip("\n")
        meta_for_codebook = metadata(workbook) if "メタデータ" in workbook.sheetnames else {}
        if meta_for_codebook.get("承認状態") == APPROVAL_PENDING and embedded_text != source_text:
            errors.append("Excel内の判定基準本文が候補原本と一致しません")
        if meta_for_codebook.get("承認状態") == APPROVAL_COMPLETE:
            approval_date = meta_for_codebook.get("承認日", "")
            if embedded_text != approved_codebook_text(approval_date):
                errors.append("Excel内の承認済み判定基準本文が固定原本と一致しません")
        approval_summary = "" if codebook_sheet["A3"].value is None else str(codebook_sheet["A3"].value)
    else:
        approval_summary = ""
    rows: list[dict[str, str]] = []
    for row_number in range(FIRST_DATA_ROW, FIRST_DATA_ROW + EXPECTED_ROWS):
        row: dict[str, str] = {}
        for column, header in enumerate(INPUT_COLUMNS, start=1):
            value = sheet.cell(row_number, column).value
            row[header] = "" if value is None else str(value).strip()
        rows.append(row)
        if not sheet.cell(row_number, 1).protection.locked or not sheet.cell(row_number, 2).protection.locked:
            errors.append(f"{row_number}行: IDまたは原文セルのロックが解除されています")
        actual_locks = [sheet.cell(row_number, column).protection.locked for column in range(3, len(INPUT_COLUMNS) + 1)]
        if any(lock != expected_answer_locked for lock in actual_locks):
            expected = "編集不可" if expected_answer_locked else "編集可能"
            errors.append(f"{row_number}行: 回答セルが承認状態どおりの{expected}ではありません")
    extra_values = [
        sheet.cell(row, column).value
        for row in range(FIRST_DATA_ROW + EXPECTED_ROWS, sheet.max_row + 1)
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(row, column).value not in {None, ""}
    ]
    if extra_values:
        errors.append("101件目以降に余剰データがあります")
    workbook.close()

    frame = pd.DataFrame(rows)
    source = read_source(reviewer)
    if not frame[[ID_COLUMN, TEXT_COLUMN]].equals(source[[ID_COLUMN, TEXT_COLUMN]]):
        errors.append("元CSVとレビューID／原文／順序が一致しません")
    if frame[ID_COLUMN].duplicated().any():
        errors.append("レビューIDが重複しています")
    if meta.get("コードブック版") != CODEBOOK_VERSION:
        errors.append("コードブック版が不一致です")
    if meta.get("評価者") != reviewer or meta.get("評価段階") != "開発用":
        errors.append("評価者または評価段階のメタデータが不一致です")
    if meta.get("元CSV SHA-256") != sha256(source_path(reviewer)):
        errors.append("元CSVのハッシュがメタデータと一致しません")
    if meta.get("コードブックMarkdown SHA-256") != sha256(CODEBOOK_SOURCE):
        errors.append("コードブックMarkdownのハッシュがメタデータと一致しません")
    if meta.get("ワークフローコード SHA-256") != sha256(Path(__file__).resolve()):
        errors.append("ワークフローコードのハッシュがメタデータと一致しません")

    row_states: list[str] = []
    row_messages: list[str] = []
    for _, row in frame.iterrows():
        values = {column: str(row[column]).strip() for column in INPUT_COLUMNS}
        filled_required = sum(bool(values[column]) for column in REQUIRED_COLUMNS)
        problems = row_problems(values)
        if filled_required == 0 and not values[COMMENT_COLUMN]:
            state = "未入力"
        elif filled_required < len(REQUIRED_COLUMNS):
            state = "入力途中"
        elif problems:
            state = "要確認"
        else:
            state = "入力完了"
        row_states.append(state)
        row_messages.append(" / ".join(problems))
    frame["入力状態"] = row_states
    frame["検査メッセージ"] = row_messages
    if any(message for message in row_messages):
        errors.append("選択肢外入力または回答矛盾があります")
    if approval_status == APPROVAL_COMPLETE and verify_approval_record:
        approval_errors = approval_record_errors(meta, approval_summary)
        errors.extend(f"承認確認: {error}" for error in approval_errors)
    if require_complete:
        if approval_status != APPROVAL_COMPLETE:
            errors.append("完了検査には承認済みExcelが必要です")
        incomplete = int(frame["入力状態"].ne("入力完了").sum())
        if incomplete:
            errors.append(f"未完了または要確認が{incomplete}件あります")
    return frame, errors


def write_validation_report(results: dict[str, pd.DataFrame]) -> Path:
    report_dir = INTERNAL / "ValidationReports"
    report_dir.mkdir(parents=True, exist_ok=True)
    report_dir.chmod(0o700)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    path = report_dir / f"Step11-22_開発用入力検査_{timestamp}.csv"
    rows: list[dict[str, str]] = []
    for reviewer, frame in results.items():
        for _, row in frame.iterrows():
            rows.append({
                "評価者": reviewer,
                ID_COLUMN: row[ID_COLUMN],
                "入力状態": row["入力状態"],
                "検査メッセージ": row["検査メッセージ"],
                "部位集約状態": (
                    "全体判定不能" if all(row[column] == "判定不能" for column in SITE_COLUMNS)
                    else "一部判定不能" if any(row[column] == "判定不能" for column in SITE_COLUMNS)
                    else "1部位以上特定" if any(row[column] == "記載あり" for column in SITE_COLUMNS)
                    else "記載なし"
                ),
            })
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")
    path.chmod(0o600)
    return path


def prepare() -> None:
    if not CODEBOOK_SOURCE.exists():
        raise WorkflowError(f"コードブック原本がありません: {CODEBOOK_SOURCE}")
    INTERNAL.mkdir(parents=True, exist_ok=True)
    INTERNAL.chmod(0o700)
    existing = list(INTERNAL.rglob("*.xlsx"))
    if SENTINEL.exists() or existing:
        raise WorkflowError("既存Excelまたは初期化記録があるため上書きしません。")
    frames = {reviewer: read_source(reviewer) for reviewer in ["A", "B"]}
    if not frames["A"].equals(frames["B"]):
        raise WorkflowError("評価者A/Bの開発用100件が一致しません")
    write_public_codebook()
    with tempfile.TemporaryDirectory(prefix="step11-review-", dir=INTERNAL) as temporary:
        temporary_path = Path(temporary)
        generated: dict[str, Path] = {}
        for reviewer in ["A", "B"]:
            target = temporary_path / workbook_path(reviewer).name
            build_workbook(frames[reviewer], reviewer, target)
            target.chmod(0o600)
            generated[reviewer] = target
        for reviewer, target in generated.items():
            os.replace(target, workbook_path(reviewer))
            workbook_path(reviewer).chmod(0o600)
    sentinel = {
        "codebook_version": CODEBOOK_VERSION,
        "approval_status": APPROVAL_PENDING,
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "workbooks": {reviewer: {"file": workbook_path(reviewer).name, "sha256": sha256(workbook_path(reviewer))} for reviewer in ["A", "B"]},
    }
    private_write_text(SENTINEL, json.dumps(sentinel, ensure_ascii=False, indent=2) + "\n")
    secure_tree(INTERNAL)
    print(f"PREPARE_OK: 開発用A/B各{EXPECTED_ROWS}件 / {CODEBOOK_VERSION} / 研究者確認待ち")


def approve(
    approved_by: str,
    reviewer_a_confirmed: str,
    reviewer_b_confirmed: str,
    adjudicator: str,
    protection_na_rule: str,
    agreement_threshold: str,
    sparse_category_rule: str,
    approval_date: str,
    approval_sheet_sha256: str = "",
) -> None:
    confirmations = {
        "研究責任者": approved_by,
        "評価者A": reviewer_a_confirmed,
        "評価者B": reviewer_b_confirmed,
        "調停担当者": adjudicator,
        "保護具『該当なし』適用条件": protection_na_rule,
        "起点機転の一致度許容水準": agreement_threshold,
        "κ算出不能・希少カテゴリの扱い": sparse_category_rule,
    }
    if any(not str(value).strip() for value in confirmations.values()):
        raise WorkflowError(
            "研究責任者、評価者A/B、調停担当者、保護具『該当なし』適用条件、"
            "一致度許容水準、κ算出不能・希少カテゴリ規則をすべて指定してください"
        )
    if APPROVAL.exists():
        raise WorkflowError("既に承認記録があります。上書きしません。")
    if approved_snapshot_path().exists():
        raise WorkflowError("承認済みコードブックsnapshotが既にあります。上書きしません。")
    try:
        sentinel_before = json.loads(SENTINEL.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise WorkflowError("承認前の初期化記録を正しく読めません") from error
    if sentinel_before.get("approval_status") != APPROVAL_PENDING:
        raise WorkflowError("初期化記録が研究者確認待ちではありません")
    expected_codebook = pd.DataFrame(codebook_rows()).astype(str)
    if not CODEBOOK_OUTPUT.exists():
        raise WorkflowError("公開コードブック候補CSVがありません")
    actual_codebook = pd.read_csv(CODEBOOK_OUTPUT, encoding="utf-8-sig", dtype=str, keep_default_na=False)
    if list(actual_codebook.columns) != list(expected_codebook.columns) or not actual_codebook.equals(expected_codebook):
        raise WorkflowError("公開コードブック候補CSVが承認対象コードと一致しません")
    public_codebook = OUTPUT / "Step11-02_歯科医師レビュー判定基準.md"
    if not public_codebook.exists() or sha256(public_codebook) != sha256(CODEBOOK_SOURCE):
        raise WorkflowError("公開判定基準Markdownが承認対象原本と一致しません")
    try:
        datetime.strptime(approval_date, "%Y-%m-%d")
    except ValueError as error:
        raise WorkflowError("承認日はYYYY-MM-DDで指定してください") from error
    pre_hashes: dict[str, str] = {}
    workbooks = {}
    for reviewer in ["A", "B"]:
        path = workbook_path(reviewer)
        frame, errors = inspect_workbook(path, reviewer, require_complete=False)
        if errors:
            raise WorkflowError(f"評価者{reviewer}Excelに問題があります: {' / '.join(errors)}")
        if frame[REQUIRED_COLUMNS + [COMMENT_COLUMN]].apply(lambda column: column.str.strip().ne("").any()).any():
            raise WorkflowError(f"評価者{reviewer}に回答開始後のセルがあるため承認処理を停止しました")
        pre_hashes[reviewer] = sha256(path)
        workbooks[reviewer] = load_workbook(path, data_only=False, keep_links=True)
    if not read_source("A").equals(read_source("B")):
        for workbook in workbooks.values():
            workbook.close()
        raise WorkflowError("評価者A/Bの開発用標本・原文・順序が一致しません")

    archive = INTERNAL / "SetupArchive"
    archive.mkdir(parents=True, exist_ok=True)
    archive.chmod(0o700)
    archive_timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    for reviewer in ["A", "B"]:
        original = workbook_path(reviewer)
        archived = archive / original.name.replace(".xlsx", f"_承認前_{archive_timestamp}.xlsx")
        shutil.copy2(original, archived)
        archived.chmod(0o600)

    # 2冊とも一時ファイルへ完全生成し、検査後にだけ現行ファイルと交換する。
    staged: dict[str, Path] = {}
    for reviewer, workbook in workbooks.items():
        meta_sheet = workbook["メタデータ"]
        for row in range(1, meta_sheet.max_row + 1):
            key = meta_sheet.cell(row, 1).value
            if key == "承認状態":
                meta_sheet.cell(row, 2, APPROVAL_COMPLETE)
            elif key == "承認者":
                meta_sheet.cell(row, 2, approved_by.strip())
            elif key == "承認日":
                meta_sheet.cell(row, 2, approval_date)
            elif key == "コードブックMarkdown SHA-256":
                meta_sheet.cell(row, 2, sha256(CODEBOOK_SOURCE))
            elif key == "ワークフローコード SHA-256":
                meta_sheet.cell(row, 2, sha256(Path(__file__).resolve()))
        input_sheet = workbook[INPUT_SHEET]
        input_sheet["A2"] = f"コードブック承認済み（{approval_date}）。固定選択肢で独立評価してください。"
        input_sheet["A2"].fill = PatternFill("solid", fgColor="C6EFCE")
        input_sheet["A2"].font = Font(bold=True, color="006100")
        codebook_sheet = workbook["判定基準"]
        codebook_sheet["A1"] = f"固定コードブック {CODEBOOK_VERSION}（承認済み：{approval_date}）"
        codebook_sheet["A1"].font = Font(bold=True, size=14, color="006100")
        for offset, value in enumerate(approved_codebook_text(approval_date).splitlines(), start=4):
            codebook_sheet.cell(offset, 1, value)
        codebook_sheet["A3"] = approval_summary_text({
            "approved_by": approved_by.strip(),
            "reviewer_a_confirmed": reviewer_a_confirmed.strip(),
            "reviewer_b_confirmed": reviewer_b_confirmed.strip(),
            "adjudicator": adjudicator.strip(),
            "protection_na_rule": protection_na_rule.strip(),
            "agreement_threshold": agreement_threshold.strip(),
            "sparse_category_rule": sparse_category_rule.strip(),
        })
        for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + EXPECTED_ROWS):
            for column in range(3, len(INPUT_COLUMNS) + 1):
                input_sheet.cell(row, column).protection = Protection(locked=False)
                input_sheet.cell(row, column).fill = PatternFill("solid", fgColor="FFF2CC")
        with tempfile.NamedTemporaryFile(prefix=f"step11-approved-{reviewer}-", suffix=".xlsx", dir=INTERNAL, delete=False) as stream:
            temporary = Path(stream.name)
        try:
            workbook.save(temporary)
            temporary.chmod(0o600)
            staged[reviewer] = temporary
        finally:
            workbook.close()
    for reviewer, temporary in staged.items():
        # 承認記録はExcel 2冊の検査後に一組としてcommitするため、ここでは
        # workbook内部だけを検査し、承認JSON/snapshotの照合はcommit後に行う。
        _, staged_errors = inspect_workbook(
            temporary, reviewer, require_complete=False, verify_approval_record=False,
        )
        if staged_errors:
            for path in staged.values():
                if path.exists():
                    path.unlink()
            raise WorkflowError(f"承認済み評価者{reviewer}Excelの事前検査に失敗しました: {' / '.join(staged_errors)}")

    record = {
        "codebook_version": CODEBOOK_VERSION,
        "status": APPROVAL_COMPLETE,
        "approved_by": approved_by.strip(),
        "reviewer_a_confirmed": reviewer_a_confirmed.strip(),
        "reviewer_b_confirmed": reviewer_b_confirmed.strip(),
        "adjudicator": adjudicator.strip(),
        "protection_na_rule": protection_na_rule.strip(),
        "agreement_threshold": agreement_threshold.strip(),
        "sparse_category_rule": sparse_category_rule.strip(),
        "approval_date": approval_date,
        "recorded_at_utc": datetime.now(timezone.utc).isoformat(),
        "codebook_markdown_sha256": sha256(CODEBOOK_SOURCE),
        "codebook_candidate_csv_sha256": sha256(CODEBOOK_OUTPUT),
        "workflow_code_sha256": sha256(Path(__file__).resolve()),
        "approval_sheet_sha256": approval_sheet_sha256,
        "preapproval_sha256": pre_hashes,
        "approved_sha256": {reviewer: sha256(staged[reviewer]) for reviewer in ["A", "B"]},
    }
    approved_frame = approved_codebook_frame(record)

    originals: dict[str, Path] = {}
    transaction = INTERNAL / f".approval-transaction-{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S%fZ')}"
    transaction.mkdir()
    transaction.chmod(0o700)
    try:
        for reviewer in ["A", "B"]:
            original_backup = transaction / workbook_path(reviewer).name
            shutil.copy2(workbook_path(reviewer), original_backup)
            original_backup.chmod(0o600)
            originals[reviewer] = original_backup
        sentinel_backup = transaction / f"original-{SENTINEL.name}"
        if SENTINEL.exists():
            shutil.copy2(SENTINEL, sentinel_backup)
            sentinel_backup.chmod(0o600)
        staged_approval = transaction / APPROVAL.name
        staged_approval.write_text(json.dumps(record, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        staged_approval.chmod(0o600)
        approved_snapshot = approved_snapshot_path()
        staged_snapshot = transaction / approved_snapshot.name
        approved_frame.to_csv(staged_snapshot, index=False, encoding="utf-8-sig")
        staged_snapshot.chmod(0o600)
        sentinel = json.loads(SENTINEL.read_text(encoding="utf-8")) if SENTINEL.exists() else {}
        sentinel.update({
            "approval_status": APPROVAL_COMPLETE,
            "approved_at_utc": datetime.now(timezone.utc).isoformat(),
            "approval_record_sha256": sha256(staged_approval),
            "approved_workbooks": record["approved_sha256"],
        })
        staged_sentinel = transaction / SENTINEL.name
        staged_sentinel.write_text(json.dumps(sentinel, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        staged_sentinel.chmod(0o600)
        for reviewer in ["A", "B"]:
            os.replace(staged[reviewer], workbook_path(reviewer))
            workbook_path(reviewer).chmod(0o600)
        os.replace(staged_approval, APPROVAL)
        APPROVAL.chmod(0o600)
        os.replace(staged_snapshot, approved_snapshot)
        approved_snapshot.chmod(0o600)
        os.replace(staged_sentinel, SENTINEL)
        SENTINEL.chmod(0o600)
    except Exception:
        for reviewer, backup in originals.items():
            if backup.exists():
                shutil.copy2(backup, workbook_path(reviewer))
                workbook_path(reviewer).chmod(0o600)
        if APPROVAL.exists():
            APPROVAL.unlink()
        approved_snapshot = approved_snapshot_path()
        if approved_snapshot.exists():
            approved_snapshot.unlink()
        if 'sentinel_backup' in locals() and sentinel_backup.exists():
            shutil.copy2(sentinel_backup, SENTINEL)
            SENTINEL.chmod(0o600)
        raise
    finally:
        for path in staged.values():
            if path.exists():
                path.unlink()
        shutil.rmtree(transaction, ignore_errors=True)
    secure_tree(INTERNAL)
    print(f"APPROVE_OK: {CODEBOOK_VERSION} / {approval_date} / 開発用A/Bへの入力を開始できます")


def approve_from_sheet(confirm_exact: str) -> None:
    """完成別紙＋CLIの二重明示確認がある場合だけ、既存の承認処理へ渡す。"""
    if confirm_exact != APPROVAL_INTENT:
        raise WorkflowError(f"--confirm-exactには『{APPROVAL_INTENT}』を完全一致で指定してください")
    values = inspect_approval_sheet(require_complete=True)
    if values["codebook_read_confirmed"] != READ_CONFIRMATION or values["approval_intent"] != APPROVAL_INTENT:
        raise WorkflowError("判定基準の確認文または明示的な承認文が不足しています")
    approve(
        values["approved_by"], values["reviewer_a_confirmed"], values["reviewer_b_confirmed"],
        values["adjudicator"], values["protection_na_rule"], values["agreement_threshold"],
        values["sparse_category_rule"], values["approval_date"], sha256(APPROVAL_SHEET),
    )


def secure_pending() -> None:
    """旧候補Excelを、回答0件を確認した上で承認前ロック仕様へ安全に移行する。"""
    if APPROVAL.exists():
        raise WorkflowError("承認済みExcelには承認前ロック移行を適用できません")
    archive = INTERNAL / "SetupArchive"
    archive.mkdir(parents=True, exist_ok=True)
    archive.chmod(0o700)
    workbooks = {}
    for reviewer in ["A", "B"]:
        path = workbook_path(reviewer)
        if not path.exists():
            raise WorkflowError(f"Excel入力票がありません: {path}")
        workbook = load_workbook(path, data_only=False, keep_links=True)
        errors = workbook_has_forbidden_content(workbook)
        sheet = workbook[INPUT_SHEET]
        headers = [sheet.cell(HEADER_ROW, column).value for column in range(1, len(INPUT_COLUMNS) + 1)]
        if headers != INPUT_COLUMNS:
            errors.append("入力列または列順が変更されています")
        meta = metadata(workbook)
        if meta.get("承認状態") != APPROVAL_PENDING:
            errors.append("研究者確認待ちのExcelではありません")
        for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + EXPECTED_ROWS):
            if any(str(sheet.cell(row, column).value or "").strip() for column in range(3, len(INPUT_COLUMNS) + 1)):
                errors.append(f"{row}行: 回答開始済みのため移行不可")
        source = read_source(reviewer)
        ids = [str(sheet.cell(row, 1).value or "") for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + EXPECTED_ROWS)]
        texts = [str(sheet.cell(row, 2).value or "") for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + EXPECTED_ROWS)]
        if ids != source[ID_COLUMN].tolist() or texts != source[TEXT_COLUMN].tolist():
            errors.append("元CSVとレビューIDまたは原文が一致しません")
        if errors:
            workbook.close()
            raise WorkflowError(f"評価者{reviewer}を移行できません: {' / '.join(errors)}")
        workbooks[reviewer] = workbook

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    for reviewer in ["A", "B"]:
        original = workbook_path(reviewer)
        archived = archive / original.name.replace(".xlsx", f"_承認前ロック移行前_{timestamp}.xlsx")
        shutil.copy2(original, archived)
        archived.chmod(0o600)
    write_public_codebook()
    public_markdown = OUTPUT / "Step11-02_歯科医師レビュー判定基準.md"
    shutil.copy2(CODEBOOK_SOURCE, public_markdown)
    for reviewer, workbook in workbooks.items():
        meta_sheet = workbook["メタデータ"]
        existing_keys = [meta_sheet.cell(row, 1).value for row in range(1, meta_sheet.max_row + 1)]
        additions = {
            "コードブックMarkdown SHA-256": sha256(CODEBOOK_SOURCE),
            "ワークフローコード SHA-256": sha256(Path(__file__).resolve()),
        }
        for key, value in additions.items():
            if key in existing_keys:
                row = existing_keys.index(key) + 1
                meta_sheet.cell(row, 2, value)
            else:
                meta_sheet.append([key, value])
        codebook_sheet = workbook["判定基準"]
        workbook.remove(codebook_sheet)
        add_codebook_sheet(workbook)
        workbook["判定基準"].protection.sheet = True
        workbook["判定基準"].protection.password = "Step11Review"
        sheet = workbook[INPUT_SHEET]
        for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + EXPECTED_ROWS):
            for column in range(3, len(INPUT_COLUMNS) + 1):
                sheet.cell(row, column).protection = Protection(locked=True)
                sheet.cell(row, column).fill = PatternFill("solid", fgColor="E7E6E6")
        with tempfile.NamedTemporaryFile(prefix="step11-pending-", suffix=".xlsx", dir=INTERNAL, delete=False) as stream:
            temporary = Path(stream.name)
        try:
            workbook.save(temporary)
            temporary.chmod(0o600)
            os.replace(temporary, workbook_path(reviewer))
            workbook_path(reviewer).chmod(0o600)
        finally:
            workbook.close()
            if temporary.exists():
                temporary.unlink()
    sentinel = json.loads(SENTINEL.read_text(encoding="utf-8")) if SENTINEL.exists() else {}
    sentinel["pending_lock_migrated_at_utc"] = datetime.now(timezone.utc).isoformat()
    sentinel["workbooks"] = {
        reviewer: {"file": workbook_path(reviewer).name, "sha256": sha256(workbook_path(reviewer))}
        for reviewer in ["A", "B"]
    }
    private_write_text(SENTINEL, json.dumps(sentinel, ensure_ascii=False, indent=2) + "\n")
    secure_tree(INTERNAL)
    print("SECURE_PENDING_OK: 回答0件を確認し、承認前A/B Excelの回答欄を編集不可にしました")


def validate(require_complete: bool, make_report: bool) -> None:
    results: dict[str, pd.DataFrame] = {}
    all_errors: list[str] = []
    for reviewer in ["A", "B"]:
        frame, errors = inspect_workbook(workbook_path(reviewer), reviewer, require_complete=require_complete)
        results[reviewer] = frame
        counts = frame["入力状態"].value_counts().to_dict()
        print(f"評価者{reviewer}: " + " / ".join(f"{state}={counts.get(state, 0)}" for state in ["入力完了", "入力途中", "未入力", "要確認"]))
        all_errors.extend(f"評価者{reviewer}: {error}" for error in errors)
    if not results["A"][[ID_COLUMN, TEXT_COLUMN]].equals(results["B"][[ID_COLUMN, TEXT_COLUMN]]):
        all_errors.append("評価者A/BでレビューID・原文・順序が一致しません")
    report = write_validation_report(results) if make_report else None
    if report:
        print(f"検査レポート: {report}")
    if all_errors:
        raise WorkflowError("\n".join(all_errors))
    status = "完了検査" if require_complete else "構造・進捗検査"
    print(f"VALIDATION_OK: {status} / A/B標本一致 / 原文改変なし / 固定選択肢・矛盾検査済み")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    subparsers.add_parser("prepare", help="開発用A/B Excelと候補コードブックを新規生成する")
    subparsers.add_parser("secure-pending", help="回答0件の旧候補Excelを承認前ロック仕様へ安全に移行する")
    subparsers.add_parser("prepare-approval-sheet", help="承認はせず、研究者用の非公開チェックシートを新規作成する")
    check_sheet_parser = subparsers.add_parser("check-approval-sheet", help="承認はせず、研究者用チェックシートを検査する")
    check_sheet_parser.add_argument("--require-complete", action="store_true", help="全承認項目の入力完了を必須とする")
    approve_parser = subparsers.add_parser("approve", help="研究者承認を記録し、開発用入力を解禁する")
    approve_parser.add_argument("--approved-by", required=True, help="承認者名または管理コード")
    approve_parser.add_argument("--reviewer-a-confirmed", required=True, help="評価者Aの確認者名または管理コード")
    approve_parser.add_argument("--reviewer-b-confirmed", required=True, help="評価者Bの確認者名または管理コード")
    approve_parser.add_argument("--adjudicator", required=True, help="第三者調停担当者名または管理コード")
    approve_parser.add_argument(
        "--protection-na-rule", required=True,
        help="ヘルメット／マウスガードで『該当なし』を選べる条件（例：事故時の活動上、当該保護具が評価対象外と原文から明記できる場合のみ）",
    )
    approve_parser.add_argument(
        "--agreement-threshold", required=True,
        help="結果を見る前に定めた起点機転の一致度許容水準（指標・数値・CI判定を含む）",
    )
    approve_parser.add_argument(
        "--sparse-category-rule", required=True,
        help="κ算出不能および希少カテゴリの事前の扱い",
    )
    approve_parser.add_argument("--approval-date", default=datetime.now(timezone.utc).date().isoformat())
    approve_sheet_parser = subparsers.add_parser(
        "approve-from-sheet", help="完成チェックシートと二重の明示確認により入力を解禁する",
    )
    approve_sheet_parser.add_argument(
        "--confirm-exact", required=True,
        help=f"安全確認のため『{APPROVAL_INTENT}』を完全一致で指定する",
    )
    for name in ["validate", "check"]:
        validate_parser = subparsers.add_parser(name, help="Excel構造・進捗・回答矛盾を読み取り検査する")
        validate_parser.add_argument("--require-complete", action="store_true", help="A/B各100件の入力完了を必須とする")
        validate_parser.add_argument("--no-report", action="store_true", help="非公開の検査レポートCSVを生成しない")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    try:
        if args.command == "prepare":
            prepare()
        elif args.command == "secure-pending":
            secure_pending()
        elif args.command == "prepare-approval-sheet":
            create_approval_sheet()
        elif args.command == "check-approval-sheet":
            inspect_approval_sheet(require_complete=args.require_complete)
        elif args.command == "approve-from-sheet":
            approve_from_sheet(args.confirm_exact)
        elif args.command == "approve":
            approve(
                args.approved_by, args.reviewer_a_confirmed, args.reviewer_b_confirmed,
                args.adjudicator, args.protection_na_rule, args.agreement_threshold,
                args.sparse_category_rule, args.approval_date,
            )
        else:
            validate(require_complete=args.require_complete, make_report=not args.no_report)
    except WorkflowError as error:
        raise SystemExit(f"WORKFLOW_STOP: {error}") from error


if __name__ == "__main__":
    main()
